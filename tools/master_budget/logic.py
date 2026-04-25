from __future__ import annotations

import csv
import gc
import re
import shutil
import subprocess
import sys
import tempfile
import time
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from toolkit.base_tool import (
    ProgressFn,  # noqa: F401 — re-exported; ToolResult/LogLine owned by frame
)
from toolkit.fills import argb, bgr_int
from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY

# ---------------------------------------------------------------------------
# pywin32 guard  (ADR-0007)
# ---------------------------------------------------------------------------
# Pre-declare as Any so the COM-path code typechecks without pywin32 stubs.
pythoncom: Any = None
pywintypes: Any = None
win32com: Any = None
win32com_client: Any = None  # the .client submodule — pywin32 quirk: importing
# ``win32com`` alone does NOT attach ``client`` as
# an attribute. We keep a separate handle so
# callers can do ``win32com_client.DispatchEx(...)``
# instead of the broken ``win32com.client...``.
win32process: Any = None
HAVE_COM: bool = False

if sys.platform == "win32":  # pragma: no cover
    try:
        import importlib as _importlib

        pythoncom = _importlib.import_module("pythoncom")
        pywintypes = _importlib.import_module("pywintypes")
        win32com = _importlib.import_module("win32com")
        # IMPORTANT: import_module("win32com.client") both imports AND attaches
        # the submodule to the parent package, so the attribute access path
        # (``win32com.client.DispatchEx``) works after this line. Without this
        # explicit import, ``win32com.client`` raises AttributeError at runtime.
        win32com_client = _importlib.import_module("win32com.client")
        win32process = _importlib.import_module("win32process")
        HAVE_COM = True
    except Exception:
        HAVE_COM = False

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
SUPPORTED_SOURCE_EXTENSIONS: set[str] = {".csv", ".xlsx", ".xlsm"}
SUPPORTED_TARGET_EXTENSIONS: set[str] = {".xlsx", ".xlsm"}
SUPPORTED_CSV_ENCODINGS: tuple[str, ...] = ("utf-8-sig", "utf-8", "cp1252")
EXCEL_RETRY_HRESULTS: set[int] = {-2147418111, -2147417846}

# openpyxl PatternFill objects — colours must match v1 exactly
_HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_MISMATCH))
_EXTRA_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_SOURCE_ONLY))

_MASTER_SHEET = "Master"
_COMPASS_SHEET = "Compass"
_PROTECTED_START_CODE = "26201"


# ---------------------------------------------------------------------------
# Public dataclass
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ImportSummary:
    matched_rows: int
    matched_cells: int
    mismatch_account_codes: list[str]  # rows: master account codes missing from source
    mismatch_subprogram_codes: list[str]  # columns: master subprogram codes missing from source
    source_only_account_codes: list[str]  # rows: source account codes missing from master
    source_only_subprogram_codes: list[str]  # columns: source subprogram codes missing from master
    output_path: Path


# ---------------------------------------------------------------------------
# Internal exceptions
# ---------------------------------------------------------------------------


class _BudgetError(Exception):
    pass


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def suggest_output_name(master_file: Path) -> str:
    """Return a suggested output file name.

    Format: ``<stem>_AUTO_<YYYYMMDD_HHMM><suffix>`` in the same directory as
    *master_file*.  Reproduces v1's 'Create suggested output name' behaviour
    exactly (see ``app.py::BudgetAutomationApp._suggest_output``).
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    suggested = master_file.with_name(f"{master_file.stem}_AUTO_{timestamp}{master_file.suffix}")
    return str(suggested)


def import_expense_sub_program(
    expense_file: Path,
    master_file: Path,
    output_file: Path,
    progress: ProgressFn,
) -> ImportSummary:
    """Port of v1's main automation routine.

    Reads the Compass Expense Sub-Program XLSX, matches account codes
    against the Master Budget XLSM template, writes an annotated copy
    to ``output_file``. Highlights mismatches with ``HL_MISMATCH`` and
    source-only rows with ``HL_SOURCE_ONLY`` (openpyxl fills).

    Preserves macro bindings by routing the final save through pywin32
    COM where available; falls back to openpyxl save on non-Windows
    (which loses macros — acceptable for CI dev loops only).

    Calls ``progress(percent, message)`` at key checkpoints.
    """
    _validate_paths(expense_file, master_file, output_file)
    progress(5, "Validating files...")

    source_data = _read_source(expense_file)
    progress(15, "Reading source data...")

    # --- first pass: read master layout (validates sheet presence) ----------
    wb = openpyxl.load_workbook(master_file, keep_vba=True, data_only=False)
    try:
        if _MASTER_SHEET not in wb.sheetnames:
            raise _BudgetError(f"Sheet '{_MASTER_SHEET}' was not found in the workbook.")
        master_ws = wb[_MASTER_SHEET]
        protected_start_row = _find_protected_start_row(master_ws)
        master_map = _read_master_layout(
            master_ws,
            editable_end_row=_editable_end_row(master_ws, protected_start_row),
        )
        if not master_map["row_map"]:
            raise _BudgetError(
                "Could not detect the account-code rows on the Master sheet. "
                "Please check that account codes are in column A."
            )
        if not master_map["subprogram_map"]:
            raise _BudgetError(
                "Could not detect the sub-program columns on the Master sheet. "
                "Please check that sub-program codes are in row 4."
            )
    finally:
        wb.close()

    # --- compute mismatch sets -----------------------------------------------
    master_codes_for_mismatch = {c for c in master_map["row_codes"] if _is_mismatch_account_code(c)}
    source_codes_for_mismatch = {
        c for c in source_data["row_codes"] if _is_mismatch_account_code(c)
    }
    missing_master_codes = sorted(
        master_codes_for_mismatch - source_codes_for_mismatch, key=_sort_key
    )
    missing_source_codes = sorted(
        source_codes_for_mismatch - master_codes_for_mismatch, key=_sort_key
    )
    missing_subprogram_codes = sorted(
        master_map["subprogram_codes"] - source_data["subprogram_codes"],
        key=_sort_key,
    )
    source_extra_subprogram_codes = sorted(
        source_data["subprogram_codes"] - master_map["subprogram_codes"],
        key=_sort_key,
    )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    progress(20, "Checking mismatches and preparing workbook...")

    # --- choose write path ---------------------------------------------------
    if _can_use_excel_native():
        matched_cells, matched_rows = _run_excel_native(  # pragma: no cover
            template_path=master_file,
            output_path=output_file,
            source_data=source_data,
            missing_master_codes=missing_master_codes,
            missing_source_codes=missing_source_codes,
            missing_subprogram_codes=missing_subprogram_codes,
            source_extra_subprogram_codes=source_extra_subprogram_codes,
            progress_callback=progress,
        )
    else:
        wb = openpyxl.load_workbook(master_file, keep_vba=True, data_only=False)
        try:
            master_ws = wb[_MASTER_SHEET]
            compass_ws = wb[_COMPASS_SHEET] if _COMPASS_SHEET in wb.sheetnames else None
            protected_start_row = _find_protected_start_row(master_ws)
            progress(35, "Updating Master sheet...")
            master_map = _insert_source_only_items_openpyxl(
                master_ws,
                source_data,
                missing_source_codes,
                source_extra_subprogram_codes,
                protected_start_row=protected_start_row,
                progress_callback=progress,
            )
            progress(55, "Writing imported values...")
            matched_cells, matched_rows = _populate_master(
                master_ws,
                master_map,
                source_data,
                progress_callback=progress,
            )
            _ensure_master_total_formulas_openpyxl(master_ws, master_map)
            if compass_ws is not None:
                progress(70, "Refreshing Compass sheet...")
                _populate_compass(compass_ws, source_data)
            progress(82, "Applying mismatch highlighting...")
            _apply_mismatch_highlights(
                master_ws,
                compass_ws,
                master_map,
                source_data,
                missing_master_codes,
                missing_source_codes,
                missing_subprogram_codes,
                source_extra_subprogram_codes,
            )
            progress(92, "Saving output workbook...")
            wb.save(output_file)
        finally:
            wb.close()

    progress(100, "Completed.")

    return ImportSummary(
        matched_rows=matched_rows,
        matched_cells=matched_cells,
        mismatch_account_codes=missing_master_codes,
        mismatch_subprogram_codes=missing_subprogram_codes,
        source_only_account_codes=missing_source_codes,
        source_only_subprogram_codes=source_extra_subprogram_codes,
        output_path=output_file,
    )


# ---------------------------------------------------------------------------
# Internal helpers — validation
# ---------------------------------------------------------------------------


def _validate_paths(source_path: Path, template_path: Path, output_path: Path) -> None:
    source_resolved = source_path.resolve()
    template_resolved = template_path.resolve()
    output_resolved = output_path.resolve()

    if not source_path.exists():
        raise _BudgetError(f"Source file not found: {source_path}")
    if not template_path.exists():
        raise _BudgetError(f"Target workbook not found: {template_path}")
    if source_path.suffix.lower() not in SUPPORTED_SOURCE_EXTENSIONS:
        raise _BudgetError("Source file must be one of: .csv, .xlsx, .xlsm")
    if template_path.suffix.lower() not in SUPPORTED_TARGET_EXTENSIONS:
        raise _BudgetError("Target workbook must be one of: .xlsx, .xlsm")
    if output_path.suffix.lower() not in SUPPORTED_TARGET_EXTENSIONS:
        raise _BudgetError("Output workbook must be one of: .xlsx, .xlsm")
    if source_resolved == template_resolved:
        raise _BudgetError("Source file and target workbook must be different files.")
    if source_resolved == output_resolved:
        raise _BudgetError("Output workbook must be different from the source file.")
    if template_resolved == output_resolved:
        raise _BudgetError(
            "Output workbook must be a new file. Please do not save over the original template."
        )
    if template_path.suffix.lower() == ".xlsm" and output_path.suffix.lower() != ".xlsm":
        raise _BudgetError(
            "Output workbook must use the .xlsm extension when the template "
            "workbook is .xlsm, otherwise macros and button bindings cannot be "
            "preserved."
        )


# ---------------------------------------------------------------------------
# Internal helpers — source reading
# ---------------------------------------------------------------------------


def _read_source(source_path: Path) -> dict[str, Any]:
    rows: list[list[str]] = (
        _read_csv_rows(source_path)
        if source_path.suffix.lower() == ".csv"
        else _read_excel_rows(source_path)
    )
    if len(rows) < 3:
        raise _BudgetError("Source file does not contain enough rows.")

    normalized = _normalize_rows(rows)
    subprogram_row = normalized[0]
    name_row = normalized[1]

    subprogram_map: dict[str, int] = {}
    subprogram_names: dict[str, str] = {}
    duplicate_subprogram_codes: list[str] = []
    for idx, code in enumerate(subprogram_row):
        code_clean = _clean_string(code)
        if code_clean:
            if code_clean in subprogram_map:
                duplicate_subprogram_codes.append(code_clean)
                continue
            subprogram_map[code_clean] = idx
            subprogram_names[code_clean] = (
                _clean_string(name_row[idx]) if idx < len(name_row) else ""
            )

    data_rows: dict[str, list[str]] = {}
    row_names: dict[str, str] = {}
    duplicate_account_codes: list[str] = []
    for row in normalized[2:]:
        account_code = _clean_string(row[0])
        if not account_code:
            continue
        if account_code in data_rows:
            duplicate_account_codes.append(account_code)
            continue
        data_rows[account_code] = row
        row_names[account_code] = _clean_string(row[1]) if len(row) > 1 else ""

    if duplicate_subprogram_codes:
        duplicates = ", ".join(sorted(set(duplicate_subprogram_codes), key=_sort_key))
        raise _BudgetError(f"Source file contains duplicate sub-program codes: {duplicates}")
    if duplicate_account_codes:
        duplicates = ", ".join(sorted(set(duplicate_account_codes), key=_sort_key))
        raise _BudgetError(f"Source file contains duplicate account codes: {duplicates}")

    return {
        "rows": data_rows,
        "row_names": row_names,
        "row_codes": set(data_rows.keys()),
        "subprogram_map": subprogram_map,
        "subprogram_names": subprogram_names,
        "subprogram_codes": set(subprogram_map.keys()) - {"Total", "EI/SP"},
    }


def _read_csv_rows(source_path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in SUPPORTED_CSV_ENCODINGS:
        try:
            with source_path.open("r", encoding=encoding, newline="") as fh:
                return list(csv.reader(fh))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise _BudgetError(
        "Could not read the CSV source file using UTF-8 or Windows encodings. "
        "Please re-save the file as UTF-8 CSV and try again."
    ) from last_error


def _read_excel_rows(source_path: Path) -> list[list[str]]:
    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if v is None else str(v) for v in row])
        return rows
    finally:
        wb.close()


def _normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    non_empty = [r for r in rows if any(str(c).strip() for c in r)]
    if not non_empty:
        raise _BudgetError("Source file is empty.")
    max_len = max(len(r) for r in non_empty)
    return [r + [""] * (max_len - len(r)) for r in non_empty]


# ---------------------------------------------------------------------------
# Internal helpers — master layout
# ---------------------------------------------------------------------------


def _find_protected_start_row(ws: Any) -> int | None:
    for row in range(1, ws.max_row + 1):
        if _clean_string(ws.cell(row, 1).value) == _PROTECTED_START_CODE:
            return row
    return None


def _editable_end_row(ws: Any, protected_start_row: int | None) -> int:
    return (protected_start_row - 1) if protected_start_row else ws.max_row


def _read_master_layout(ws: Any, editable_end_row: int | None = None) -> dict[str, Any]:
    row_map: dict[str, int] = {}
    row_names: dict[str, str] = {}
    subprogram_map: dict[str, int] = {}
    subprogram_names: dict[str, str] = {}

    for col in range(4, ws.max_column + 1):
        code = _clean_string(ws.cell(4, col).value)
        if code:
            subprogram_map[code] = col
            subprogram_names[code] = _clean_string(ws.cell(5, col).value)

    max_row = editable_end_row if editable_end_row is not None else ws.max_row
    for row in range(1, max_row + 1):
        account_code = _clean_string(ws.cell(row, 1).value)
        if account_code.isdigit() and len(account_code) >= 5:
            row_map[account_code] = row
            row_names[account_code] = _clean_string(ws.cell(row, 2).value)

    return {
        "row_map": row_map,
        "row_names": row_names,
        "subprogram_map": subprogram_map,
        "subprogram_names": subprogram_names,
        "row_codes": set(row_map.keys()),
        "subprogram_codes": set(subprogram_map.keys()),
    }


# ---------------------------------------------------------------------------
# Internal helpers — master population (openpyxl path)
# ---------------------------------------------------------------------------


def _populate_master(
    ws: Any,
    master_map: dict[str, Any],
    source_data: dict[str, Any],
    progress_callback: ProgressFn | None = None,
) -> tuple[int, int]:
    matched_cells = 0
    matched_rows = 0
    row_items = sorted(master_map["row_map"].items(), key=lambda item: item[1])
    total_rows = max(1, len(row_items))
    for index, (account_code, target_row) in enumerate(row_items, start=1):
        source_row = source_data["rows"].get(account_code)
        if source_row is not None:
            matched_rows += 1
        for subprogram_code, target_col in master_map["subprogram_map"].items():
            cell = ws.cell(target_row, target_col)
            if source_row is None:
                cell.value = None
                continue
            source_idx = source_data["subprogram_map"].get(subprogram_code)
            if source_idx is None or source_idx >= len(source_row):
                cell.value = None
                continue
            parsed = _parse_source_number(source_row[source_idx])
            cell.value = parsed
            if parsed is not None:
                matched_cells += 1
        if progress_callback is not None and (index == total_rows or index % 10 == 0):
            pct = 55 + int(20 * index / total_rows)
            progress_callback(pct, f"Writing imported values... ({index}/{total_rows} rows)")
    return matched_cells, matched_rows


def _populate_compass(ws: Any, source_data: dict[str, Any]) -> None:
    max_existing_row = ws.max_row
    max_existing_col = ws.max_column
    for row in ws.iter_rows(
        min_row=1, max_row=max_existing_row, min_col=1, max_col=max_existing_col
    ):
        for cell in row:
            cell.value = None

    ordered_subprogram_codes = [
        code for code in source_data["subprogram_map"].keys() if code not in {"Total", "EI/SP"}
    ]
    ordered_account_codes = list(source_data["rows"].keys())

    for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
        ws.cell(1, col_idx).value = (
            int(subprogram_code) if subprogram_code.isdigit() else subprogram_code
        )
        ws.cell(2, col_idx).value = source_data["subprogram_names"].get(subprogram_code, "")

    ws.cell(2, 2).value = "EI/SP"
    ws.cell(2, 3).value = "Total"

    for row_idx, account_code in enumerate(ordered_account_codes, start=3):
        source_row = source_data["rows"][account_code]
        ws.cell(row_idx, 1).value = int(account_code) if account_code.isdigit() else account_code
        ws.cell(row_idx, 2).value = source_data["row_names"].get(account_code, "")
        total_idx = source_data["subprogram_map"].get("Total")
        total_value = (
            _parse_source_number(source_row[total_idx])
            if total_idx is not None and total_idx < len(source_row)
            else None
        )
        ws.cell(row_idx, 3).value = total_value
        for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
            source_idx = source_data["subprogram_map"].get(subprogram_code)
            value = (
                _parse_source_number(source_row[source_idx])
                if source_idx is not None and source_idx < len(source_row)
                else None
            )
            ws.cell(row_idx, col_idx).value = value

    last_used_row = len(ordered_account_codes) + 2
    last_used_col = len(ordered_subprogram_codes) + 3
    for row in range(1, max_existing_row + 1):
        for col in range(1, max_existing_col + 1):
            if row <= last_used_row and col <= last_used_col:
                continue
            ws.cell(row, col).value = None


def _apply_mismatch_highlights(
    master_ws: Any,
    compass_ws: Any,
    master_map: dict[str, Any],
    source_data: dict[str, Any],
    missing_master_codes: list[str],
    missing_source_codes: list[str],
    missing_subprogram_codes: list[str],
    source_extra_subprogram_codes: list[str],
) -> None:
    master_last_col = (
        max(master_map["subprogram_map"].values()) if master_map["subprogram_map"] else 3
    )
    master_last_row = max(master_map["row_map"].values()) if master_map["row_map"] else 5

    for code in missing_master_codes:
        row_idx = master_map["row_map"].get(code)
        if row_idx:
            for col in range(1, master_last_col + 1):
                _apply_fill(master_ws.cell(row_idx, col))

    for code in missing_subprogram_codes:
        col_idx = master_map["subprogram_map"].get(code)
        if col_idx:
            for row in range(4, master_last_row + 1):
                _apply_fill(master_ws.cell(row, col_idx))

    for code in missing_source_codes:
        row_idx = master_map["row_map"].get(code)
        if row_idx:
            for col in range(1, master_last_col + 1):
                _apply_extra_fill(master_ws.cell(row_idx, col))

    for code in source_extra_subprogram_codes:
        col_idx = master_map["subprogram_map"].get(code)
        if col_idx:
            for row in range(4, master_last_row + 1):
                _apply_extra_fill(master_ws.cell(row, col_idx))

    if compass_ws is None:
        return

    ordered_subprogram_codes = [
        code for code in source_data["subprogram_map"].keys() if code not in {"Total", "EI/SP"}
    ]
    ordered_account_codes = list(source_data["rows"].keys())
    extra_rows = set(missing_source_codes)
    extra_cols = set(source_extra_subprogram_codes)

    for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
        if subprogram_code in extra_cols:
            for row_idx in range(1, len(ordered_account_codes) + 3):
                _apply_extra_fill(compass_ws.cell(row_idx, col_idx))

    last_col = len(ordered_subprogram_codes) + 3
    for row_idx, account_code in enumerate(ordered_account_codes, start=3):
        if account_code in extra_rows:
            for col_idx in range(1, last_col + 1):
                _apply_extra_fill(compass_ws.cell(row_idx, col_idx))


def _apply_fill(cell: Any) -> None:
    cell.fill = copy(_HIGHLIGHT_FILL)


def _apply_extra_fill(cell: Any) -> None:
    cell.fill = copy(_EXTRA_FILL)


# ---------------------------------------------------------------------------
# Internal helpers — insert source-only items (openpyxl path)
# ---------------------------------------------------------------------------


def _insert_source_only_items_openpyxl(
    ws: Any,
    source_data: dict[str, Any],
    missing_source_codes: list[str],
    source_extra_subprogram_codes: list[str],
    protected_start_row: int | None = None,
    progress_callback: ProgressFn | None = None,
) -> dict[str, Any]:
    editable_end = _editable_end_row(ws, protected_start_row)
    master_map = _read_master_layout(ws, editable_end_row=editable_end)
    total_steps = max(1, len(source_extra_subprogram_codes) + len(missing_source_codes))
    step_index = 0

    for code in sorted(source_extra_subprogram_codes, key=_sort_key):
        insert_col = _find_insert_col(master_map["subprogram_map"], code)
        source_col = insert_col - 1 if insert_col > 4 else insert_col + 1
        _insert_partial_column_openpyxl(ws, insert_col, 1, editable_end)
        _copy_column_format_openpyxl(ws, source_col, insert_col)
        _copy_column_formulas_openpyxl(ws, source_col, insert_col, 1, editable_end)
        ws.cell(4, insert_col).value = int(code) if code.isdigit() else code
        ws.cell(5, insert_col).value = source_data["subprogram_names"].get(code, "")
        master_map = _read_master_layout(
            ws,
            editable_end_row=_editable_end_row(ws, protected_start_row),
        )
        step_index += 1
        if progress_callback is not None:
            pct = 35 + int(18 * step_index / total_steps)
            progress_callback(pct, f"Updating Master sheet... ({step_index}/{total_steps})")

    for code in sorted(missing_source_codes, key=_sort_key):
        insert_row = _find_insert_row(master_map["row_map"], code)
        if protected_start_row is not None and insert_row >= protected_start_row:
            continue
        source_row_num = insert_row - 1 if insert_row > 6 else insert_row + 1
        ws.insert_rows(insert_row, 1)
        _copy_row_format_openpyxl(ws, source_row_num, insert_row)
        _copy_row_formulas_openpyxl(ws, source_row_num, insert_row)
        ws.cell(insert_row, 1).value = int(code) if code.isdigit() else code
        ws.cell(insert_row, 2).value = source_data["row_names"].get(code, "")
        master_map = _read_master_layout(
            ws,
            editable_end_row=_editable_end_row(ws, protected_start_row),
        )
        _set_row_total_formula_openpyxl(ws, insert_row, max(master_map["subprogram_map"].values()))
        step_index += 1
        if progress_callback is not None:
            pct = 35 + int(18 * step_index / total_steps)
            progress_callback(pct, f"Updating Master sheet... ({step_index}/{total_steps})")

    return _read_master_layout(ws, editable_end_row=_editable_end_row(ws, protected_start_row))


def _insert_partial_column_openpyxl(ws: Any, insert_col: int, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    max_col_before = ws.max_column
    ws.move_range(
        f"{get_column_letter(insert_col)}{start_row}:{get_column_letter(max_col_before)}{end_row}",
        rows=0,
        cols=1,
        translate=True,
    )
    for row in range(start_row, end_row + 1):
        ws.cell(row, insert_col).value = None


def _copy_column_formulas_openpyxl(
    ws: Any,
    source_col: int,
    target_col: int,
    start_row: int,
    end_row: int,
) -> None:
    if source_col < 1 or end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        source_value = ws.cell(row, source_col).value
        if isinstance(source_value, str) and source_value.startswith("="):
            try:
                translated = Translator(
                    source_value,
                    origin=f"{get_column_letter(source_col)}{row}",
                ).translate_formula(f"{get_column_letter(target_col)}{row}")
            except Exception:
                translated = source_value
            ws.cell(row, target_col).value = translated


def _copy_row_formulas_openpyxl(ws: Any, source_row: int, target_row: int) -> None:
    if source_row < 1 or source_row > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        source_value = ws.cell(source_row, col).value
        if isinstance(source_value, str) and source_value.startswith("="):
            try:
                translated = Translator(
                    source_value,
                    origin=f"{get_column_letter(col)}{source_row}",
                ).translate_formula(f"{get_column_letter(col)}{target_row}")
            except Exception:
                translated = source_value
            ws.cell(target_row, col).value = translated


def _copy_column_format_openpyxl(ws: Any, source_col: int, target_col: int) -> None:
    if source_col < 1 or source_col > ws.max_column:
        return
    for row in range(1, ws.max_row + 1):
        ws.cell(row, target_col)._style = copy(  # noqa: SLF001
            ws.cell(row, source_col)._style  # noqa: SLF001
        )
        if ws.row_dimensions[row].height is not None:
            ws.row_dimensions[row].height = ws.row_dimensions[row].height
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    ws.column_dimensions[target_letter].hidden = ws.column_dimensions[source_letter].hidden


def _copy_row_format_openpyxl(ws: Any, source_row: int, target_row: int) -> None:
    if source_row < 1 or source_row > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        ws.cell(target_row, col)._style = copy(  # noqa: SLF001
            ws.cell(source_row, col)._style  # noqa: SLF001
        )
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden


def _ensure_master_total_formulas_openpyxl(ws: Any, master_map: dict[str, Any]) -> None:
    if not master_map["row_map"] or not master_map["subprogram_map"]:
        return
    last_col = max(master_map["subprogram_map"].values())
    for row_idx in master_map["row_map"].values():
        _set_row_total_formula_openpyxl(ws, row_idx, last_col)


def _set_row_total_formula_openpyxl(ws: Any, row_idx: int, last_col: int) -> None:
    last_col_letter = get_column_letter(last_col)
    ws.cell(row_idx, 3).value = f"=SUM(D{row_idx}:{last_col_letter}{row_idx})"


# ---------------------------------------------------------------------------
# Internal helpers — positional utilities
# ---------------------------------------------------------------------------


def _find_insert_col(subprogram_map: dict[str, int], new_code: str) -> int:
    ordered = sorted(subprogram_map.items(), key=lambda item: item[1])
    for code, col in ordered:
        if _sort_key(new_code) < _sort_key(code):
            return col
    return (ordered[-1][1] + 1) if ordered else 4


def _find_insert_row(row_map: dict[str, int], new_code: str) -> int:
    ordered = sorted(row_map.items(), key=lambda item: item[1])
    for code, row in ordered:
        if _sort_key(new_code) < _sort_key(code):
            return row
    return (ordered[-1][1] + 1) if ordered else 6


# ---------------------------------------------------------------------------
# Internal helpers — Excel COM path (Windows + pywin32 only)
# ---------------------------------------------------------------------------


def _can_use_excel_native() -> bool:
    import os

    return os.name == "nt" and HAVE_COM


def _is_retryable_excel_error(exc: Exception) -> bool:  # pragma: no cover
    if pywintypes is not None and isinstance(exc, pywintypes.com_error):
        hresult = getattr(exc, "hresult", None)
        if hresult in EXCEL_RETRY_HRESULTS:
            return True
    args = getattr(exc, "args", ())
    if args and isinstance(args[0], int) and args[0] in EXCEL_RETRY_HRESULTS:
        return True
    text = " ".join(str(part) for part in args).lower()
    return "call was rejected by callee" in text or "server busy" in text or "拒绝接收呼叫" in text


def _call_excel_with_retries(  # pragma: no cover
    action: Any,
    attempts: int = 20,
    initial_delay_seconds: float = 0.2,
) -> Any:
    delay = initial_delay_seconds
    for attempt in range(1, attempts + 1):
        try:
            return action()
        except Exception as exc:
            if not _is_retryable_excel_error(exc) or attempt == attempts:
                raise
            try:
                if pythoncom is not None:
                    pythoncom.PumpWaitingMessages()
            except Exception:
                pass
            time.sleep(delay)
            delay = min(delay * 1.5, 1.5)
    raise RuntimeError("Excel retry loop ended unexpectedly.")


def _copy_file_with_retries(  # pragma: no cover
    source_path: Path,
    target_path: Path,
    attempts: int = 20,
    initial_delay_seconds: float = 0.2,
) -> None:
    delay = initial_delay_seconds
    for attempt in range(1, attempts + 1):
        try:
            if target_path.exists():
                target_path.unlink()
            shutil.copy2(source_path, target_path)
            return
        except PermissionError:
            if attempt == attempts:
                raise
            time.sleep(delay)
            delay = min(delay * 1.5, 1.5)


def _run_excel_native(  # pragma: no cover
    template_path: Path,
    output_path: Path,
    source_data: dict[str, Any],
    missing_master_codes: list[str],
    missing_source_codes: list[str],
    missing_subprogram_codes: list[str],
    source_extra_subprogram_codes: list[str],
    progress_callback: ProgressFn | None = None,
) -> tuple[int, int]:
    pythoncom.CoInitialize()
    excel = None
    wb = None
    master_ws = None
    compass_ws = None
    temp_dir = None
    temp_template_path: Path | None = None
    temp_output_path: Path | None = None
    excel_pid: int | None = None
    matched_cells = 0
    matched_rows = 0
    save_completed = False

    def _progress(percent: int, message: str) -> None:
        if progress_callback is not None:
            progress_callback(percent, message)

    try:
        if output_path.exists():
            output_path.unlink()

        temp_dir = Path(tempfile.mkdtemp(prefix="budget_automation_"))
        temp_template_path = temp_dir / template_path.name
        temp_output_path = temp_dir / output_path.name
        shutil.copy2(template_path, temp_template_path)
        if temp_output_path.exists():
            temp_output_path.unlink()

        excel = win32com_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.UserControl = False
        try:
            excel.AskToUpdateLinks = False
        except Exception:
            pass
        try:
            excel.AutomationSecurity = 3
        except Exception:
            pass
        if win32process is not None:
            try:
                excel_pid = win32process.GetWindowThreadProcessId(excel.Hwnd)[1]
            except Exception:
                excel_pid = None

        _progress(30, "Opening workbook in Excel...")
        wb = _call_excel_with_retries(
            lambda: excel.Workbooks.Open(
                str(temp_template_path.resolve()),
                UpdateLinks=0,
                ReadOnly=False,
                IgnoreReadOnlyRecommended=True,
                AddToMru=False,
                Notify=False,
            )
        )
        master_ws = wb.Worksheets(_MASTER_SHEET)
        compass_ws = (
            wb.Worksheets(_COMPASS_SHEET)
            if _COMPASS_SHEET in [ws.Name for ws in wb.Worksheets]
            else None
        )
        protected_start_row = _find_protected_start_row_excel(master_ws)

        _progress(42, "Updating Master sheet...")
        master_map = _insert_source_only_items_excel(
            master_ws,
            source_data,
            missing_source_codes,
            source_extra_subprogram_codes,
            protected_start_row=protected_start_row,
            progress_callback=_progress,
        )
        _progress(58, "Writing imported values...")
        matched_cells, matched_rows = _populate_master_excel(
            master_ws,
            master_map,
            source_data,
            progress_callback=_progress,
        )
        _ensure_master_total_formulas_excel(master_ws, master_map)
        if compass_ws is not None:
            _progress(72, "Refreshing Compass sheet...")
            _populate_compass_excel(compass_ws, source_data)
        _progress(84, "Applying mismatch highlighting...")
        _apply_mismatch_highlights_excel(
            master_ws,
            compass_ws,
            master_map,
            source_data,
            missing_master_codes,
            missing_source_codes,
            missing_subprogram_codes,
            source_extra_subprogram_codes,
        )

        file_format = 52 if output_path.suffix.lower() == ".xlsm" else 51
        _progress(94, "Saving output workbook...")
        _call_excel_with_retries(
            lambda: wb.SaveAs(
                str(temp_output_path.resolve()),
                FileFormat=file_format,
                ConflictResolution=2,
            )
        )
        rebound_buttons = _rebind_macro_buttons_excel(
            wb,
            workbook_names=[template_path.name, temp_template_path.name],
            output_workbook_name=output_path.name,
        )
        if rebound_buttons:
            _progress(
                96,
                f"Updating macro button bindings... ({rebound_buttons} item(s))",
            )
            _call_excel_with_retries(lambda: wb.Save())
        save_completed = True
    except Exception as exc:
        raise _BudgetError(f"Excel save failed: {exc}") from exc
    finally:
        compass_ws = None
        master_ws = None
        gc.collect()
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        wb = None
        gc.collect()
        try:
            if excel is not None:
                excel.ScreenUpdating = True
                excel.EnableEvents = True
                excel.Quit()
        except Exception:
            pass
        excel = None
        gc.collect()
        if excel_pid is not None:
            _ensure_excel_process_exited(excel_pid)
    try:
        if save_completed and temp_output_path is not None:
            _progress(98, "Copying workbook to the selected output folder...")
            _copy_file_with_retries(temp_output_path, output_path)
    except Exception as exc:
        raise _BudgetError(
            "Workbook was saved in Excel but could not be copied to the "
            f"selected output folder: {exc}"
        ) from exc
    finally:
        if temp_dir is not None:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return matched_cells, matched_rows


def _ensure_excel_process_exited(  # pragma: no cover
    pid: int, timeout_seconds: float = 5.0
) -> None:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if not _is_process_running(pid):
            return
        time.sleep(0.25)
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/T", "/F"],
        capture_output=True,
        text=True,
        check=False,
    )


def _is_process_running(pid: int) -> bool:  # pragma: no cover
    import ctypes

    process_query_limited_information = 0x1000
    synchronize = 0x00100000
    handle = ctypes.windll.kernel32.OpenProcess(  # type: ignore[attr-defined]
        process_query_limited_information | synchronize, False, pid
    )
    if not handle:
        return False
    try:
        return True
    finally:
        ctypes.windll.kernel32.CloseHandle(handle)  # type: ignore[attr-defined]


def _rebind_macro_buttons_excel(  # pragma: no cover
    workbook: Any,
    workbook_names: list[str],
    output_workbook_name: str,
) -> int:
    updated = 0
    candidates = [name for name in workbook_names if name]
    if not candidates:
        return 0

    sheet_count = int(_call_excel_with_retries(lambda: workbook.Sheets.Count))
    for sheet_index in range(1, sheet_count + 1):
        try:
            sheet = _call_excel_with_retries(lambda idx=sheet_index: workbook.Sheets(idx))
            shapes = _call_excel_with_retries(lambda current_sheet=sheet: current_sheet.Shapes)
            shape_count = int(
                _call_excel_with_retries(lambda current_shapes=shapes: current_shapes.Count)
            )
        except Exception:
            continue

        for index in range(1, shape_count + 1):
            try:
                shape = _call_excel_with_retries(
                    lambda idx=index, current_shapes=shapes: current_shapes.Item(idx)
                )
                on_action = str(
                    _call_excel_with_retries(
                        lambda current_shape=shape: current_shape.OnAction or ""
                    )
                ).strip()
            except Exception:
                continue

            new_on_action = _rewrite_shape_on_action(
                on_action=on_action,
                workbook_names=candidates,
                output_workbook_name=output_workbook_name,
            )
            if new_on_action and new_on_action != on_action:
                try:
                    _call_excel_with_retries(
                        lambda current_shape=shape, action=new_on_action: setattr(
                            current_shape, "OnAction", action
                        )
                    )
                    updated += 1
                except Exception:
                    pass

    return updated


def _rewrite_shape_on_action(
    on_action: str,
    workbook_names: list[str],
    output_workbook_name: str,
) -> str:
    action = on_action.strip()
    if not action or "!" not in action:
        return action
    workbook_part, macro_part = action.rsplit("!", 1)
    workbook_part_lower = workbook_part.lower()
    workbook_match = re.match(r"^'?([^']+?)'?$", workbook_part.strip())
    normalized_workbook_part = workbook_match.group(1) if workbook_match else workbook_part.strip()
    normalized_workbook_lower = normalized_workbook_part.lower()
    for workbook_name in workbook_names:
        workbook_name_lower = workbook_name.lower()
        if (
            workbook_name_lower == normalized_workbook_lower
            or workbook_name_lower in workbook_part_lower
            or f"[{workbook_name_lower}]" in workbook_part_lower
        ):
            return f"'{output_workbook_name}'!{macro_part}"
    return action


# ---------------------------------------------------------------------------
# Internal helpers — Excel COM population
# ---------------------------------------------------------------------------


def _find_protected_start_row_excel(ws: Any) -> int | None:  # pragma: no cover
    max_row = int(ws.UsedRange.Rows.Count)
    for row in range(1, max_row + 1):
        if _clean_string(ws.Cells(row, 1).Value) == _PROTECTED_START_CODE:
            return row
    return None


def _editable_end_row_excel(  # pragma: no cover
    ws: Any, protected_start_row: int | None
) -> int:
    max_row = int(ws.UsedRange.Rows.Count)
    return (protected_start_row - 1) if protected_start_row else max_row


def _read_master_layout_excel(  # pragma: no cover
    ws: Any, editable_end_row: int | None = None
) -> dict[str, Any]:
    row_map: dict[str, int] = {}
    row_names: dict[str, str] = {}
    subprogram_map: dict[str, int] = {}
    subprogram_names: dict[str, str] = {}
    max_col = int(ws.UsedRange.Columns.Count)
    max_row = editable_end_row if editable_end_row is not None else int(ws.UsedRange.Rows.Count)
    for col in range(4, max_col + 1):
        code = _clean_string(ws.Cells(4, col).Value)
        if code:
            subprogram_map[code] = col
            subprogram_names[code] = _clean_string(ws.Cells(5, col).Value)
    for row in range(1, max_row + 1):
        account_code = _clean_string(ws.Cells(row, 1).Value)
        if account_code.isdigit() and len(account_code) >= 5:
            row_map[account_code] = row
            row_names[account_code] = _clean_string(ws.Cells(row, 2).Value)
    return {
        "row_map": row_map,
        "row_names": row_names,
        "subprogram_map": subprogram_map,
        "subprogram_names": subprogram_names,
        "row_codes": set(row_map.keys()),
        "subprogram_codes": set(subprogram_map.keys()),
    }


def _populate_master_excel(  # pragma: no cover
    ws: Any,
    master_map: dict[str, Any],
    source_data: dict[str, Any],
    progress_callback: ProgressFn | None = None,
) -> tuple[int, int]:
    sorted_rows = sorted(master_map["row_map"].items(), key=lambda item: item[1])
    sorted_cols = sorted(master_map["subprogram_map"].items(), key=lambda item: item[1])
    if not sorted_rows or not sorted_cols:
        return 0, 0

    matched_cells = 0
    matched_rows = 0
    segments = _contiguous_row_segments(sorted_rows)
    total_segments = max(1, len(segments))
    for seg_index, segment in enumerate(segments, start=1):
        first_row = segment[0][1]
        last_row = segment[-1][1]
        first_col = sorted_cols[0][1]
        last_col = sorted_cols[-1][1]
        ws.Range(ws.Cells(first_row, first_col), ws.Cells(last_row, last_col)).ClearContents()

        matrix = []
        for account_code, _target_row in segment:
            source_row = source_data["rows"].get(account_code)
            if source_row is not None:
                matched_rows += 1
            row_values = []
            for subprogram_code, _target_col in sorted_cols:
                value: Any = ""
                if source_row is not None:
                    source_idx = source_data["subprogram_map"].get(subprogram_code)
                    if source_idx is not None and source_idx < len(source_row):
                        parsed = _parse_source_number(source_row[source_idx])
                        value = "" if parsed is None else parsed
                        if parsed is not None:
                            matched_cells += 1
                row_values.append(value)
            matrix.append(tuple(row_values))

        ws.Range(ws.Cells(first_row, first_col), ws.Cells(last_row, last_col)).Value = tuple(matrix)
        if progress_callback is not None:
            pct = 58 + int(18 * seg_index / total_segments)
            progress_callback(
                pct,
                f"Writing imported values... ({seg_index}/{total_segments} sections)",
            )
    return matched_cells, matched_rows


def _contiguous_row_segments(
    sorted_rows: list[tuple[str, int]],
) -> list[list[tuple[str, int]]]:
    if not sorted_rows:
        return []
    segments: list[list[tuple[str, int]]] = [[sorted_rows[0]]]
    for item in sorted_rows[1:]:
        if item[1] == segments[-1][-1][1] + 1:
            segments[-1].append(item)
        else:
            segments.append([item])
    return segments


def _populate_compass_excel(  # pragma: no cover
    ws: Any, source_data: dict[str, Any]
) -> None:
    max_existing_row = int(ws.UsedRange.Rows.Count)
    max_existing_col = int(ws.UsedRange.Columns.Count)
    ws.Cells.ClearContents()

    ordered_subprogram_codes = [
        code for code in source_data["subprogram_map"].keys() if code not in {"Total", "EI/SP"}
    ]
    ordered_account_codes = list(source_data["rows"].keys())

    if ordered_subprogram_codes:
        header_row_1 = []
        header_row_2 = []
        for subprogram_code in ordered_subprogram_codes:
            header_row_1.append(
                int(subprogram_code) if subprogram_code.isdigit() else subprogram_code
            )
            header_row_2.append(source_data["subprogram_names"].get(subprogram_code, ""))
        ws.Range(ws.Cells(1, 4), ws.Cells(1, 3 + len(header_row_1))).Value = (tuple(header_row_1),)
        ws.Range(ws.Cells(2, 4), ws.Cells(2, 3 + len(header_row_2))).Value = (tuple(header_row_2),)

    ws.Cells(2, 2).Value = "EI/SP"
    ws.Cells(2, 3).Value = "Total"

    data_matrix = []
    for account_code in ordered_account_codes:
        source_row = source_data["rows"][account_code]
        total_idx = source_data["subprogram_map"].get("Total")
        total_value: Any = ""
        if total_idx is not None and total_idx < len(source_row):
            parsed_total = _parse_source_number(source_row[total_idx])
            total_value = "" if parsed_total is None else parsed_total

        row_values: list[Any] = [
            int(account_code) if account_code.isdigit() else account_code,
            source_data["row_names"].get(account_code, ""),
            total_value,
        ]
        for subprogram_code in ordered_subprogram_codes:
            source_idx = source_data["subprogram_map"].get(subprogram_code)
            value = ""
            if source_idx is not None and source_idx < len(source_row):
                parsed_value = _parse_source_number(source_row[source_idx])
                value = "" if parsed_value is None else parsed_value
            row_values.append(value)
        data_matrix.append(tuple(row_values))

    if data_matrix:
        ws.Range(
            ws.Cells(3, 1),
            ws.Cells(2 + len(data_matrix), len(data_matrix[0])),
        ).Value = tuple(data_matrix)

    last_used_row = len(ordered_account_codes) + 2
    last_used_col = len(ordered_subprogram_codes) + 3
    if max_existing_row > last_used_row:
        ws.Range(
            ws.Cells(last_used_row + 1, 1),
            ws.Cells(max_existing_row, max_existing_col),
        ).ClearContents()
    if max_existing_col > last_used_col and last_used_row >= 1:
        ws.Range(
            ws.Cells(1, last_used_col + 1),
            ws.Cells(last_used_row, max_existing_col),
        ).ClearContents()


def _apply_mismatch_highlights_excel(  # pragma: no cover
    master_ws: Any,
    compass_ws: Any,
    master_map: dict[str, Any],
    source_data: dict[str, Any],
    missing_master_codes: list[str],
    missing_source_codes: list[str],
    missing_subprogram_codes: list[str],
    source_extra_subprogram_codes: list[str],
) -> None:
    master_last_col = (
        max(master_map["subprogram_map"].values()) if master_map["subprogram_map"] else 3
    )
    master_last_row = max(master_map["row_map"].values()) if master_map["row_map"] else 5
    for code in missing_master_codes:
        row_idx = master_map["row_map"].get(code)
        if row_idx:
            master_ws.Range(
                master_ws.Cells(row_idx, 1),
                master_ws.Cells(row_idx, master_last_col),
            ).Interior.Color = bgr_int(HL_MISMATCH)
    for code in missing_subprogram_codes:
        col_idx = master_map["subprogram_map"].get(code)
        if col_idx:
            master_ws.Range(
                master_ws.Cells(4, col_idx),
                master_ws.Cells(master_last_row, col_idx),
            ).Interior.Color = bgr_int(HL_MISMATCH)
    for code in missing_source_codes:
        row_idx = master_map["row_map"].get(code)
        if row_idx:
            master_ws.Range(
                master_ws.Cells(row_idx, 1),
                master_ws.Cells(row_idx, master_last_col),
            ).Interior.Color = bgr_int(HL_SOURCE_ONLY)
    for code in source_extra_subprogram_codes:
        col_idx = master_map["subprogram_map"].get(code)
        if col_idx:
            master_ws.Range(
                master_ws.Cells(4, col_idx),
                master_ws.Cells(master_last_row, col_idx),
            ).Interior.Color = bgr_int(HL_SOURCE_ONLY)

    if compass_ws is None:
        return

    ordered_subprogram_codes = [
        code for code in source_data["subprogram_map"].keys() if code not in {"Total", "EI/SP"}
    ]
    ordered_account_codes = list(source_data["rows"].keys())
    extra_rows = set(missing_source_codes)
    extra_cols = set(source_extra_subprogram_codes)

    for col_idx, subprogram_code in enumerate(ordered_subprogram_codes, start=4):
        if subprogram_code in extra_cols:
            compass_ws.Range(
                compass_ws.Cells(1, col_idx),
                compass_ws.Cells(len(ordered_account_codes) + 2, col_idx),
            ).Interior.Color = bgr_int(HL_SOURCE_ONLY)

    last_col = len(ordered_subprogram_codes) + 3
    for row_idx, account_code in enumerate(ordered_account_codes, start=3):
        if account_code in extra_rows:
            compass_ws.Range(
                compass_ws.Cells(row_idx, 1),
                compass_ws.Cells(row_idx, last_col),
            ).Interior.Color = bgr_int(HL_SOURCE_ONLY)


def _insert_source_only_items_excel(  # pragma: no cover
    ws: Any,
    source_data: dict[str, Any],
    missing_source_codes: list[str],
    source_extra_subprogram_codes: list[str],
    protected_start_row: int | None = None,
    progress_callback: ProgressFn | None = None,
) -> dict[str, Any]:
    editable_end = _editable_end_row_excel(ws, protected_start_row)
    master_map = _read_master_layout_excel(ws, editable_end_row=editable_end)
    total_steps = max(1, len(source_extra_subprogram_codes) + len(missing_source_codes))
    step_index = 0

    for code in sorted(source_extra_subprogram_codes, key=_sort_key):
        insert_col = _find_insert_col(master_map["subprogram_map"], code)
        copy_from_col = insert_col - 1 if insert_col > 4 else insert_col + 1
        _insert_partial_column_excel(ws, insert_col, 1, editable_end)
        try:
            ws.Columns(insert_col).ColumnWidth = ws.Columns(copy_from_col).ColumnWidth
        except Exception:
            pass
        _copy_column_formulas_excel(ws, copy_from_col, insert_col, 1, editable_end)
        ws.Cells(4, insert_col).Value = int(code) if code.isdigit() else code
        ws.Cells(5, insert_col).Value = source_data["subprogram_names"].get(code, "")
        master_map = _read_master_layout_excel(
            ws,
            editable_end_row=_editable_end_row_excel(ws, protected_start_row),
        )
        step_index += 1
        if progress_callback is not None:
            pct = 42 + int(14 * step_index / total_steps)
            progress_callback(pct, f"Updating Master sheet... ({step_index}/{total_steps})")

    for code in sorted(missing_source_codes, key=_sort_key):
        insert_row = _find_insert_row(master_map["row_map"], code)
        if protected_start_row is not None and insert_row >= protected_start_row:
            continue
        copy_from_row = insert_row - 1 if insert_row > 6 else insert_row + 1
        ws.Rows(insert_row).Insert()
        try:
            ws.Rows(insert_row).RowHeight = ws.Rows(copy_from_row).RowHeight
        except Exception:
            pass
        _copy_row_formulas_excel(ws, copy_from_row, insert_row, int(ws.UsedRange.Columns.Count))
        ws.Cells(insert_row, 1).Value = int(code) if code.isdigit() else code
        ws.Cells(insert_row, 2).Value = source_data["row_names"].get(code, "")
        master_map = _read_master_layout_excel(
            ws,
            editable_end_row=_editable_end_row_excel(ws, protected_start_row),
        )
        _set_row_total_formula_excel(ws, insert_row, max(master_map["subprogram_map"].values()))
        step_index += 1
        if progress_callback is not None:
            pct = 42 + int(14 * step_index / total_steps)
            progress_callback(pct, f"Updating Master sheet... ({step_index}/{total_steps})")

    return _read_master_layout_excel(
        ws, editable_end_row=_editable_end_row_excel(ws, protected_start_row)
    )


def _insert_partial_column_excel(  # pragma: no cover
    ws: Any, insert_col: int, start_row: int, end_row: int
) -> None:
    if end_row < start_row:
        return
    ws.Range(ws.Cells(start_row, insert_col), ws.Cells(end_row, insert_col)).Insert(Shift=-4161)


def _copy_column_formulas_excel(  # pragma: no cover
    ws: Any,
    source_col: int,
    target_col: int,
    start_row: int,
    end_row: int,
) -> None:
    if source_col < 1 or end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        try:
            if bool(ws.Cells(row, source_col).HasFormula):
                ws.Cells(row, target_col).FormulaR1C1 = ws.Cells(row, source_col).FormulaR1C1
        except Exception:
            pass


def _copy_row_formulas_excel(  # pragma: no cover
    ws: Any, source_row: int, target_row: int, end_col: int
) -> None:
    if source_row < 1:
        return
    for col in range(1, int(end_col) + 1):
        try:
            if bool(ws.Cells(source_row, col).HasFormula):
                ws.Cells(target_row, col).FormulaR1C1 = ws.Cells(source_row, col).FormulaR1C1
        except Exception:
            pass


def _ensure_master_total_formulas_excel(  # pragma: no cover
    ws: Any, master_map: dict[str, Any]
) -> None:
    if not master_map["row_map"] or not master_map["subprogram_map"]:
        return
    last_col = max(master_map["subprogram_map"].values())
    for row_idx in master_map["row_map"].values():
        _set_row_total_formula_excel(ws, row_idx, last_col)


def _set_row_total_formula_excel(  # pragma: no cover
    ws: Any, row_idx: int, last_col: int
) -> None:
    last_col_letter = get_column_letter(last_col)
    ws.Cells(row_idx, 3).Formula = f"=SUM(D{row_idx}:{last_col_letter}{row_idx})"


# ---------------------------------------------------------------------------
# Internal helpers — pure utility
# ---------------------------------------------------------------------------


def _is_mismatch_account_code(code: str) -> bool:
    return code.isdigit() and len(code) == 5 and code[0] in {"7", "8"}


def _clean_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _sort_key(code: str) -> tuple[int, Any]:
    text = str(code).strip()
    if text.isdigit():
        return (0, int(text))
    return (1, text)


def _parse_source_number(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if text in {"#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}:
        return None
    text = text.replace(",", "")
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return text
