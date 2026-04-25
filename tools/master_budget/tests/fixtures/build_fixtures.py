"""Fixture builders for Master Budget Compass Autofill integration tests.

LAYOUT ASSUMPTIONS (documented for code-review audit at end of M3):

Expense Sub-Program XLSX (source / Compass export)
---------------------------------------------------
Sheet 0 (any name).
Row 1 — sub-program codes: col 0 empty (account-code header), col 1 is the
        account-name column, cols 2+ hold sub-program codes (e.g. "11234").
        "Total" and "EI/SP" are special sentinel codes that logic.py excludes
        from mismatch tracking.
Row 2 — sub-program names (human-readable labels aligned under row-1 codes).
Row 3+ — data rows: col 0 = account code (5+ digit string), col 1 = name,
          cols 2+ = amounts matching the sub-program code columns.

This matches the v1 _read_source() logic in budget_automation.py: row 0 is
the subprogram_row, row 1 is the name_row, rows 2+ are data_rows keyed by
col-0 account code.

Master Budget XLSM (template)
------------------------------
Sheet name: "Master"  (required by logic._MASTER_SHEET == "Master")
Row 4 — sub-program codes starting from col 4 (D).  _read_master_layout()
        reads ws.cell(4, col) for col in range(4, max_column+1).
Row 5 — sub-program names (human-readable labels).
Rows 6+ — account-code rows: col 1 (A) = account code (isdigit, len>=5),
           col 2 (B) = account name.  Values land in subprogram columns.
Col 3 (C) — row totals (formula injected by logic, not needed in template).

Why these choices:
- Sheet name "Master" is hardcoded as logic._MASTER_SHEET.
- Sub-program header at row 4 and name row at row 5 are hardcoded in
  _read_master_layout() (both openpyxl and Excel-COM variants).
- Sub-program columns start at col 4: _read_master_layout() loops from 4.
- Account codes detected by: isdigit() and len >= 5.  Rows 1-5 are reserved
  for headers/subprogram labels, so account codes start at row 6.
- Amounts are written by _populate_master() into ws.cell(target_row, target_col)
  where target_col is the subprogram column (>=4).

No macros are included — this builder produces a plain .xlsm saved with
keep_vba=False (openpyxl just writes the correct extension).  Macro
preservation is Windows-only and tested separately.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl  # noqa: E402

# ---------------------------------------------------------------------------
# Public builders
# ---------------------------------------------------------------------------


def build_expense_file(path: Path, rows: list[tuple[str, Decimal]]) -> None:
    """Write a minimal Compass Expense Sub-Program XLSX to *path*.

    Each entry in *rows* is ``(account_code, amount)``.  A single synthetic
    sub-program column "11111" is used so the logic has at least one
    sub-program to map.

    Layout written
    --------------
    Row 1 (index 0): ["", "",    "11111"]        <- sub-program codes
    Row 2 (index 1): ["", "",    "Test SP"]       <- sub-program names
    Row 3+ (index 2+): [account_code, name, amount]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    sp_code = "11111"
    sp_name = "Test SP"

    # Row 1: sub-program code header
    ws.cell(1, 1).value = ""  # account-code column placeholder
    ws.cell(1, 2).value = ""  # name column placeholder
    ws.cell(1, 3).value = sp_code

    # Row 2: sub-program name header
    ws.cell(2, 1).value = ""
    ws.cell(2, 2).value = ""
    ws.cell(2, 3).value = sp_name

    # Data rows (from row 3 onwards)
    for row_idx, (account_code, amount) in enumerate(rows, start=3):
        ws.cell(row_idx, 1).value = account_code
        ws.cell(row_idx, 2).value = f"Account {account_code}"
        ws.cell(row_idx, 3).value = float(amount)

    wb.save(path)


def build_master_template(path: Path, account_codes: list[str]) -> None:
    """Write a minimal Master Budget XLSM template to *path*.

    The template has a "Master" sheet containing the given account codes.
    A single synthetic sub-program column "11111" is placed in col D (col 4)
    so the tool has something to populate.

    Layout written (Master sheet)
    ------------------------------
    Row 4, col D (4): sub-program code "11111"
    Row 5, col D (4): sub-program name "Test SP"
    Row 6+:  col A (1) = account code, col B (2) = account name
    """
    wb = openpyxl.Workbook()
    # Rename default sheet to "Master"
    ws = wb.active
    assert ws is not None
    ws.title = "Master"

    sp_code = "11111"
    sp_name = "Test SP"

    # Row 4: sub-program code header (col D = col 4)
    ws.cell(4, 4).value = sp_code
    # Row 5: sub-program name header
    ws.cell(5, 4).value = sp_name

    # Account code rows starting at row 6
    for row_idx, account_code in enumerate(account_codes, start=6):
        ws.cell(row_idx, 1).value = account_code
        ws.cell(row_idx, 2).value = f"Account {account_code}"

    # Save as .xlsm (openpyxl will write the correct content type)
    wb.save(path)


# ---------------------------------------------------------------------------
# Multi-subprogram variant (used by mismatch tests)
# ---------------------------------------------------------------------------


def build_expense_file_multi(
    path: Path,
    account_rows: list[tuple[str, dict[str, Decimal]]],
    subprogram_codes: list[str],
) -> None:
    """Write a Compass XLSX with multiple sub-program columns.

    Parameters
    ----------
    account_rows:
        List of ``(account_code, {sp_code: amount})``.
    subprogram_codes:
        Ordered list of sub-program codes to appear as columns (col 3+).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    # Row 1: sub-program code header
    ws.cell(1, 1).value = ""
    ws.cell(1, 2).value = ""
    for col_offset, sp_code in enumerate(subprogram_codes):
        ws.cell(1, 3 + col_offset).value = sp_code

    # Row 2: sub-program name header
    ws.cell(2, 1).value = ""
    ws.cell(2, 2).value = ""
    for col_offset, sp_code in enumerate(subprogram_codes):
        ws.cell(2, 3 + col_offset).value = f"SP {sp_code}"

    # Data rows (from row 3 onwards)
    for row_idx, (account_code, amounts) in enumerate(account_rows, start=3):
        ws.cell(row_idx, 1).value = account_code
        ws.cell(row_idx, 2).value = f"Account {account_code}"
        for col_offset, sp_code in enumerate(subprogram_codes):
            amount = amounts.get(sp_code)
            ws.cell(row_idx, 3 + col_offset).value = float(amount) if amount is not None else None

    wb.save(path)


def build_master_template_multi(
    path: Path,
    account_codes: list[str],
    subprogram_codes: list[str],
) -> None:
    """Write a Master Budget template with multiple sub-program columns.

    Sub-program codes go into row 4 starting at col D (col 4).
    Account codes go into col A starting at row 6.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Master"

    # Row 4+5: sub-program header rows
    for col_offset, sp_code in enumerate(subprogram_codes):
        ws.cell(4, 4 + col_offset).value = sp_code
        ws.cell(5, 4 + col_offset).value = f"SP {sp_code}"

    # Account code rows starting at row 6
    for row_idx, account_code in enumerate(account_codes, start=6):
        ws.cell(row_idx, 1).value = account_code
        ws.cell(row_idx, 2).value = f"Account {account_code}"

    wb.save(path)


# ---------------------------------------------------------------------------
# Type-checking helper (not used at runtime, but satisfies mypy --strict)
# ---------------------------------------------------------------------------

__all__: list[Any] = [
    "build_expense_file",
    "build_master_template",
    "build_expense_file_multi",
    "build_master_template_multi",
]
