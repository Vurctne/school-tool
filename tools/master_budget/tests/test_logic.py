from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

import pytest  # noqa: F401 — used by pytest.raises / pytest.mark

from tools.master_budget.logic import (
    ImportSummary,
    _clean_string,
    _contiguous_row_segments,
    _find_insert_col,
    _find_insert_row,
    _is_mismatch_account_code,
    _normalize_rows,
    _parse_source_number,
    _rewrite_shape_on_action,
    _sort_key,
    suggest_output_name,
)

# ---------------------------------------------------------------------------
# suggest_output_name
# ---------------------------------------------------------------------------


def test_suggest_output_name_format(tmp_path: Path) -> None:
    """Name must match ``<stem>_AUTO_YYYYMMDD_HHMM.xlsm`` in the source dir."""
    master = tmp_path / "Master_Budget_2025.xlsm"
    master.touch()
    result = suggest_output_name(master)
    pattern = re.compile(r"Master_Budget_2025_AUTO_\d{8}_\d{4}\.xlsm$")
    assert pattern.search(result), f"Unexpected name: {result}"


def test_suggest_output_name_preserves_directory(tmp_path: Path) -> None:
    """Suggested name must live in the same directory as the master file."""
    master = tmp_path / "sub" / "Budget.xlsm"
    master.parent.mkdir()
    master.touch()
    result = suggest_output_name(master)
    assert result.startswith(str(tmp_path / "sub"))


def test_suggest_output_name_preserves_suffix(tmp_path: Path) -> None:
    """Suffix is inherited from the master file (xlsm or xlsx)."""
    master_xlsm = tmp_path / "Budget.xlsm"
    master_xlsm.touch()
    assert suggest_output_name(master_xlsm).endswith(".xlsm")


def test_suggest_compare_output_name_format(tmp_path: Path) -> None:
    """Round 39 — Compare output filename pins to
    ``MasterBudget_Compare_<YYYYMMDD_HHMM>.xlsx`` next to file A."""
    from tools.master_budget.logic import suggest_compare_output_name

    a = tmp_path / "Master_Budget_2025.xlsm"
    a.touch()
    result = suggest_compare_output_name(a)
    pattern = re.compile(r"MasterBudget_Compare_\d{8}_\d{4}\.xlsx$")
    assert pattern.search(result), f"Unexpected compare name: {result}"
    # Lives next to the source file, not in cwd.
    assert result.startswith(str(tmp_path))

    master_xlsx = tmp_path / "Budget.xlsx"
    master_xlsx.touch()
    assert suggest_output_name(master_xlsx).endswith(".xlsx")


# ---------------------------------------------------------------------------
# Account-code matching algorithm
# ---------------------------------------------------------------------------


def test_account_code_matching_algorithm_matched() -> None:
    """Codes present in both master and source → neither mismatch set."""
    master_codes: set[str] = {"71001", "71002", "80001"}
    source_codes: set[str] = {"71001", "71002", "80001"}
    missing_master = sorted(master_codes - source_codes)
    missing_source = sorted(source_codes - master_codes)
    assert missing_master == []
    assert missing_source == []


def test_account_code_matching_algorithm_mismatch_master() -> None:
    """Code in master but not source → shows up in missing_master."""
    master_codes: set[str] = {"71001", "71002"}
    source_codes: set[str] = {"71001"}
    missing_master = sorted(master_codes - source_codes)
    assert missing_master == ["71002"]


def test_account_code_matching_algorithm_source_only() -> None:
    """Code in source but not master → shows up in missing_source."""
    master_codes: set[str] = {"71001"}
    source_codes: set[str] = {"71001", "80999"}
    missing_source = sorted(source_codes - master_codes)
    assert missing_source == ["80999"]


def test_is_mismatch_account_code_filters_correctly() -> None:
    """Only 5-digit codes starting with 7 or 8 are tracked."""
    assert _is_mismatch_account_code("71001") is True
    assert _is_mismatch_account_code("80001") is True
    assert _is_mismatch_account_code("26201") is False  # doesn't start with 7/8
    assert _is_mismatch_account_code("7100") is False  # only 4 digits
    assert _is_mismatch_account_code("710011") is False  # 6 digits
    assert _is_mismatch_account_code("ABCDE") is False  # non-numeric


# ---------------------------------------------------------------------------
# Amount / number parsing
# ---------------------------------------------------------------------------


def test_parse_source_number_integer() -> None:
    """Whole-number floats come back as int."""
    assert _parse_source_number("1000") == 1000
    assert isinstance(_parse_source_number("1000"), int)


def test_parse_source_number_currency_string() -> None:
    """Comma-separated strings are stripped and parsed."""
    assert _parse_source_number("1,234,567") == 1234567


def test_parse_source_number_float() -> None:
    """Non-integer floats come back as float."""
    result = _parse_source_number("1234.56")
    assert result == 1234.56
    assert isinstance(result, float)


def test_parse_source_number_errors_return_none() -> None:
    """Excel error strings must yield None."""
    for err in ("#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"):
        assert _parse_source_number(err) is None


def test_parse_source_number_empty() -> None:
    """Empty / None input → None."""
    assert _parse_source_number(None) is None
    assert _parse_source_number("") is None
    assert _parse_source_number("   ") is None


def test_parse_source_number_non_numeric_text_passthrough() -> None:
    """Non-numeric, non-error strings are returned as-is."""
    assert _parse_source_number("N/A") == "N/A"


# ---------------------------------------------------------------------------
# _clean_string
# ---------------------------------------------------------------------------


def test_clean_string_float_integer() -> None:
    """Float that is a whole number should return str(int)."""
    assert _clean_string(12345.0) == "12345"


def test_clean_string_none() -> None:
    assert _clean_string(None) == ""


def test_clean_string_strips_whitespace() -> None:
    assert _clean_string("  hello  ") == "hello"


# ---------------------------------------------------------------------------
# _sort_key
# ---------------------------------------------------------------------------


def test_sort_key_numeric_before_alpha() -> None:
    """Numeric codes sort before alphabetic ones."""
    keys = [_sort_key(c) for c in ["ABC", "12345", "99999", "DEF"]]
    assert keys[1] < keys[0]  # "12345" < "ABC"
    assert keys[2] < keys[3]  # "99999" < "DEF"


def test_sort_key_numeric_ordering() -> None:
    codes = ["80001", "71002", "71001"]
    assert sorted(codes, key=_sort_key) == ["71001", "71002", "80001"]


# ---------------------------------------------------------------------------
# _find_insert_col / _find_insert_row
# ---------------------------------------------------------------------------


def test_find_insert_col_in_order() -> None:
    subprogram_map = {"100": 4, "200": 5, "300": 6}
    # 150 < 200 → insert at col 5
    assert _find_insert_col(subprogram_map, "150") == 5


def test_find_insert_col_at_end() -> None:
    subprogram_map = {"100": 4, "200": 5}
    # 999 > all → insert after last (col 6)
    assert _find_insert_col(subprogram_map, "999") == 6


def test_find_insert_row_in_order() -> None:
    row_map = {"71001": 7, "71003": 8}
    # 71002 between 71001 and 71003 → insert at row 8
    assert _find_insert_row(row_map, "71002") == 8


def test_find_insert_row_at_end() -> None:
    row_map = {"71001": 7, "71002": 8}
    assert _find_insert_row(row_map, "80999") == 9


# ---------------------------------------------------------------------------
# _contiguous_row_segments
# ---------------------------------------------------------------------------


def test_contiguous_row_segments_single_block() -> None:
    rows: list[tuple[str, int]] = [("A", 5), ("B", 6), ("C", 7)]
    result = _contiguous_row_segments(rows)
    assert len(result) == 1
    assert result[0] == rows


def test_contiguous_row_segments_two_blocks() -> None:
    rows: list[tuple[str, int]] = [("A", 5), ("B", 6), ("C", 10), ("D", 11)]
    result = _contiguous_row_segments(rows)
    assert len(result) == 2
    assert result[0] == [("A", 5), ("B", 6)]
    assert result[1] == [("C", 10), ("D", 11)]


def test_contiguous_row_segments_empty() -> None:
    assert _contiguous_row_segments([]) == []


# ---------------------------------------------------------------------------
# _normalize_rows
# ---------------------------------------------------------------------------


def test_normalize_rows_pads_short_rows() -> None:
    rows = [["A", "B", "C"], ["X"]]
    result = _normalize_rows(rows)
    assert all(len(r) == 3 for r in result)
    assert result[1] == ["X", "", ""]


def test_normalize_rows_drops_blank_rows() -> None:
    rows = [["", ""], ["A", "B"], ["", "  "]]
    result = _normalize_rows(rows)
    assert len(result) == 1
    assert result[0] == ["A", "B"]


# ---------------------------------------------------------------------------
# _rewrite_shape_on_action
# ---------------------------------------------------------------------------


def test_rewrite_shape_on_action_matches_workbook() -> None:
    result = _rewrite_shape_on_action(
        on_action="'template.xlsm'!RunMacro",
        workbook_names=["template.xlsm"],
        output_workbook_name="output.xlsm",
    )
    assert result == "'output.xlsm'!RunMacro"


def test_rewrite_shape_on_action_no_match() -> None:
    result = _rewrite_shape_on_action(
        on_action="'other.xlsm'!RunMacro",
        workbook_names=["template.xlsm"],
        output_workbook_name="output.xlsm",
    )
    assert result == "'other.xlsm'!RunMacro"


def test_rewrite_shape_on_action_no_bang() -> None:
    result = _rewrite_shape_on_action(
        on_action="RunMacro",
        workbook_names=["template.xlsm"],
        output_workbook_name="output.xlsm",
    )
    assert result == "RunMacro"


# ---------------------------------------------------------------------------
# ImportSummary dataclass
# ---------------------------------------------------------------------------


def test_import_summary_is_frozen(tmp_path: Path) -> None:
    """ImportSummary must be frozen (immutable)."""
    summary = ImportSummary(
        matched_rows=10,
        matched_cells=50,
        mismatch_account_codes=["71001"],
        mismatch_subprogram_codes=[],
        source_only_account_codes=[],
        source_only_subprogram_codes=[],
        output_path=tmp_path / "out.xlsm",
    )
    with pytest.raises((AttributeError, TypeError)):
        summary.matched_rows = 99  # type: ignore[misc]


# ---------------------------------------------------------------------------
# pywin32 guard — import smoke (non-Windows must not raise)
# ---------------------------------------------------------------------------


@pytest.mark.skipif(sys.platform == "win32", reason="non-Windows only")
def test_logic_imports_cleanly_on_non_windows() -> None:
    """Module must import without error on non-Windows (HAVE_COM == False)."""
    import tools.master_budget.logic as logic

    assert logic.HAVE_COM is False


# ---------------------------------------------------------------------------
# Regression: Excel COM path highlight colours match openpyxl path + tokens
# ---------------------------------------------------------------------------


def test_com_interior_colour_matches_hl_mismatch() -> None:
    """The Win32 COM highlight path must derive its ``Interior.Color`` BGR
    integer from ``bgr_int(HL_MISMATCH)`` / ``bgr_int(HL_SOURCE_ONLY)`` —
    never from a hard-coded literal. A 1-off silently paints the wrong
    shade on Windows (the openpyxl CI path won't catch it).

    Bug fixed 2026-04-24 after the M3 code review; the test was migrated
    on 2026-04-25 from a source-scan of the hard-coded integers to a
    token+helper derivation check so it survives the fills-module refactor
    while preserving the same regression guarantee.
    """
    import inspect

    from toolkit.fills import bgr_int
    from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY
    from tools.master_budget import logic as mb_logic

    # 1. Math: the encoding of each token is exactly the historical value.
    assert bgr_int(HL_MISMATCH) == 13421812, "HL_MISMATCH (#F4CCCC) must BGR-encode to 13421812"
    assert bgr_int(HL_SOURCE_ONLY) == 14282978, (
        "HL_SOURCE_ONLY (#E2F0D9) must BGR-encode to 14282978"
    )
    # 1a. Historical off-by-one drift: 13421823 decodes to #FFCCCC, NOT
    # HL_MISMATCH.  Guard against it reappearing if someone fat-fingers the
    # token.
    assert bgr_int(HL_MISMATCH) != 13421823

    # 2. Wiring: the COM path must derive from the helpers; no hard-coded
    # integer may reappear in the function body.
    src = inspect.getsource(mb_logic._apply_mismatch_highlights_excel)
    assert "bgr_int(HL_MISMATCH)" in src, (
        "COM highlight path must use bgr_int(HL_MISMATCH), not a literal."
    )
    assert "bgr_int(HL_SOURCE_ONLY)" in src, (
        "COM highlight path must use bgr_int(HL_SOURCE_ONLY), not a literal."
    )
    assert "13421812" not in src, (
        "COM path must derive from bgr_int(HL_MISMATCH); literal forbidden."
    )
    assert "14282978" not in src, (
        "COM path must derive from bgr_int(HL_SOURCE_ONLY); literal forbidden."
    )


def test_win32com_client_submodule_is_explicitly_imported() -> None:
    """Regression: the v1 code path used ``win32com.client.DispatchEx(...)``
    after only importing ``win32com``, which fails at runtime with
    ``AttributeError: module 'win32com' has no attribute 'client'``. The fix
    is to also import ``win32com.client`` (which both imports the submodule
    AND attaches it to the parent package as a side effect), and to use a
    dedicated ``win32com_client`` handle at the call site.

    This test inspects the module source to ensure the fix stays in place
    across future refactors; a Linux CI runner can't import pywin32 at all,
    so we verify via string match rather than import.
    """
    import inspect

    from tools.master_budget import logic as mb_logic

    src = inspect.getsource(mb_logic)
    # The explicit submodule import is mandatory.
    assert 'import_module("win32com.client")' in src, (
        "Missing: _importlib.import_module('win32com.client'). "
        "Without it win32com.client raises AttributeError on Windows."
    )
    # The COM-dispatch call site must use the explicit handle, not the
    # attribute-access path that used to break.
    assert "win32com_client.DispatchEx(" in src, (
        "DispatchEx should be called via win32com_client, not win32com.client."
    )
    assert "win32com.client.DispatchEx(" not in src, (
        "The attribute-access path win32com.client.DispatchEx is the broken "
        "pattern we're guarding against. Use win32com_client.DispatchEx instead."
    )


# ---------------------------------------------------------------------------
# Round 27 — Compare two Master Budget files
# ---------------------------------------------------------------------------


def _build_compare_master(
    path: Path,
    sp_data: dict[str, dict[str, Any]],
    *,
    revenue_label: str = "Total Estimated Revenue",
    expenditure_label: str = "Total Proposed Expenditure Current Year",
    funds_label: str = "Total Estimated Funds Held future years",
) -> None:
    """Write a synthetic Master Budget XLSM for the compare tests.

    Layout (Master sheet):
      Row 4, cols D+: sub-program codes
      Row 5, cols D+: sub-program names
      Row 6: revenue_label in col B, values per sp in cols D+
      Row 7: expenditure_label in col B, values per sp in cols D+
      Row 8: funds_label in col B, values per sp in cols D+
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Master"

    sp_codes = list(sp_data.keys())
    for col_offset, sp in enumerate(sp_codes):
        ws.cell(4, 4 + col_offset).value = sp
        ws.cell(5, 4 + col_offset).value = f"SP {sp}"

    rows_meta = [
        (6, "revenue", revenue_label),
        (7, "expenditure", expenditure_label),
        (8, "funds_held", funds_label),
    ]
    for row_idx, key, label in rows_meta:
        ws.cell(row_idx, 2).value = label
        for col_offset, sp in enumerate(sp_codes):
            v = sp_data[sp].get(key)
            ws.cell(row_idx, 4 + col_offset).value = v

    wb.save(path)


def test_compare_no_differences(tmp_path: Path) -> None:
    """Two identical files yield zero diff rows but track sub-program names."""
    from tools.master_budget.logic import compare_master_budgets

    sp_data = {
        "1100": {"revenue": 10000.0, "expenditure": 8000.0, "funds_held": 2000.0},
        "1200": {"revenue": 5000.0, "expenditure": 5500.0, "funds_held": -500.0},
    }
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, sp_data)
    _build_compare_master(b, sp_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    assert summary.rows == []
    assert summary.only_in_a == []
    assert summary.only_in_b == []


def test_compare_value_differences(tmp_path: Path) -> None:
    """Sub-programs with differing values appear in summary.rows; identical
    sub-programs are filtered out (per "show only the differences" spec)."""
    from tools.master_budget.logic import compare_master_budgets

    a_data = {
        "1100": {"revenue": 10000.0, "expenditure": 8000.0, "funds_held": 2000.0},
        "1200": {"revenue": 5000.0, "expenditure": 5500.0, "funds_held": -500.0},
        "1300": {"revenue": 0.0, "expenditure": 0.0, "funds_held": 0.0},
    }
    b_data = {
        "1100": {"revenue": 12000.0, "expenditure": 8000.0, "funds_held": 2000.0},
        "1200": {"revenue": 5000.0, "expenditure": 5500.0, "funds_held": -500.0},
        "1300": {"revenue": 0.0, "expenditure": 100.0, "funds_held": 0.0},
    }
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, a_data)
    _build_compare_master(b, b_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    sp_codes_in_diffs = {r.sub_program for r in summary.rows}
    assert sp_codes_in_diffs == {"1100", "1300"}
    assert summary.only_in_a == []
    assert summary.only_in_b == []

    row_1100 = next(r for r in summary.rows if r.sub_program == "1100")
    assert row_1100.a_revenue == 10000.0
    assert row_1100.b_revenue == 12000.0
    assert row_1100.only_in is None


def test_compare_sub_program_only_in_one_file(tmp_path: Path) -> None:
    """Sub-programs unique to A or B are always reported via only_in flag."""
    from tools.master_budget.logic import compare_master_budgets

    a_data = {
        "1100": {"revenue": 1000.0, "expenditure": 500.0, "funds_held": 500.0},
        "1200": {"revenue": 2000.0, "expenditure": 0.0, "funds_held": 2000.0},
    }
    b_data = {
        "1100": {"revenue": 1000.0, "expenditure": 500.0, "funds_held": 500.0},
        "1300": {"revenue": 3000.0, "expenditure": 1000.0, "funds_held": 2000.0},
    }
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, a_data)
    _build_compare_master(b, b_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    assert summary.only_in_a == ["1200"]
    assert summary.only_in_b == ["1300"]

    only_a_row = next(r for r in summary.rows if r.sub_program == "1200")
    assert only_a_row.only_in == "A"
    assert only_a_row.a_revenue == 2000.0
    assert only_a_row.b_revenue is None

    only_b_row = next(r for r in summary.rows if r.sub_program == "1300")
    assert only_b_row.only_in == "B"
    assert only_b_row.b_revenue == 3000.0
    assert only_b_row.a_revenue is None


def test_compare_label_match_substring_fallback(tmp_path: Path) -> None:
    """Substring-matching label kicks in when the exact label isn't found."""
    from tools.master_budget.logic import compare_master_budgets

    sp_data = {"1100": {"revenue": 10.0, "expenditure": 5.0, "funds_held": 5.0}}
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, sp_data, revenue_label="Total Estimated Revenue 2026")
    _build_compare_master(b, sp_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    matched_label, kind = summary.label_match_a["revenue"]
    assert kind == "substring"
    assert "Revenue" in matched_label

    matched_b_label, kind_b = summary.label_match_b["revenue"]
    assert kind_b == "exact"


def test_compare_label_match_missing_treated_as_blank(tmp_path: Path) -> None:
    """When a target label is absent, that metric reads as None for every
    sub-program in that file (and the label_match entry shows ``missing``)."""
    from tools.master_budget.logic import compare_master_budgets

    sp_data = {"1100": {"revenue": 10.0, "expenditure": 5.0, "funds_held": 5.0}}
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, sp_data, funds_label="Some Unrelated Column Name")
    _build_compare_master(b, sp_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    _, kind = summary.label_match_a["funds_held"]
    assert kind == "missing"


def test_compare_write_xlsx_creates_file(tmp_path: Path) -> None:
    """write_compare_xlsx returns a fresh summary with output_path set and
    actually writes a readable .xlsx with the expected header row."""
    import openpyxl

    from tools.master_budget.logic import compare_master_budgets, write_compare_xlsx

    a_data = {"1100": {"revenue": 100.0, "expenditure": 50.0, "funds_held": 50.0}}
    b_data = {"1100": {"revenue": 200.0, "expenditure": 50.0, "funds_held": 50.0}}
    a = tmp_path / "a.xlsm"
    b = tmp_path / "b.xlsm"
    _build_compare_master(a, a_data)
    _build_compare_master(b, b_data)

    summary = compare_master_budgets(a, b, lambda *_: None)
    out = tmp_path / "compare.xlsx"
    new_summary = write_compare_xlsx(summary, out)
    assert new_summary.output_path == out
    assert out.exists()

    wb = openpyxl.load_workbook(out, data_only=True)
    try:
        ws = wb["Compare"]
        assert ws.cell(2, 1).value == "Sub-program"
        assert ws.cell(3, 1).value == "1100"
        assert ws.cell(3, 5).value == 100.0
    finally:
        wb.close()


def test_compare_rejects_same_file(tmp_path: Path) -> None:
    """Passing the same path for A and B raises a clear error."""
    from tools.master_budget.logic import compare_master_budgets

    sp_data = {"1100": {"revenue": 10.0, "expenditure": 5.0, "funds_held": 5.0}}
    a = tmp_path / "same.xlsm"
    _build_compare_master(a, sp_data)

    with pytest.raises(Exception, match="different files"):
        compare_master_budgets(a, a, lambda *_: None)
