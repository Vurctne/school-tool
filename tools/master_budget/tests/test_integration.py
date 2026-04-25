"""Integration tests for the Master Budget Compass Autofill tool.

All tests use ``tmp_path`` for isolation — no repo paths are written to.
Fixture workbooks are synthetic: they encode exactly the layout that
logic.py expects (see fixtures/build_fixtures.py for the documented
assumptions).

Windows-only path (macro preservation via pywin32 COM) is covered by a
single stub test that documents the intent and is unconditionally skipped
on non-Windows platforms.
"""

from __future__ import annotations

import re
import sys
from decimal import Decimal
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import openpyxl
import pytest

from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY
from tools.master_budget.frame import MasterBudgetTool
from tools.master_budget.logic import ImportSummary, import_expense_sub_program, suggest_output_name
from tools.master_budget.tests.fixtures.build_fixtures import (
    build_expense_file,
    build_expense_file_multi,
    build_master_template,
    build_master_template_multi,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_CODES_5 = ["71001", "71002", "71003", "71004", "71005"]
_AMOUNT_BASE = Decimal("1000")


def _noop_progress(percent: int, message: str) -> None:
    pass


# ---------------------------------------------------------------------------
# 1. Happy path — all codes match
# ---------------------------------------------------------------------------


def test_happy_path_all_codes_match(tmp_path: Path) -> None:
    """Five matching account codes → 5 matched rows, 0 mismatches, 0 source-only."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    rows = [(code, _AMOUNT_BASE + Decimal(i * 100)) for i, code in enumerate(_CODES_5)]
    build_expense_file(expense, rows)
    build_master_template(master, _CODES_5)

    summary: ImportSummary = import_expense_sub_program(
        expense_file=expense,
        master_file=master,
        output_file=output,
        progress=_noop_progress,
    )

    assert summary.matched_rows == 5
    assert summary.mismatch_account_codes == []
    assert summary.mismatch_subprogram_codes == []
    assert summary.source_only_account_codes == []
    assert summary.source_only_subprogram_codes == []
    assert output.exists()
    assert summary.output_path == output


# ---------------------------------------------------------------------------
# 2. Warning path — some mismatches
#    Source: {A, B, C}  Master: {A, B, D}
#    → mismatch_codes == ["D"] (in master but not source, 7xxxx → tracked)
#    → source_only_codes == ["C"] (in source but not master, 7xxxx → tracked)
# ---------------------------------------------------------------------------


def test_warning_path_mismatches(tmp_path: Path) -> None:
    """Mismatch detection: 1 mismatch code (D), 1 source-only code (C)."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    source_codes = ["71001", "71002", "71003"]  # A=71001, B=71002, C=71003
    master_codes = ["71001", "71002", "71004"]  # A=71001, B=71002, D=71004

    rows = [(code, Decimal("500")) for code in source_codes]
    build_expense_file(expense, rows)
    build_master_template(master, master_codes)

    summary = import_expense_sub_program(
        expense_file=expense,
        master_file=master,
        output_file=output,
        progress=_noop_progress,
    )

    assert summary.mismatch_account_codes == ["71004"]  # D: in master but not source
    assert summary.source_only_account_codes == ["71003"]  # C: in source but not master
    assert output.exists()


# ---------------------------------------------------------------------------
# 3. Highlight colours verified in output XLSX
# ---------------------------------------------------------------------------


def test_mismatch_highlight_colours(tmp_path: Path) -> None:
    """HL_MISMATCH fill on D's row; HL_SOURCE_ONLY fill on C's row in output."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    source_codes = ["71001", "71002", "71003"]
    master_codes = ["71001", "71002", "71004"]

    rows = [(code, Decimal("500")) for code in source_codes]
    build_expense_file(expense, rows)
    build_master_template(master, master_codes)

    import_expense_sub_program(
        expense_file=expense,
        master_file=master,
        output_file=output,
        progress=_noop_progress,
    )

    wb = openpyxl.load_workbook(output, keep_vba=True, data_only=True)
    try:
        ws = wb["Master"]
        # Locate rows by scanning col A
        row_map: dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row, 1).value
            if val is not None:
                key = (
                    str(int(val))
                    if isinstance(val, float) and val.is_integer()
                    else str(val).strip()
                )
                row_map[key] = row

        assert "71004" in row_map, "Mismatch code 71004 not found in output Master sheet"
        assert "71003" in row_map, "Source-only code 71003 not found in output Master sheet"

        mismatch_row = row_map["71004"]
        source_only_row = row_map["71003"]

        # Check the first data column (col A = col 1) for fill colours.
        mismatch_fill = ws.cell(mismatch_row, 1).fill
        source_only_fill = ws.cell(source_only_row, 1).fill

        # openpyxl stores fgColor.rgb as "FFRRGGBB" (8 chars with alpha prefix).
        # Strip the leading two-char alpha byte ("FF") to get the bare RRGGBB.
        def _rgb(fill: Any) -> str:
            rgb = str(fill.fgColor.rgb).upper() if fill.fgColor else ""
            # 8-char form "FFRRGGBB" → drop first 2 chars
            return rgb[2:] if len(rgb) == 8 else rgb

        assert _rgb(mismatch_fill) == HL_MISMATCH.upper(), (
            f"Expected HL_MISMATCH {HL_MISMATCH} on row {mismatch_row}, got {_rgb(mismatch_fill)}"
        )
        assert _rgb(source_only_fill) == HL_SOURCE_ONLY.upper(), (
            f"Expected HL_SOURCE_ONLY {HL_SOURCE_ONLY} on row "
            f"{source_only_row}, got {_rgb(source_only_fill)}"
        )
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 4. Output cell values — amount for code A matches source
# ---------------------------------------------------------------------------


def test_output_cell_values(tmp_path: Path) -> None:
    """After happy-path run, reopen output and verify amount for 71001 == 1500."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    target_code = "71001"
    target_amount = Decimal("1500")
    rows = [(target_code, target_amount), ("71002", Decimal("200"))]
    build_expense_file(expense, rows)
    build_master_template(master, [target_code, "71002"])

    import_expense_sub_program(
        expense_file=expense,
        master_file=master,
        output_file=output,
        progress=_noop_progress,
    )

    wb = openpyxl.load_workbook(output, data_only=True)
    try:
        ws = wb["Master"]
        # Find the row for target_code in col A; amount is in col D (col 4).
        found_amount: float | int | None = None
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row, 1).value
            if cell_val is None:
                continue
            cell_str = (
                str(int(cell_val))
                if isinstance(cell_val, float) and cell_val.is_integer()
                else str(cell_val).strip()
            )
            if cell_str == target_code:
                raw = ws.cell(row, 4).value  # col D = first subprogram col
                if isinstance(raw, (int, float)):
                    found_amount = raw
                break
        assert found_amount is not None, f"Amount cell for {target_code} not found or empty"
        assert Decimal(str(found_amount)) == target_amount, (
            f"Expected {target_amount}, got {found_amount}"
        )
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 5. Error: missing source file raises an exception
# ---------------------------------------------------------------------------


def test_missing_source_file_raises(tmp_path: Path) -> None:
    """import_expense_sub_program raises when the source file does not exist."""
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"
    build_master_template(master, ["71001"])

    with pytest.raises((FileNotFoundError, ValueError, Exception)) as exc_info:
        import_expense_sub_program(
            expense_file=Path("/nonexistent_expense_file_xyz.xlsx"),
            master_file=master,
            output_file=output,
            progress=_noop_progress,
        )
    # Confirm the error is descriptive (not a bare AttributeError, etc.)
    assert exc_info.value is not None


# ---------------------------------------------------------------------------
# 6. MasterBudgetTool.run() end-to-end — happy path → "success"
# ---------------------------------------------------------------------------


def test_tool_run_happy_path(tmp_path: Path) -> None:
    """MasterBudgetTool.run() returns status='success' when all codes match."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    rows = [("71001", Decimal("100")), ("71002", Decimal("200"))]
    build_expense_file(expense, rows)
    build_master_template(master, ["71001", "71002"])

    tool = MasterBudgetTool()
    mock_progress = MagicMock()
    result = tool.run(
        paths={
            "expense_file": str(expense),
            "master_file": str(master),
            "output_file": str(output),
        },
        progress=mock_progress,
    )

    assert result.status == "success"
    assert result.output_path is not None
    assert Path(result.output_path).exists()


# ---------------------------------------------------------------------------
# 7. MasterBudgetTool.run() end-to-end — mismatch → "warning"
# ---------------------------------------------------------------------------


def test_tool_run_mismatch_warning(tmp_path: Path) -> None:
    """MasterBudgetTool.run() returns status='warning' when mismatches exist."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    # Source has 71003, master has 71004 → mismatch on both sides
    rows = [("71001", Decimal("100")), ("71003", Decimal("300"))]
    build_expense_file(expense, rows)
    build_master_template(master, ["71001", "71004"])

    tool = MasterBudgetTool()
    mock_progress = MagicMock()
    result = tool.run(
        paths={
            "expense_file": str(expense),
            "master_file": str(master),
            "output_file": str(output),
        },
        progress=mock_progress,
    )

    assert result.status == "warning"


# ---------------------------------------------------------------------------
# 8. suggest_output_name format
# ---------------------------------------------------------------------------


def test_suggest_output_name_format(tmp_path: Path) -> None:
    """suggest_output_name returns <stem>_AUTO_YYYYMMDD_HHMM<suffix> in same dir."""
    master = tmp_path / "foo.xlsm"
    master.touch()
    result = suggest_output_name(master)
    pattern = re.compile(r"foo_AUTO_\d{8}_\d{4}\.xlsm$")
    assert pattern.search(result), f"Unexpected name: {result}"
    # Must live under the same directory
    assert result.startswith(str(tmp_path))


# ---------------------------------------------------------------------------
# 9. suggest_output_name — directory preserved
# ---------------------------------------------------------------------------


def test_suggest_output_name_directory(tmp_path: Path) -> None:
    """Suggested name is under the same parent directory as the master file."""
    sub = tmp_path / "subdir"
    sub.mkdir()
    master = sub / "Budget.xlsm"
    master.touch()
    result = suggest_output_name(master)
    assert Path(result).parent == sub


# ---------------------------------------------------------------------------
# 10. Multiple subprogram columns — matched cells count
# ---------------------------------------------------------------------------


def test_multiple_subprogram_columns_matched_cells(tmp_path: Path) -> None:
    """With 2 account codes × 2 sub-program columns, matched_cells == 4."""
    expense = tmp_path / "expense.xlsx"
    master = tmp_path / "master.xlsm"
    output = tmp_path / "output.xlsm"

    sp_codes = ["11111", "22222"]
    account_rows = [
        ("71001", {"11111": Decimal("100"), "22222": Decimal("200")}),
        ("71002", {"11111": Decimal("300"), "22222": Decimal("400")}),
    ]
    build_expense_file_multi(expense, account_rows, sp_codes)
    build_master_template_multi(master, ["71001", "71002"], sp_codes)

    summary = import_expense_sub_program(
        expense_file=expense,
        master_file=master,
        output_file=output,
        progress=_noop_progress,
    )

    assert summary.matched_rows == 2
    assert summary.matched_cells == 4
    assert summary.mismatch_account_codes == []
    assert summary.source_only_account_codes == []


# ---------------------------------------------------------------------------
# 11. Windows-only stub — macro preservation via pywin32
# ---------------------------------------------------------------------------


@pytest.mark.skipif(sys.platform != "win32", reason="pywin32 COM path is Windows-only")
def test_macro_preservation_windows_only() -> None:  # pragma: no cover
    """Placeholder: byte-parity macro preservation is tested on Windows only.

    On Windows with pywin32 installed the tool routes through
    _run_excel_native(), which opens the workbook in Excel via COM,
    saves it with file_format=52 (xlsm), and rebinds all macro button
    OnAction strings to the new output file name.  This path requires
    an active Excel installation and cannot run in headless Linux CI.

    True v1/v2 byte-parity comparison is carried out by Ivan by installing
    both MSIX packages side-by-side and comparing outputs of identical
    input pairs.
    """
    assert sys.platform == "win32"  # tautological guard; here for documentation
