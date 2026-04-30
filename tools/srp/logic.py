from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Literal

import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from toolkit.base_tool import ProgressFn
from toolkit.fills import argb
from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY

# ---------------------------------------------------------------------------
# Currency parsing
# ---------------------------------------------------------------------------

_DASH_RE = re.compile(r"^[—–\-]{1,2}$")


def parse_decimal(raw: str) -> Decimal:
    """Convert an SRP currency string to Decimal."""
    text = raw.strip() if raw else ""
    if not text or _DASH_RE.match(text):
        return Decimal("0")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    text = text.lstrip("$").strip()
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse {raw!r} as a decimal: {exc}") from exc


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

# Simplified 4-category scheme: the XLSX columns already show per-version
# values, so the category just flags *whether* anything changed.
SrpCategory = Literal["unchanged", "changed", "new", "removed"]


@dataclass(frozen=True)
class SrpLine:
    ref: int
    section: str
    description: str
    # Per-slot values (None when that version was not provided / line absent)
    indicative: Decimal | None
    confirmed: Decimal | None
    revised1: Decimal | None
    revised2: Decimal | None
    category: SrpCategory
    # Variances between adjacent provided versions (generalised list)
    # Each entry: (label, value)
    adjacent_variances: list[tuple[str, Decimal | None]]
    # Legacy convenience fields retained for backward-compat / frame.py use
    variance_ind_to_conf: Decimal | None
    variance_conf_to_rev1: Decimal | None
    variance_rev1_to_rev2: Decimal | None
    # "variance" = variance between the first and last provided values
    variance: Decimal | None
    pct: Decimal | None


@dataclass(frozen=True)
class SrpSummary:
    lines: list[SrpLine]
    # Totals for the FIRST and LAST provided version
    total_first: Decimal
    total_last: Decimal
    # Legacy field aliases pointing at first/last for backward-compat
    total_indicative: Decimal  # = total_first
    total_confirmed: Decimal  # = total_last
    counts: dict[str, int]
    output_path: Path
    # Ordered list of provided version labels e.g. ["Indicative", "Confirmed"]
    version_labels: list[str]
    # Legacy boolean flags (derived from version_labels)
    has_revised1: bool
    has_revised2: bool


# ---------------------------------------------------------------------------
# Skip / header patterns
# ---------------------------------------------------------------------------

_SKIP_ABOVE_RE = re.compile(
    r"^("
    r"Department of Education"
    r"|Student Resource Package"
    r"|Host School"
    r"|Budget Type"
    r"|School\s.+SFO Index"
    r"|Type Secondary"
    r"|Secondary Students"
    r"|Equity \(Social Disadvantage\) Students"
    r"|Primary Level"
    r"|Policy and Advisory"
    r"|Ref\s+Students"
    r")",
    re.IGNORECASE,
)

_SUBTOTAL_RE = re.compile(r"^\$[\d,]+\.\d{2}")
_GRAND_TOTAL_RE = re.compile(r"^TOTAL STUDENT RESOURCE PACKAGE", re.IGNORECASE)
_FOOTER_RE = re.compile(r"^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}")
_EQUITY_REFORM_RE = re.compile(r"^Equity Reform Implementation Statement", re.IGNORECASE)


def _is_section_header(line: str) -> bool:
    if _SKIP_ABOVE_RE.match(line):
        return False
    if _SUBTOTAL_RE.match(line):
        return False
    if _GRAND_TOTAL_RE.match(line):
        return False
    if _FOOTER_RE.match(line):
        return False
    if _EQUITY_REFORM_RE.match(line):
        return False
    if "$" in line:
        return False
    return True


# ---------------------------------------------------------------------------
# PDF parser
# ---------------------------------------------------------------------------


def parse_srp_pdf(
    pdf_path: Path,
) -> dict[tuple[int, str], tuple[str, Decimal]]:
    """Parse a VIC DoE SRP Budget Details PDF."""
    if not pdf_path.exists():
        raise ValueError(f"File not found: {pdf_path}")

    result: dict[tuple[int, str], tuple[str, Decimal]] = {}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                raise ValueError(
                    "SRP PDF appears empty or unrecognised; "
                    "check the file is a VIC DoE SRP Budget Details export"
                )

            for page in pdf.pages:
                found_tables = page.find_tables()

                for t_idx, found_table in enumerate(found_tables):
                    prev_bottom = found_tables[t_idx - 1].bbox[3] if t_idx > 0 else 0.0
                    above_crop = page.crop((0.0, prev_bottom, page.width, found_table.bbox[1]))
                    above_text = above_crop.extract_text() or ""
                    section = _extract_section_header(above_text)

                    for row in found_table.extract():
                        if not row or not row[0]:
                            continue
                        desc = row[0].strip()
                        ref_raw = row[1].strip() if len(row) > 1 and row[1] else ""
                        if not ref_raw.isdigit():
                            continue
                        ref = int(ref_raw)
                        total_raw = row[-1].strip() if row[-1] else "$0.00"
                        try:
                            total = parse_decimal(total_raw)
                        except ValueError:
                            total = Decimal("0")
                        key = (ref, desc)
                        if key not in result:
                            result[key] = (section, total)

    except OSError as exc:
        raise ValueError(
            "SRP PDF appears empty or unrecognised; "
            "check the file is a VIC DoE SRP Budget Details export"
        ) from exc

    if not result:
        raise ValueError(
            "SRP PDF appears empty or unrecognised; "
            "check the file is a VIC DoE SRP Budget Details export"
        )

    return result


def parse_srp_pdf_ordered(
    pdf_path: Path,
) -> tuple[dict[tuple[int, str], tuple[str, Decimal]], list[tuple[int, str]]]:
    """Like parse_srp_pdf but also returns parse-order key list."""
    if not pdf_path.exists():
        raise ValueError(f"File not found: {pdf_path}")

    result: dict[tuple[int, str], tuple[str, Decimal]] = {}
    ordered_keys: list[tuple[int, str]] = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                raise ValueError(
                    "SRP PDF appears empty or unrecognised; "
                    "check the file is a VIC DoE SRP Budget Details export"
                )

            for page in pdf.pages:
                found_tables = page.find_tables()

                for t_idx, found_table in enumerate(found_tables):
                    prev_bottom = found_tables[t_idx - 1].bbox[3] if t_idx > 0 else 0.0
                    above_crop = page.crop((0.0, prev_bottom, page.width, found_table.bbox[1]))
                    above_text = above_crop.extract_text() or ""
                    section = _extract_section_header(above_text)

                    for row in found_table.extract():
                        if not row or not row[0]:
                            continue
                        desc = row[0].strip()
                        ref_raw = row[1].strip() if len(row) > 1 and row[1] else ""
                        if not ref_raw.isdigit():
                            continue
                        ref = int(ref_raw)
                        total_raw = row[-1].strip() if row[-1] else "$0.00"
                        try:
                            total = parse_decimal(total_raw)
                        except ValueError:
                            total = Decimal("0")
                        key = (ref, desc)
                        if key not in result:
                            result[key] = (section, total)
                            ordered_keys.append(key)

    except OSError as exc:
        raise ValueError(
            "SRP PDF appears empty or unrecognised; "
            "check the file is a VIC DoE SRP Budget Details export"
        ) from exc

    if not result:
        raise ValueError(
            "SRP PDF appears empty or unrecognised; "
            "check the file is a VIC DoE SRP Budget Details export"
        )

    return result, ordered_keys


def _extract_section_header(above_text: str) -> str:
    for line in reversed(above_text.splitlines()):
        stripped = line.strip()
        if not stripped:
            continue
        if _is_section_header(stripped):
            return stripped
    return "Unknown"


# ---------------------------------------------------------------------------
# Diff / compare (any 2-4 versions)
# ---------------------------------------------------------------------------

# Canonical slot keys in order
_SLOT_KEYS = ("indicative", "confirmed", "revised1", "revised2")

# Map slot key -> default label when that slot is used
_SLOT_LABELS: dict[str, str] = {
    "indicative": "Indicative",
    "confirmed": "Confirmed",
    "revised1": "1st Revised",
    "revised2": "2nd Revised",
}


def _make_srp_line_generic(
    ref: int,
    section: str,
    desc: str,
    slot_values: dict[str, Decimal | None],
    version_labels: list[str],
    slot_keys: list[str],
) -> SrpLine:
    """Build an SrpLine from per-version slot values."""
    _tol = Decimal("0.01")

    ind_val = slot_values.get("indicative")
    conf_val = slot_values.get("confirmed")
    rev1_val = slot_values.get("revised1")
    rev2_val = slot_values.get("revised2")

    # Build adjacent variances between consecutive provided versions
    adjacent_variances: list[tuple[str, Decimal | None]] = []
    vals_in_order: list[Decimal | None] = [slot_values.get(k) for k in slot_keys]
    for i in range(len(slot_keys) - 1):
        lbl = version_labels[i] + "→" + version_labels[i + 1]
        v_a = vals_in_order[i]
        v_b = vals_in_order[i + 1]
        var: Decimal | None = (v_b - v_a) if (v_a is not None and v_b is not None) else None
        adjacent_variances.append((lbl, var))

    # Legacy variance fields
    has_ind = "indicative" in slot_keys
    has_conf = "confirmed" in slot_keys
    has_rev1 = "revised1" in slot_keys
    has_rev2 = "revised2" in slot_keys

    v_ind_conf: Decimal | None = None
    if has_ind and has_conf and ind_val is not None and conf_val is not None:
        v_ind_conf = conf_val - ind_val

    v_conf_rev1: Decimal | None = None
    if has_conf and has_rev1 and conf_val is not None and rev1_val is not None:
        v_conf_rev1 = rev1_val - conf_val

    v_rev1_rev2: Decimal | None = None
    if has_rev1 and has_rev2 and rev1_val is not None and rev2_val is not None:
        v_rev1_rev2 = rev2_val - rev1_val

    # "variance" = first provided value -> last provided value
    first_val = next((v for v in vals_in_order if v is not None), None)
    last_val = next((v for v in reversed(vals_in_order) if v is not None), None)
    variance_total: Decimal | None = (
        (last_val - first_val) if (first_val is not None and last_val is not None) else None
    )

    # pct based on first->last
    pct: Decimal | None = None
    if variance_total is not None and first_val is not None and first_val != Decimal("0"):
        pct = (variance_total / first_val * Decimal("100")).quantize(Decimal("0.01"))
    elif variance_total is not None and first_val is not None:
        pct = Decimal("0")

    # Determine presence across provided versions
    present = [slot_values.get(k) is not None for k in slot_keys]

    # Simplified category logic
    cat: SrpCategory
    if present[0] and not any(present[1:]):
        cat = "removed"
    elif not present[0] and any(present[1:]):
        cat = "new"
    elif variance_total is not None and abs(variance_total) >= _tol:
        cat = "changed"
    else:
        any_change = any(v is not None and abs(v) >= _tol for _, v in adjacent_variances)
        cat = "changed" if any_change else "unchanged"

    return SrpLine(
        ref=ref,
        section=section,
        description=desc,
        indicative=ind_val,
        confirmed=conf_val,
        revised1=rev1_val if has_rev1 else None,
        revised2=rev2_val if has_rev2 else None,
        category=cat,
        adjacent_variances=adjacent_variances,
        variance_ind_to_conf=v_ind_conf,
        variance_conf_to_rev1=v_conf_rev1,
        variance_rev1_to_rev2=v_rev1_rev2,
        variance=variance_total,
        pct=pct,
    )


def compare_srp(
    indicative: dict[tuple[int, str], tuple[str, Decimal]] | None = None,
    confirmed: dict[tuple[int, str], tuple[str, Decimal]] | None = None,
    revised1: dict[tuple[int, str], tuple[str, Decimal]] | None = None,
    revised2: dict[tuple[int, str], tuple[str, Decimal]] | None = None,
    indicative_order: list[tuple[int, str]] | None = None,
    versions: list[tuple[str, dict[tuple[int, str], tuple[str, Decimal]]]] | None = None,
) -> list[SrpLine]:
    """Diff two to four parsed SRP dicts and return a list of SrpLine.

    Can be called with legacy keyword arguments (indicative, confirmed,
    revised1, revised2) or with the generic ``versions`` list of
    (slot_key, data) pairs.
    """
    if versions is not None:
        slot_keys = [k for k, _ in versions]
        version_labels = [_SLOT_LABELS[k] for k in slot_keys]
        dicts: list[dict[tuple[int, str], tuple[str, Decimal]]] = [d for _, d in versions]
    else:
        slot_keys = []
        dicts = []
        if indicative is not None:
            slot_keys.append("indicative")
            dicts.append(indicative)
        if confirmed is not None:
            slot_keys.append("confirmed")
            dicts.append(confirmed)
        if revised1 is not None:
            slot_keys.append("revised1")
            dicts.append(revised1)
        if revised2 is not None:
            slot_keys.append("revised2")
            dicts.append(revised2)
        version_labels = [_SLOT_LABELS[k] for k in slot_keys]

    if len(slot_keys) < 2:
        raise ValueError(
            "At least 2 SRP versions required; got " + str(len(slot_keys)) + ": " + str(slot_keys)
        )

    all_keys: set[tuple[int, str]] = set()
    for d in dicts:
        all_keys |= set(d)

    # Determine row order
    first_order = indicative_order if indicative_order is not None else None
    if first_order is None and dicts:
        first_order = list(dicts[0].keys())

    if first_order is not None:
        ordered = [k for k in first_order if k in all_keys]
        seen: set[tuple[int, str]] = set(ordered)
        for d in dicts[1:]:
            for k in d:
                if k not in seen:
                    ordered.append(k)
                    seen.add(k)
        sorted_keys = ordered
    else:
        sorted_keys = sorted(all_keys, key=lambda k: (k[0], k[1]))

    lines: list[SrpLine] = []
    for key in sorted_keys:
        ref, desc = key
        slot_values: dict[str, Decimal | None] = {}
        section_candidates: list[str] = []

        for slot_key, d in zip(slot_keys, dicts, strict=False):
            entry = d.get(key)
            if entry is not None:
                slot_values[slot_key] = entry[1]
                section_candidates.append(entry[0])
            else:
                slot_values[slot_key] = None

        section = next((s for s in section_candidates if s), "")

        lines.append(
            _make_srp_line_generic(
                ref=ref,
                section=section,
                desc=desc,
                slot_values=slot_values,
                version_labels=version_labels,
                slot_keys=slot_keys,
            )
        )

    return lines


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

_MISMATCH_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_MISMATCH))
_SOURCE_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_SOURCE_ONLY))

_DECREASED_CATS: frozenset[str] = frozenset({"removed"})
_INCREASED_CATS: frozenset[str] = frozenset({"new"})


def _fmt_dollar(value: Decimal) -> str:
    return f"${value:,.2f}"


def _fmt_pct(value: Decimal) -> str:
    return f"{value:,.2f}%"


def _write_xlsx(
    lines: list[SrpLine],
    output_file: Path,
    version_labels: list[str],
    slot_keys: list[str],
) -> None:
    """Write lines to a new XLSX workbook at output_file."""
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    wb = Workbook()
    active = wb.active
    assert isinstance(active, Worksheet)
    ws: Worksheet = active
    ws.title = "SRP Comparison"

    headers: list[str] = ["Ref", "Section", "Description"]
    col_widths: list[int] = [8, 36, 44]

    for label in version_labels:
        headers.append(label)
        col_widths.append(18)

    for i in range(len(version_labels) - 1):
        a = version_labels[i]
        b = version_labels[i + 1]
        headers.append("Variance (" + a + "→" + b + ")")
        col_widths.append(24)

    headers += ["%", "Category"]
    col_widths += [10, 18]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, ln in enumerate(lines, start=2):
        slot_val_strs: list[str] = []
        for sk in slot_keys:
            v = (
                getattr(ln, sk)
                if sk in ("indicative", "confirmed", "revised1", "revised2")
                else None
            )
            slot_val_strs.append(_fmt_dollar(v) if v is not None else "")

        var_strs: list[str] = []
        for _lbl, v in ln.adjacent_variances:
            var_strs.append(_fmt_dollar(v) if v is not None else "")

        pct_str = _fmt_pct(ln.pct) if ln.pct is not None else ""

        row_prefix: list[str | int] = [ln.ref, ln.section, ln.description]
        values: list[str | int] = row_prefix + slot_val_strs + var_strs + [pct_str, ln.category]

        if ln.category in _DECREASED_CATS:
            fill: PatternFill | None = _MISMATCH_FILL
        elif ln.category in _INCREASED_CATS:
            fill = _SOURCE_FILL
        elif ln.category == "changed":
            if ln.variance is not None and ln.variance > Decimal("0"):
                fill = _SOURCE_FILL
            elif ln.variance is not None and ln.variance < Decimal("0"):
                fill = _MISMATCH_FILL
            else:
                fill = None
        else:
            fill = None

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill is not None:
                cell.fill = fill

    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def generate_srp_comparison(
    *,
    output_file: Path,
    progress: ProgressFn,
    indicative_pdf: Path | None = None,
    confirmed_pdf: Path | None = None,
    revised1_pdf: Path | None = None,
    revised2_pdf: Path | None = None,
) -> SrpSummary:
    """Parse two to four SRP PDFs, diff them, write an XLSX, and return SrpSummary.

    All four inputs are optional kwargs; at least two must be provided.
    Any combination of two or more versions is accepted.
    """
    provided_pairs: list[tuple[str, Path]] = [
        (slot, p)
        for slot, p in [
            ("indicative", indicative_pdf),
            ("confirmed", confirmed_pdf),
            ("revised1", revised1_pdf),
            ("revised2", revised2_pdf),
        ]
        if p is not None
    ]

    if len(provided_pairs) < 2:
        provided_labels = [_SLOT_LABELS[s] for s, _ in provided_pairs]
        raise ValueError(
            "At least 2 SRP versions required; got "
            + str(len(provided_pairs))
            + ": "
            + str(provided_labels)
        )

    slot_keys = [s for s, _ in provided_pairs]
    version_labels = [_SLOT_LABELS[s] for s in slot_keys]

    parse_end = 55
    parse_start = 10
    n = len(provided_pairs)
    step = max(1, (parse_end - parse_start) // n)

    parsed_dicts: list[dict[tuple[int, str], tuple[str, Decimal]]] = []
    first_order: list[tuple[int, str]] | None = None

    for i, (_slot, pdf_path) in enumerate(provided_pairs):
        pct = parse_start + i * step
        progress(pct, "Reading " + version_labels[i] + " SRP...")
        if i == 0:
            d, first_order = parse_srp_pdf_ordered(pdf_path)
        else:
            d = parse_srp_pdf(pdf_path)
        parsed_dicts.append(d)

    progress(55, "Comparing line by line...")
    versions_arg = list(zip(slot_keys, parsed_dicts, strict=False))
    lines = compare_srp(
        versions=versions_arg,
        indicative_order=first_order,
    )

    progress(75, "Writing workbook...")
    _write_xlsx(lines, output_file, version_labels=version_labels, slot_keys=slot_keys)

    total_first = sum((v for (_, v) in parsed_dicts[0].values()), Decimal("0"))
    total_last = sum((v for (_, v) in parsed_dicts[-1].values()), Decimal("0"))

    counts: dict[str, int] = {}
    for ln in lines:
        counts[ln.category] = counts.get(ln.category, 0) + 1

    progress(100, "Done.")

    return SrpSummary(
        lines=lines,
        total_first=total_first,
        total_last=total_last,
        total_indicative=total_first,
        total_confirmed=total_last,
        counts=counts,
        output_path=output_file,
        version_labels=version_labels,
        has_revised1="revised1" in slot_keys,
        has_revised2="revised2" in slot_keys,
    )
