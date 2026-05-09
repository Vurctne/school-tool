"""Tests for tools/srp/logic.py -- parser, diff, and XLSX writer."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY
from tools.srp.logic import (
    SrpSummary,
    compare_srp,
    generate_srp_comparison,
    parse_decimal,
    parse_srp_pdf,
    parse_srp_pdf_ordered,
)

# ---------------------------------------------------------------------------
# Sample PDF paths (must exist on disk; tests fail, never skip, if absent)
# ---------------------------------------------------------------------------

_INDICATIVE_PDF = Path("Samples/SRP budget Report/2026 Indicative Budget SRP.pdf")
_CONFIRMED_PDF = Path("Samples/SRP budget Report/2026 confirmed SRP Budget.pdf")

# ---------------------------------------------------------------------------
# Test: currency parser
# ---------------------------------------------------------------------------


class TestParseDecimal:
    def test_dollar_prefix(self) -> None:
        assert parse_decimal("$16,172,029.00") == Decimal("16172029.00")

    def test_plain_number(self) -> None:
        assert parse_decimal("1,234.56") == Decimal("1234.56")

    def test_zero(self) -> None:
        assert parse_decimal("$0.00") == Decimal("0.00")

    def test_em_dash(self) -> None:
        assert parse_decimal("\u2014") == Decimal("0")

    def test_blank(self) -> None:
        assert parse_decimal("") == Decimal("0")

    def test_negative_parens(self) -> None:
        assert parse_decimal("(500.00)") == Decimal("-500.00")

    def test_invalid_raises(self) -> None:
        with pytest.raises(ValueError):
            parse_decimal("not-a-number!!")


# ---------------------------------------------------------------------------
# Test: parse_srp_pdf against real sample PDFs
# ---------------------------------------------------------------------------


class TestParseSrpPdfIndicative:
    @pytest.fixture(scope="class")
    def data(self) -> dict[tuple[int, str], tuple[str, Decimal]]:
        return parse_srp_pdf(_INDICATIVE_PDF)

    def test_minimum_key_count(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        assert len(data) >= 20, f"Expected >= 20 keys, got {len(data)}"

    def test_all_values_are_decimal(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        for (ref, desc), (_section, total) in data.items():
            assert isinstance(total, Decimal), f"Non-Decimal total for ({ref}, {desc!r})"

    def test_all_refs_are_int(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        for ref, _ in data:
            assert isinstance(ref, int), f"Non-int ref: {ref!r}"

    def test_all_descriptions_nonempty(
        self, data: dict[tuple[int, str], tuple[str, Decimal]]
    ) -> None:
        bad = [k for k in data if not k[1].strip()]
        assert not bad, f"Empty descriptions: {bad[:3]}"

    def test_ref15_multiplicity(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        """Ref 15 must appear as two distinct (ref, description) keys in Indicative."""
        ref15_keys = [k for k in data if k[0] == 15]
        assert len(ref15_keys) >= 2, (
            f"Expected at least 2 keys for Ref 15 in Indicative, got {ref15_keys}"
        )
        descriptions = {k[1] for k in ref15_keys}
        assert len(descriptions) >= 2, (
            f"Expected distinct descriptions for Ref 15, got {descriptions}"
        )

    def test_known_key_present(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        assert (1, "Years 7 - 12 Students") in data

    def test_known_amount(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        key = (1, "Years 7 - 12 Students")
        assert key in data
        _, total = data[key]
        assert total == Decimal("16172029.00"), f"Unexpected total for {key}: {total}"

    def test_section_assigned(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        sections = {section for (_, (section, _)) in data.items()}
        assert sections - {"Unknown"}, "All sections are 'Unknown' -- section detection broken"

    def test_core_student_allocation_section(
        self, data: dict[tuple[int, str], tuple[str, Decimal]]
    ) -> None:
        key = (1, "Years 7 - 12 Students")
        section, _ = data[key]
        assert "Core Student Learning" in section, f"Unexpected section: {section!r}"

    def test_total_sum_positive(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        total = sum((v for (_, v) in data.values()), Decimal("0"))
        assert total > Decimal("0"), f"Total should be positive, got {total}"


class TestParseSrpPdfConfirmed:
    @pytest.fixture(scope="class")
    def data(self) -> dict[tuple[int, str], tuple[str, Decimal]]:
        return parse_srp_pdf(_CONFIRMED_PDF)

    def test_minimum_key_count(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        assert len(data) >= 20, f"Expected >= 20 keys, got {len(data)}"

    def test_new_lines_present(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        """Confirmed has lines not in Indicative (Quality Improvements, Language Assistants)."""
        assert (160, "Quality Improvements") in data or (
            42,
            "Language Assistants Program",
        ) in data, (
            "Expected at least one of (160, 'Quality Improvements') or "
            "(42, 'Language Assistants Program') in Confirmed"
        )

    def test_tier3_present(self, data: dict[tuple[int, str], tuple[str, Decimal]]) -> None:
        assert (138, "Tier 3 Individualised Funding") in data


# ---------------------------------------------------------------------------
# Test: parse_srp_pdf_ordered
# ---------------------------------------------------------------------------


class TestParseSrpPdfOrdered:
    def test_ordered_keys_match_dict_keys(self) -> None:
        data, ordered = parse_srp_pdf_ordered(_INDICATIVE_PDF)
        assert set(ordered) == set(data.keys()), "ordered keys must be the same set as dict keys"

    def test_ordered_keys_no_duplicates(self) -> None:
        _data, ordered = parse_srp_pdf_ordered(_INDICATIVE_PDF)
        assert len(ordered) == len(set(ordered)), "ordered keys must have no duplicates"

    def test_ordered_length_matches_dict(self) -> None:
        data, ordered = parse_srp_pdf_ordered(_INDICATIVE_PDF)
        assert len(ordered) == len(data)

    def test_first_key_has_low_ref(self) -> None:
        """The first key in the PDF should typically have a low Ref number (1 or close)."""
        _data, ordered = parse_srp_pdf_ordered(_INDICATIVE_PDF)
        assert ordered, "expected at least one key"
        first_ref = ordered[0][0]
        assert first_ref <= 10, f"Expected first ref <= 10, got {first_ref}"


# ---------------------------------------------------------------------------
# Test: compare_srp with synthetic fixtures
# ---------------------------------------------------------------------------


def _mk(
    ref: int, desc: str, section: str, amount: str
) -> tuple[tuple[int, str], tuple[str, Decimal]]:
    return (ref, desc), (section, Decimal(amount))


class TestCompareSrp:
    def _make_data(
        self,
    ) -> tuple[
        dict[tuple[int, str], tuple[str, Decimal]],
        dict[tuple[int, str], tuple[str, Decimal]],
    ]:
        indicative = dict(
            [
                _mk(1, "Item A", "Sec1", "1000.00"),
                _mk(2, "Item B", "Sec1", "2000.00"),
                _mk(3, "Item C", "Sec2", "3000.00"),
                _mk(4, "Item D", "Sec2", "4000.00"),
                _mk(5, "Item E", "Sec3", "500.00"),
            ]
        )
        confirmed = dict(
            [
                _mk(1, "Item A", "Sec1", "1000.00"),  # unchanged
                _mk(2, "Item B", "Sec1", "2500.00"),  # changed (increased net)
                _mk(3, "Item C", "Sec2", "2800.00"),  # changed (decreased net)
                # Item D removed
                _mk(6, "Item F", "Sec3", "750.00"),  # new
                _mk(5, "Item E", "Sec3", "500.00"),  # unchanged
            ]
        )
        return indicative, confirmed

    def test_categories_present(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        cats = {ln.category for ln in lines}
        assert "unchanged" in cats
        assert "changed" in cats
        assert "new" in cats
        assert "removed" in cats

    def test_unchanged_variance_zero(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        unchanged = [ln for ln in lines if ln.category == "unchanged"]
        assert unchanged
        for ln in unchanged:
            assert ln.variance == Decimal("0") or ln.variance is None

    def test_changed_increased_variance_positive(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        # Item B: 2000 -> 2500
        b_line = next(ln for ln in lines if ln.description == "Item B")
        assert b_line.category == "changed"
        assert b_line.variance is not None and b_line.variance > Decimal("0")

    def test_changed_decreased_variance_negative(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        # Item C: 3000 -> 2800
        c_line = next(ln for ln in lines if ln.description == "Item C")
        assert c_line.category == "changed"
        assert c_line.variance is not None and c_line.variance < Decimal("0")

    def test_new_no_first_version(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        new_lines = [ln for ln in lines if ln.category == "new"]
        assert new_lines
        for ln in new_lines:
            assert ln.indicative is None
            assert ln.confirmed is not None

    def test_removed_no_second_version(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        removed = [ln for ln in lines if ln.category == "removed"]
        assert removed
        for ln in removed:
            assert ln.confirmed is None
            assert ln.indicative is not None

    def test_ref_desc_join_key(self) -> None:
        """Two lines with the same Ref but different descriptions must be separate."""
        indicative = dict(
            [
                _mk(15, "Integration Level 1", "PSD", "8813.00"),
                _mk(15, "Integration Level 2", "PSD", "20382.00"),
            ]
        )
        confirmed = dict(
            [
                # Level 1 removed, Level 2 changed (increased)
                _mk(15, "Integration Level 2", "PSD", "40764.00"),
            ]
        )
        lines = compare_srp(indicative=indicative, confirmed=confirmed)
        assert len(lines) == 2
        cats = {(ln.ref, ln.description): ln.category for ln in lines}
        assert cats[(15, "Integration Level 1")] == "removed"
        assert cats[(15, "Integration Level 2")] == "changed"

    def test_pct_calculation(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        # Item B: indicative=2000, confirmed=2500, variance=500, pct=25.00%
        b_line = next(ln for ln in lines if ln.description == "Item B")
        assert b_line.pct is not None
        assert b_line.pct == Decimal("25.00"), f"Expected 25.00%, got {b_line.pct}"

    def test_total_line_count(self) -> None:
        ind, conf = self._make_data()
        lines = compare_srp(indicative=ind, confirmed=conf)
        # 5 indicative + 1 new = 6 unique (ref, desc) pairs
        assert len(lines) == 6

    def test_indicative_order_preserved(self) -> None:
        """When indicative_order is supplied, output rows follow Indicative order."""
        indicative = dict(
            [
                _mk(3, "Item C", "Sec2", "3000.00"),
                _mk(1, "Item A", "Sec1", "1000.00"),
                _mk(2, "Item B", "Sec1", "2000.00"),
            ]
        )
        confirmed = dict(
            [
                _mk(1, "Item A", "Sec1", "1000.00"),
                _mk(2, "Item B", "Sec1", "2500.00"),
                _mk(3, "Item C", "Sec2", "3200.00"),
                _mk(4, "Item D", "Sec3", "400.00"),  # new
            ]
        )
        ind_order = [(3, "Item C"), (1, "Item A"), (2, "Item B")]
        lines = compare_srp(indicative=indicative, confirmed=confirmed, indicative_order=ind_order)
        assert lines[0].ref == 3
        assert lines[1].ref == 1
        assert lines[2].ref == 2
        assert lines[3].ref == 4

    def test_three_way_compare_unchanged_and_changed(self) -> None:
        """3-way comparison uses changed / unchanged categories."""
        indicative = dict([_mk(1, "Alpha", "S1", "1000.00"), _mk(2, "Beta", "S1", "2000.00")])
        confirmed = dict([_mk(1, "Alpha", "S1", "1500.00"), _mk(2, "Beta", "S1", "2000.00")])
        revised1 = dict(
            [
                ((1, "Alpha"), ("S1", Decimal("1500.00"))),  # same as confirmed
                ((2, "Beta"), ("S1", Decimal("2300.00"))),  # changed at step 2
            ]
        )
        lines = compare_srp(indicative=indicative, confirmed=confirmed, revised1=revised1)
        by_desc = {ln.description: ln for ln in lines}
        # Alpha changed 1000->1500 at step1, unchanged step2 -> net "changed"
        assert by_desc["Alpha"].category == "changed"
        # Beta unchanged at step1, changed 2000->2300 at step2 -> net "changed"
        assert by_desc["Beta"].category == "changed"

    def test_revised1_none_fields_without_revised1(self) -> None:
        """Without revised1, revised1/variance_conf_to_rev1 fields are None."""
        indicative = dict([_mk(1, "A", "S1", "100.00")])
        confirmed = dict([_mk(1, "A", "S1", "100.00")])
        lines = compare_srp(indicative=indicative, confirmed=confirmed)
        assert lines[0].revised1 is None
        assert lines[0].revised2 is None
        assert lines[0].variance_conf_to_rev1 is None
        assert lines[0].variance_rev1_to_rev2 is None

    def test_generic_versions_interface_two_versions(self) -> None:
        """compare_srp(versions=...) accepts any 2 slot keys."""
        conf_data = dict([_mk(1, "A", "S1", "1000.00"), _mk(2, "B", "S1", "2000.00")])
        rev1_data = dict([_mk(1, "A", "S1", "1200.00"), _mk(2, "B", "S1", "2000.00")])
        lines = compare_srp(versions=[("confirmed", conf_data), ("revised1", rev1_data)])
        assert len(lines) == 2
        by_desc = {ln.description: ln for ln in lines}
        assert by_desc["A"].category == "changed"
        assert by_desc["B"].category == "unchanged"

    def test_generic_versions_interface_raises_on_one(self) -> None:
        conf_data = dict([_mk(1, "A", "S1", "1000.00")])
        with pytest.raises(ValueError, match="At least 2"):
            compare_srp(versions=[("confirmed", conf_data)])

    def test_generic_versions_interface_raises_on_zero(self) -> None:
        with pytest.raises(ValueError, match="At least 2"):
            compare_srp(versions=[])

    def test_adjacent_variances_populated(self) -> None:
        """adjacent_variances list has one entry per adjacent pair."""
        indicative = dict([_mk(1, "A", "S1", "1000.00")])
        confirmed = dict([_mk(1, "A", "S1", "1200.00")])
        lines = compare_srp(indicative=indicative, confirmed=confirmed)
        assert len(lines[0].adjacent_variances) == 1
        _lbl, var = lines[0].adjacent_variances[0]
        assert var == Decimal("200.00")

    def test_three_way_adjacent_variances(self) -> None:
        """3-way compare has 2 adjacent_variances entries."""
        ind = dict([_mk(1, "A", "S1", "1000.00")])
        conf = dict([_mk(1, "A", "S1", "1100.00")])
        rev1 = dict([_mk(1, "A", "S1", "1300.00")])
        lines = compare_srp(indicative=ind, confirmed=conf, revised1=rev1)
        assert len(lines[0].adjacent_variances) == 2
        _, v1 = lines[0].adjacent_variances[0]
        _, v2 = lines[0].adjacent_variances[1]
        assert v1 == Decimal("100.00")
        assert v2 == Decimal("200.00")


# ---------------------------------------------------------------------------
# Test: any-2-of-4 via compare_srp (synthetic)
# ---------------------------------------------------------------------------


class TestCompareSrpAnyTwoVersions:
    def test_confirmed_and_revised1_only(self) -> None:
        conf = dict([_mk(1, "A", "S1", "1000.00"), _mk(2, "B", "S1", "2000.00")])
        rev1 = dict([_mk(1, "A", "S1", "1200.00"), _mk(3, "C", "S1", "500.00")])
        lines = compare_srp(confirmed=conf, revised1=rev1)
        cats = {ln.description: ln.category for ln in lines}
        assert cats["A"] == "changed"
        assert cats["B"] == "removed"  # in conf (first ver) only
        assert cats["C"] == "new"  # in rev1 (second ver) only

    def test_revised1_and_revised2_only(self) -> None:
        rev1 = dict([_mk(1, "X", "S1", "500.00")])
        rev2 = dict([_mk(1, "X", "S1", "500.00"), _mk(2, "Y", "S1", "300.00")])
        lines = compare_srp(revised1=rev1, revised2=rev2)
        cats = {ln.description: ln.category for ln in lines}
        assert cats["X"] == "unchanged"
        assert cats["Y"] == "new"

    def test_indicative_and_revised2_only(self) -> None:
        ind = dict([_mk(1, "P", "S1", "1000.00"), _mk(2, "Q", "S1", "200.00")])
        rev2 = dict([_mk(1, "P", "S1", "1000.00"), _mk(3, "R", "S1", "100.00")])
        lines = compare_srp(indicative=ind, revised2=rev2)
        cats = {ln.description: ln.category for ln in lines}
        assert cats["P"] == "unchanged"
        assert cats["Q"] == "removed"
        assert cats["R"] == "new"

    def test_ind_rev1_rev2_skip_confirmed(self) -> None:
        """3 versions: Indicative + Rev1 + Rev2 (skipping Confirmed)."""
        ind = dict([_mk(1, "A", "S1", "1000.00")])
        rev1 = dict([_mk(1, "A", "S1", "1100.00")])
        rev2 = dict([_mk(1, "A", "S1", "1300.00")])
        lines = compare_srp(indicative=ind, revised1=rev1, revised2=rev2)
        assert len(lines) == 1
        assert lines[0].category == "changed"
        assert lines[0].variance == Decimal("300.00")
        # 3 slots -> 2 adjacent pairs
        assert len(lines[0].adjacent_variances) == 2

    def test_version_labels_reflect_provided_slots(self) -> None:
        """SrpLine.adjacent_variances labels must name the provided versions."""
        conf = dict([_mk(1, "A", "S1", "1000.00")])
        rev2 = dict([_mk(1, "A", "S1", "1200.00")])
        lines = compare_srp(confirmed=conf, revised2=rev2)
        lbl, _ = lines[0].adjacent_variances[0]
        assert "Confirmed" in lbl
        assert "Previous Year Revised" in lbl

    def test_one_version_raises(self) -> None:
        ind = dict([_mk(1, "A", "S1", "1000.00")])
        with pytest.raises(ValueError, match="At least 2"):
            compare_srp(indicative=ind)

    def test_zero_versions_raises(self) -> None:
        with pytest.raises(ValueError, match="At least 2"):
            compare_srp()

    def test_first_provided_version_sets_row_order(self) -> None:
        """Row order comes from the FIRST provided version, regardless of slot."""
        conf = dict(
            [
                _mk(3, "C", "S", "300.00"),
                _mk(1, "A", "S", "100.00"),
                _mk(2, "B", "S", "200.00"),
            ]
        )
        rev1 = dict([_mk(1, "A", "S", "150.00"), _mk(2, "B", "S", "200.00")])
        lines = compare_srp(confirmed=conf, revised1=rev1)
        # Order should follow conf (C, A, B -- insertion order in dict)
        assert lines[0].description == "C"
        assert lines[1].description == "A"
        assert lines[2].description == "B"


# ---------------------------------------------------------------------------
# Test: generate_srp_comparison against real sample PDFs
# ---------------------------------------------------------------------------


class TestGenerateSrpComparison:
    def test_round_trip_produces_xlsx(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_out.xlsx"
        progress_calls: list[tuple[int, str]] = []

        def progress(pct: int, msg: str) -> None:
            progress_calls.append((pct, msg))

        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=progress,
        )

        assert output.exists(), "Output XLSX was not created"
        assert isinstance(summary, SrpSummary)
        assert len(summary.lines) >= 20

    def test_all_four_categories_present(self, tmp_path: Path) -> None:
        """At least one line in each of the 4 categories must exist in the real diff."""
        output = tmp_path / "srp_cat.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        cats = {ln.category for ln in summary.lines}
        assert "unchanged" in cats, f"Expected 'unchanged' in {cats}"
        assert "changed" in cats, f"Expected 'changed' in {cats}"
        assert "new" in cats, f"Expected 'new' in {cats}"
        assert "removed" in cats, f"Expected 'removed' in {cats}"

    def test_version_labels_two_versions(self, tmp_path: Path) -> None:
        """Without revised PDFs, version_labels has 2 entries."""
        output = tmp_path / "srp_rev.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        assert summary.version_labels == ["Indicative", "Confirmed"]
        assert not summary.has_revised1
        assert not summary.has_revised2

    def test_indicative_order_preserved_in_summary(self, tmp_path: Path) -> None:
        """Lines in summary must start in Indicative PDF order."""
        output = tmp_path / "srp_order.xlsx"
        from tools.srp.logic import parse_srp_pdf_ordered

        _, ind_order = parse_srp_pdf_ordered(_INDICATIVE_PDF)
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        ind_keys_in_summary = [
            (ln.ref, ln.description) for ln in summary.lines if ln.indicative is not None
        ]
        ind_keys_in_order = [k for k in ind_order if k in set(ind_keys_in_summary)]
        assert ind_keys_in_summary == ind_keys_in_order, (
            "Summary lines do not follow Indicative PDF source order"
        )

    def test_xlsx_header_correct(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_hdr.xlsx"
        generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        header_vals = [ws.cell(1, c).value for c in range(1, (ws.max_column or 1) + 1)]  # type: ignore[union-attr]
        assert "Ref" in header_vals
        assert "Description" in header_vals
        assert "Indicative" in header_vals
        assert "Confirmed" in header_vals

    def test_xlsx_row_count_matches_lines(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_rows.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        data_rows = (ws.max_row or 1) - 1  # type: ignore[union-attr]
        assert data_rows == len(summary.lines)

    def test_mismatch_fill_on_removed_rows(self, tmp_path: Path) -> None:
        """Rows with category 'removed' must carry HL_MISMATCH fill."""
        output = tmp_path / "srp_fill_dec.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )

        rem_lines = [ln for ln in summary.lines if ln.category == "removed"]
        assert rem_lines, (
            "No 'removed' lines found -- the sample diff must produce at least one "
            "such line for this test to be meaningful."
        )

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        target_argb = ("FF" + HL_MISMATCH).upper()

        for ln in rem_lines[:3]:
            for row_idx in range(2, (ws.max_row or 2) + 1):  # type: ignore[union-attr]
                ref_val = ws.cell(row_idx, 1).value  # type: ignore[union-attr]
                desc_val = ws.cell(row_idx, 3).value  # type: ignore[union-attr]
                if str(ref_val) == str(ln.ref) and str(desc_val) == ln.description:
                    fill = ws.cell(row_idx, 1).fill  # type: ignore[union-attr]
                    assert fill and fill.fgColor and fill.fgColor.rgb, (
                        f"Removed row ({ln.ref}, {ln.description!r}) has no fill"
                    )
                    assert target_argb in fill.fgColor.rgb.upper(), (
                        f"Removed row ({ln.ref}, {ln.description!r}) has wrong fill "
                        f"{fill.fgColor.rgb!r}; expected {target_argb!r}"
                    )
                    break

    def test_source_fill_on_new_rows(self, tmp_path: Path) -> None:
        """Rows with category 'new' must carry HL_SOURCE_ONLY fill."""
        output = tmp_path / "srp_fill_inc.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )

        new_lines = [ln for ln in summary.lines if ln.category == "new"]
        assert new_lines, (
            "No 'new' lines found -- the sample diff must produce at least one "
            "such line for this test to be meaningful."
        )

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        target_argb = ("FF" + HL_SOURCE_ONLY).upper()

        for ln in new_lines[:3]:
            for row_idx in range(2, (ws.max_row or 2) + 1):  # type: ignore[union-attr]
                ref_val = ws.cell(row_idx, 1).value  # type: ignore[union-attr]
                desc_val = ws.cell(row_idx, 3).value  # type: ignore[union-attr]
                if str(ref_val) == str(ln.ref) and str(desc_val) == ln.description:
                    fill = ws.cell(row_idx, 1).fill  # type: ignore[union-attr]
                    assert fill and fill.fgColor and fill.fgColor.rgb, (
                        f"New row ({ln.ref}, {ln.description!r}) has no fill"
                    )
                    assert target_argb in fill.fgColor.rgb.upper(), (
                        f"New row ({ln.ref}, {ln.description!r}) has wrong fill "
                        f"{fill.fgColor.rgb!r}; expected {target_argb!r}"
                    )
                    break

    def test_unchanged_row_has_no_fill(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_fill_unch.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )

        unchanged = [ln for ln in summary.lines if ln.category == "unchanged"]
        assert unchanged, "Expected at least one unchanged line"

        wb = openpyxl.load_workbook(output)
        ws = wb.active

        for ln in unchanged[:3]:
            for row_idx in range(2, (ws.max_row or 2) + 1):  # type: ignore[union-attr]
                ref_val = ws.cell(row_idx, 1).value  # type: ignore[union-attr]
                desc_val = ws.cell(row_idx, 3).value  # type: ignore[union-attr]
                if str(ref_val) == str(ln.ref) and str(desc_val) == ln.description:
                    fill = ws.cell(row_idx, 1).fill  # type: ignore[union-attr]
                    no_fill = (
                        fill is None
                        or fill.fill_type in (None, "none")
                        or (fill.fgColor is not None and fill.fgColor.type == "none")
                    )
                    assert no_fill, (
                        f"Unchanged row ({ln.ref}, {ln.description!r}) should have no fill"
                    )
                    break

    def test_progress_callbacks_fired(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_prog.xlsx"
        pcts: list[int] = []

        def progress(pct: int, msg: str) -> None:
            pcts.append(pct)

        generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=progress,
        )

        assert 10 in pcts
        assert 100 in pcts

    def test_totals_computed(self, tmp_path: Path) -> None:
        output = tmp_path / "srp_totals.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        assert summary.total_first > Decimal("0")
        assert summary.total_last > Decimal("0")
        # Legacy aliases
        assert summary.total_indicative == summary.total_first
        assert summary.total_confirmed == summary.total_last

    def test_missing_indicative_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="not found"):
            generate_srp_comparison(
                indicative_pdf=tmp_path / "missing.pdf",
                confirmed_pdf=_CONFIRMED_PDF,
                output_file=tmp_path / "out.xlsx",
                progress=lambda p, m: None,
            )

    def test_one_pdf_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="At least 2"):
            generate_srp_comparison(
                indicative_pdf=_INDICATIVE_PDF,
                output_file=tmp_path / "out.xlsx",
                progress=lambda p, m: None,
            )

    def test_zero_pdfs_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="At least 2"):
            generate_srp_comparison(
                output_file=tmp_path / "out.xlsx",
                progress=lambda p, m: None,
            )

    def test_ind_and_confirmed_version_labels(self, tmp_path: Path) -> None:
        """Ind + Conf -> version_labels == [Indicative, Confirmed]."""
        output = tmp_path / "out.xlsx"
        summary = generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        assert summary.version_labels == ["Indicative", "Confirmed"]

    def test_xlsx_two_versions_column_count(self, tmp_path: Path) -> None:
        """2 versions -> 8 columns: Ref+Sec+Desc+V1+V2+Var+%+Cat."""
        output = tmp_path / "out.xlsx"
        generate_srp_comparison(
            indicative_pdf=_INDICATIVE_PDF,
            confirmed_pdf=_CONFIRMED_PDF,
            output_file=output,
            progress=lambda p, m: None,
        )
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws is not None
        assert ws.max_column == 8
