from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from toolkit.base_tool import ProgressFn
from toolkit.fills import argb
from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY

# ---------------------------------------------------------------------------
# Currency parsing
# ---------------------------------------------------------------------------

# Bare hyphen, en-dash, em-dash (alone) -> zero
_DASH_RE = re.compile(r"^[—–\-]{1,2}$")


def _parse_opstat_decimal(raw: str) -> Decimal:
    """Convert a GL21150 numeric token to Decimal.

    Extends the sub_program parser to handle bare ``-`` (hyphen) as zero,
    negative values written as ``-987,012``, and all existing cases:
    commas, parentheses-negative, blank.
    """
    text = raw.strip() if raw else ""

    if not text or _DASH_RE.match(text):
        return Decimal("0")

    # parentheses -> negative
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    # strip leading $
    text = text.lstrip("$").strip()

    # remove thousands separators
    text = text.replace(",", "")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse {raw!r} as a decimal: {exc}") from exc


# ---------------------------------------------------------------------------
# PDF structure constants
# ---------------------------------------------------------------------------

# GL code prefix for 5-digit data rows
_DATA_ROW_RE = re.compile(r"^(\d{5})\s+(.*)")

# Section header lines (exact, after strip)
_SECTION_NAMES = frozenset(
    [
        "REVENUE",
        "EXPENDITURE",
        "CAPITAL EXPENDITURE",
        "ASSET WRITE-DOWNS",
        "PRIOR YEAR ADJUSTMENTS",
    ]
)

# Lines to skip outright (not section headers, not data rows)
_SKIP_RE = re.compile(
    r"^("
    r"\d{4,}:\w"  # school header e.g. "8819:Melbourne"
    r"|General Ledger"
    r"|Operating Statement"
    r"|for the period"
    r"|Current Month"
    r"|GL Code Account Title"
    r"|Total Operating Revenue"
    r"|Total Operating Expenditure"
    r"|Total Capital Expenditure"
    r"|Total Asset Write-Downs"
    r"|Total Prior Year Adjustments"
    r"|Net Operating Surplus"
    r"|Outstanding Orders"
    r"|We certify"
    r"|School Principal"
    r"|School Council"
    r"|\d+ \w+ \d{4} \d+:\d+"  # date+time footer e.g. "3 March 2026 9:20"
    r"|\d+ \[GL\d+"  # page footer e.g. "1 [GL21150]"
    r"|Page \d+ of \d+"
    r")",
    re.IGNORECASE,
)

# Period label pattern
_PERIOD_RE = re.compile(r"for the period ending\s+(\d+\s+\w+\s+\d{4})", re.IGNORECASE)

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class OpStatRow:
    """A single GL account row extracted from one PDF."""

    gl_code: int
    description: str
    section: str
    subsection: str
    ytd_actual: Decimal


@dataclass(frozen=True)
class OpStatLine:
    """A diff line produced by comparing prior vs current periods."""

    gl_code: int
    description: str
    section: str
    subsection: str
    ytd_prior: Decimal | None
    ytd_current: Decimal | None
    movement: Decimal
    pct: Decimal | None
    is_favourable: bool | None
    exceeds_threshold: bool


@dataclass(frozen=True)
class OpStatSummary:
    """Result of a full comparison run."""

    lines: list[OpStatLine]
    revenue_movement: Decimal
    expenditure_movement: Decimal
    operating_result_movement: Decimal
    period_current: str
    period_prior: str
    output_path: Path


# ---------------------------------------------------------------------------
# PDF parser
# ---------------------------------------------------------------------------

# Sections where positive movement is adverse (costs)
_COST_SECTIONS = frozenset(
    [
        "EXPENDITURE",
        "CAPITAL EXPENDITURE",
        "ASSET WRITE-DOWNS",
        "PRIOR YEAR ADJUSTMENTS",
    ]
)


def _is_subsection_candidate(line: str) -> bool:
    """Return True if *line* looks like a sub-section header.

    Sub-section headers are non-empty, do not start with a digit (so they are
    not data rows or sub-total rows), are not all-caps section names, and are
    not in the skip list.
    """
    if not line:
        return False
    if _DATA_ROW_RE.match(line):
        return False
    if line in _SECTION_NAMES:
        return False
    if _SKIP_RE.match(line):
        return False
    # Sub-total rows start with a digit followed by comma/space (e.g. "16,622")
    if re.match(r"^\d", line):
        return False
    return True


def _parse_data_row(
    gl_code_str: str,
    rest: str,
    section: str,
    subsection: str,
) -> OpStatRow | None:
    """Extract an OpStatRow from the right-hand side of a data line.

    The rest string contains: description + 9 space-separated numeric tokens.
    Tokens: CM Actual | CM Budget | CM Variance | YTD Actual | YTD Budget |
            YTD Variance | Annual Budget | % | Last Year Actual.

    Returns None if fewer than 4 numeric tokens are found (malformed row).
    """
    parts = rest.split()

    # Split description from numeric tokens.
    # Scan left-to-right; the first token that matches a standalone numeric
    # (pure digits/commas, optional leading minus; or bare "-") is the
    # boundary between description text and the numeric block.
    desc_parts: list[str] = []
    num_parts: list[str] = []
    found_num = False
    for part in parts:
        if not found_num:
            # A standalone numeric: optional minus, then digits with optional commas.
            # Bare "-" is also numeric. Part like "<$5,000" is NOT standalone numeric.
            if re.match(r"^-?\d[\d,]*(\.\d+)?$|^-$", part):
                found_num = True
                num_parts.append(part)
            else:
                desc_parts.append(part)
        else:
            num_parts.append(part)

    if len(num_parts) < 4:
        return None

    description = " ".join(desc_parts)
    ytd_actual = _parse_opstat_decimal(num_parts[3])

    return OpStatRow(
        gl_code=int(gl_code_str),
        description=description,
        section=section,
        subsection=subsection,
        ytd_actual=ytd_actual,
    )


def parse_opstat_pdf(pdf_path: Path) -> tuple[list[OpStatRow], str]:
    """Parse a CASES21 GL21150 Operating Statement PDF.

    Returns ``(rows, period_label)`` where *period_label* is extracted from
    the "for the period ending ..." header line, e.g. ``"28 February 2026"``.
    Raises :class:`ValueError` if the file is absent or unrecognisable.
    """
    if not pdf_path.exists():
        raise ValueError(f"File not found: {pdf_path}")

    rows: list[OpStatRow] = []
    period_label = ""
    section = "REVENUE"
    subsection = ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                raise ValueError(f"Operating Statement PDF appears empty: {pdf_path}")

            for page in pdf.pages:
                text = page.extract_text() or ""

                for raw_line in text.splitlines():
                    line = raw_line.strip()
                    if not line:
                        continue

                    # Extract period label (grab on first match)
                    if not period_label:
                        pm = _PERIOD_RE.match(line)
                        if pm:
                            period_label = pm.group(1)
                            continue

                    # Section header?
                    if line in _SECTION_NAMES:
                        section = line
                        subsection = ""
                        continue

                    # Skip line?
                    if _SKIP_RE.match(line):
                        continue

                    # Data row?
                    dm = _DATA_ROW_RE.match(line)
                    if dm:
                        row = _parse_data_row(dm.group(1), dm.group(2), section, subsection)
                        if row is not None:
                            rows.append(row)
                        continue

                    # Sub-section candidate -- update context
                    if _is_subsection_candidate(line):
                        subsection = line

    except OSError as exc:
        raise ValueError(f"Cannot read Operating Statement PDF: {pdf_path}") from exc

    if not rows:
        raise ValueError(f"No data rows found in Operating Statement PDF: {pdf_path}")

    return rows, period_label


# ---------------------------------------------------------------------------
# Diff / compare
# ---------------------------------------------------------------------------

_ZERO = Decimal("0")


def compare_opstat(
    prior: list[OpStatRow],
    current: list[OpStatRow],
    threshold_dollars: Decimal,
    threshold_pct: int,
) -> list[OpStatLine]:
    """Diff two lists of OpStatRow objects.

    Returns a list of :class:`OpStatLine` objects -- one per unique GL code in
    the union of prior and current -- sorted by (section_order, gl_code).
    """
    # Build lookup dicts keyed by gl_code.  Last occurrence wins if duplicates.
    prior_map: dict[int, OpStatRow] = {r.gl_code: r for r in prior}
    current_map: dict[int, OpStatRow] = {r.gl_code: r for r in current}

    # Canonical section order for sorting
    _section_order = {
        "REVENUE": 0,
        "EXPENDITURE": 1,
        "CAPITAL EXPENDITURE": 2,
        "ASSET WRITE-DOWNS": 3,
        "PRIOR YEAR ADJUSTMENTS": 4,
    }

    all_codes = sorted(
        prior_map.keys() | current_map.keys(),
        key=lambda c: (
            _section_order.get((current_map.get(c) or prior_map[c]).section, 9),
            c,
        ),
    )

    lines: list[OpStatLine] = []
    for code in all_codes:
        p_row = prior_map.get(code)
        c_row = current_map.get(code)

        # Use whichever row has more context for description/section/subsection
        ref_row: OpStatRow = c_row if c_row is not None else p_row  # type: ignore[assignment]

        ytd_prior: Decimal | None = p_row.ytd_actual if p_row is not None else None
        ytd_current: Decimal | None = c_row.ytd_actual if c_row is not None else None

        prior_val = ytd_prior if ytd_prior is not None else _ZERO
        current_val = ytd_current if ytd_current is not None else _ZERO

        movement = current_val - prior_val

        pct: Decimal | None
        if prior_val != _ZERO:
            pct = movement / prior_val * Decimal("100")
        else:
            pct = None

        # Determine favourability
        is_favourable: bool | None
        if movement == _ZERO:
            is_favourable = None
        elif ref_row.section in _COST_SECTIONS:
            # Lower costs = favourable
            is_favourable = movement < _ZERO
        else:
            # Higher revenue = favourable
            is_favourable = movement > _ZERO

        # Threshold check
        abs_movement = abs(movement)
        if pct is not None:
            abs_pct = abs(pct)
            exceeds_threshold = bool(
                abs_movement >= threshold_dollars or abs_pct >= Decimal(str(threshold_pct))
            )
        else:
            exceeds_threshold = bool(abs_movement >= threshold_dollars)

        lines.append(
            OpStatLine(
                gl_code=code,
                description=ref_row.description,
                section=ref_row.section,
                subsection=ref_row.subsection,
                ytd_prior=ytd_prior,
                ytd_current=ytd_current,
                movement=movement,
                pct=pct,
                is_favourable=is_favourable,
                exceeds_threshold=exceeds_threshold,
            )
        )

    return lines


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

_HEADERS = [
    "Account",
    "Description",
    "Section",
    "Sub-section",
    "YTD Prior",
    "YTD Current",
    "Movement",
    "%",
    "Favourable?",
]

_COL_WIDTHS = [10, 42, 24, 28, 14, 14, 14, 10, 12]

_FAVOURABLE_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_SOURCE_ONLY))
_ADVERSE_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_MISMATCH))

_MONO_FONT_NAME = "Cascadia Mono"
_MONO_FALLBACK = "Consolas"

_NUMERIC_COLS = {5, 6, 7, 8}  # 1-indexed: YTD Prior, YTD Current, Movement, %


def _xlsx_dollar(value: Decimal | None) -> float | str:
    """Return a float for numeric columns so openpyxl stores them as numbers."""
    if value is None:
        return ""
    return float(value)


def _write_xlsx(
    lines: list[OpStatLine],
    output_file: Path,
    period_prior: str,
    period_current: str,
) -> None:
    """Write comparison data to *output_file*."""
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    wb = Workbook()
    active = wb.active
    assert isinstance(active, Worksheet)
    ws: Worksheet = active
    ws.title = "OS Comparison"

    # Header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Sub-header row with period labels in the YTD columns
    ws.cell(row=2, column=5, value=f"({period_prior})")
    ws.cell(row=2, column=6, value=f"({period_current})")
    for col_idx in range(1, len(_HEADERS) + 1):
        ws.cell(row=2, column=col_idx).font = Font(italic=True, size=9)

    ws.freeze_panes = "A3"

    # Data rows (start at row 3)
    for row_idx, line in enumerate(lines, start=3):
        fav_str: Any = (
            "Yes" if line.is_favourable is True else ("No" if line.is_favourable is False else "")
        )
        pct_val: Any = float(line.pct) if line.pct is not None else ""

        values: list[Any] = [
            str(line.gl_code).zfill(5),
            line.description,
            line.section,
            line.subsection,
            _xlsx_dollar(line.ytd_prior),
            _xlsx_dollar(line.ytd_current),
            float(line.movement),
            pct_val,
            fav_str,
        ]

        # Determine fill
        if line.exceeds_threshold and line.is_favourable is True:
            fill: PatternFill | None = _FAVOURABLE_FILL
        elif line.exceeds_threshold and line.is_favourable is False:
            fill = _ADVERSE_FILL
        else:
            fill = None

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if fill is not None:
                cell.fill = fill
            if col_idx in _NUMERIC_COLS:
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(name=_MONO_FONT_NAME)
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"

    # Column widths
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def generate_opstat_comparison(
    current_file: Path,
    prior_file: Path,
    output_file: Path,
    threshold_dollars: Decimal,
    threshold_pct: int,
    progress: ProgressFn,
) -> OpStatSummary:
    """Orchestrate parse + diff + XLSX write.

    Parameters
    ----------
    current_file:
        CASES21 GL21150 PDF for the current period.
    prior_file:
        CASES21 GL21150 PDF for the prior period.
    output_file:
        Destination ``.xlsx`` path.
    threshold_dollars:
        Minimum absolute movement in dollars to warrant highlighting.
    threshold_pct:
        Minimum absolute percentage change to warrant highlighting.
    progress:
        Callback ``(percent: int, message: str) -> None``.
    """
    progress(10, "Reading current period PDF…")
    current_rows, period_current = parse_opstat_pdf(current_file)

    progress(30, "Reading prior period PDF…")
    prior_rows, period_prior = parse_opstat_pdf(prior_file)

    progress(50, "Comparing periods…")
    lines = compare_opstat(
        prior=prior_rows,
        current=current_rows,
        threshold_dollars=threshold_dollars,
        threshold_pct=threshold_pct,
    )

    progress(70, "Writing workbook…")
    _write_xlsx(
        lines=lines,
        output_file=output_file,
        period_prior=period_prior,
        period_current=period_current,
    )

    # Compute section-level movement summaries
    _zero = Decimal("0")

    revenue_movement: Decimal = sum((ln.movement for ln in lines if ln.section == "REVENUE"), _zero)
    expenditure_movement: Decimal = sum(
        (ln.movement for ln in lines if ln.section == "EXPENDITURE"), _zero
    )
    operating_result_movement: Decimal = revenue_movement - expenditure_movement

    progress(100, "Done.")

    return OpStatSummary(
        lines=lines,
        revenue_movement=revenue_movement,
        expenditure_movement=expenditure_movement,
        operating_result_movement=operating_result_movement,
        period_current=period_current,
        period_prior=period_prior,
        output_path=output_file,
    )
