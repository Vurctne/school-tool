"""Tests for tools/operating/logic.py — parser, diff, and XLSX writer."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY
from tools.operating.logic import (
    OpStatRow,
    OpStatSummary,
    _parse_opstat_decimal,
    compare_opstat,
    generate_opstat_comparison,
    parse_opstat_pdf,
)

# ---------------------------------------------------------------------------
# Sample PDF paths
# ---------------------------------------------------------------------------

_PRIOR_PDF = Path("Samples/Operating Statement/GL21150_Operating Statement Detailed.pdf")
_CURRENT_PDF = Path("Samples/Operating Statement/GL21150_Operating Statement Detailed 2.pdf")

# ---------------------------------------------------------------------------
# 1. Decimal parser
# ---------------------------------------------------------------------------


class TestParseOpstatDecimal:
    def test_bare_hyphen_is_zero(self) -> None:
        assert _parse_opstat_decimal("-") == Decimal("0")

    def test_em_dash_is_zero(self) -> None:
        assert _parse_opstat_decimal("—") == Decimal("0")

    def test_en_dash_is_zero(self) -> None:
        assert _parse_opstat_decimal("–") == Decimal("0")

    def test_blank_is_zero(self) -> None:
        assert _parse_opstat_decimal("") == Decimal("0")

    def test_plain_number(self) -> None:
        assert _parse_opstat_decimal("325,979") == Decimal("325979")

    def test_negative_with_minus(self) -> None:
        assert _parse_opstat_decimal("-987,012") == Decimal("-987012")

    def test_negative_integer(self) -> None:
        assert _parse_opstat_decimal("-55") == Decimal("-55")

    def test_commas_stripped(self) -> None:
        assert _parse_opstat_decimal("1,312,991") == Decimal("1312991")

    def test_parentheses_negative(self) -> None:
        assert _parse_opstat_decimal("(500)") == Decimal("-500")

    def test_dollar_prefix_stripped(self) -> None:
        assert _parse_opstat_decimal("$1,234.56") == Decimal("1234.56")

    def test_integer_string(self) -> None:
        assert _parse_opstat_decimal("42") == Decimal("42")

    def test_invalid_raises(self) -> None:
        with pytest.raises(ValueError):
            _parse_opstat_decimal("not-a-number!!")


# ---------------------------------------------------------------------------
# 2. parse_opstat_pdf — real PDF fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def prior_rows() -> list[OpStatRow]:
    return parse_opstat_pdf(_PRIOR_PDF)[0]


@pytest.fixture(scope="module")
def current_rows() -> list[OpStatRow]:
    return parse_opstat_pdf(_CURRENT_PDF)[0]


class TestParsePriorPdf:
    def test_minimum_row_count(self, prior_rows: list[OpStatRow]) -> None:
        assert len(prior_rows) >= 50, f"Expected >= 50 rows, got {len(prior_rows)}"

    def test_period_label_extracted(self) -> None:
        _, period = parse_opstat_pdf(_PRIOR_PDF)
        assert period == "28 February 2026", f"Got {period!r}"

    def test_has_revenue_section(self, prior_rows: list[OpStatRow]) -> None:
        revenue_rows = [r for r in prior_rows if r.section == "REVENUE"]
        assert len(revenue_rows) > 0, "No REVENUE rows found"

    def test_has_expenditure_section(self, prior_rows: list[OpStatRow]) -> None:
        exp_rows = [r for r in prior_rows if r.section == "EXPENDITURE"]
        assert len(exp_rows) > 0, "No EXPENDITURE rows found"

    def test_has_capital_expenditure_section(self, prior_rows: list[OpStatRow]) -> None:
        cap_rows = [r for r in prior_rows if r.section == "CAPITAL EXPENDITURE"]
        assert len(cap_rows) > 0, "No CAPITAL EXPENDITURE rows found"

    def test_known_gl_code_present(self, prior_rows: list[OpStatRow]) -> None:
        # GL 70001 is Cash SRP Funding — present in both PDFs
        codes = {r.gl_code for r in prior_rows}
        assert 70001 in codes, "GL 70001 not found in prior PDF"

    def test_known_gl_ytd_value(self, prior_rows: list[OpStatRow]) -> None:
        # GL 70001 YTD Actual = 325,979 in the prior PDF
        row = next(r for r in prior_rows if r.gl_code == 70001)
        assert row.ytd_actual == Decimal("325979")

    def test_subsection_assigned(self, prior_rows: list[OpStatRow]) -> None:
        # GL 70001 is under "Dep't Grants" sub-section
        row = next(r for r in prior_rows if r.gl_code == 70001)
        assert "grants" in row.subsection.lower() or "dep" in row.subsection.lower(), (
            f"Unexpected subsection {row.subsection!r} for GL 70001"
        )

    def test_no_total_rows_included(self, prior_rows: list[OpStatRow]) -> None:
        desc_lower = {r.description.lower() for r in prior_rows}
        for bad in (
            "total operating revenue",
            "net operating surplus",
            "we certify",
            "outstanding orders",
        ):
            assert not any(bad in d for d in desc_lower), f"Excluded description found: {bad!r}"

    def test_all_gl_codes_are_5_digit(self, prior_rows: list[OpStatRow]) -> None:
        for row in prior_rows:
            assert 10000 <= row.gl_code <= 99999, f"GL code {row.gl_code} is not 5-digit"

    def test_ytd_actuals_are_decimal(self, prior_rows: list[OpStatRow]) -> None:
        for row in prior_rows:
            assert isinstance(row.ytd_actual, Decimal)


class TestParseCurrentPdf:
    def test_period_label_extracted(self) -> None:
        _, period = parse_opstat_pdf(_CURRENT_PDF)
        assert period == "31 March 2026", f"Got {period!r}"

    def test_minimum_row_count(self, current_rows: list[OpStatRow]) -> None:
        assert len(current_rows) >= 50

    def test_known_gl_ytd_value(self, current_rows: list[OpStatRow]) -> None:
        # GL 73002 Interest Received: YTD Actual = 83,895 in the current PDF
        row = next((r for r in current_rows if r.gl_code == 73002), None)
        assert row is not None, "GL 73002 not found in current PDF"
        assert row.ytd_actual == Decimal("83895")


# ---------------------------------------------------------------------------
# 3. compare_opstat — synthetic fixtures (no PDF needed)
# ---------------------------------------------------------------------------


def _make_row(
    gl_code: int,
    description: str = "Test",
    section: str = "REVENUE",
    subsection: str = "Grants",
    ytd_actual: str = "1000",
) -> OpStatRow:
    return OpStatRow(
        gl_code=gl_code,
        description=description,
        section=section,
        subsection=subsection,
        ytd_actual=Decimal(ytd_actual),
    )


class TestCompareOpstat:
    def test_both_periods_movement(self) -> None:
        prior = [_make_row(70001, ytd_actual="100000")]
        current = [_make_row(70001, ytd_actual="110000")]
        lines = compare_opstat(prior, current, Decimal("5000"), 10)
        assert len(lines) == 1
        ln = lines[0]
        assert ln.movement == Decimal("10000")
        assert ln.pct is not None
        assert abs(ln.pct - Decimal("10")) < Decimal("0.01")

    def test_current_only_row(self) -> None:
        prior: list[OpStatRow] = []
        current = [_make_row(89999, ytd_actual="20000")]
        lines = compare_opstat(prior, current, Decimal("5000"), 10)
        assert len(lines) == 1
        ln = lines[0]
        assert ln.ytd_prior is None
        assert ln.movement == Decimal("20000")
        assert ln.pct is None

    def test_prior_only_row(self) -> None:
        prior = [_make_row(70001, ytd_actual="30000")]
        current: list[OpStatRow] = []
        lines = compare_opstat(prior, current, Decimal("5000"), 10)
        assert len(lines) == 1
        ln = lines[0]
        assert ln.ytd_current is None
        assert ln.movement == Decimal("-30000")

    def test_revenue_positive_movement_is_favourable(self) -> None:
        prior = [_make_row(70001, section="REVENUE", ytd_actual="100000")]
        current = [_make_row(70001, section="REVENUE", ytd_actual="110000")]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is True

    def test_revenue_negative_movement_is_adverse(self) -> None:
        prior = [_make_row(70001, section="REVENUE", ytd_actual="110000")]
        current = [_make_row(70001, section="REVENUE", ytd_actual="100000")]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is False

    def test_expenditure_positive_movement_is_adverse(self) -> None:
        prior = [
            _make_row(80001, section="EXPENDITURE", subsection="Salaries", ytd_actual="100000")
        ]
        current = [
            _make_row(80001, section="EXPENDITURE", subsection="Salaries", ytd_actual="110000")
        ]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is False

    def test_expenditure_negative_movement_is_favourable(self) -> None:
        prior = [
            _make_row(80001, section="EXPENDITURE", subsection="Salaries", ytd_actual="110000")
        ]
        current = [
            _make_row(80001, section="EXPENDITURE", subsection="Salaries", ytd_actual="100000")
        ]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is True

    def test_zero_movement_is_not_favourable_or_adverse(self) -> None:
        prior = [_make_row(70001, ytd_actual="100000")]
        current = [_make_row(70001, ytd_actual="100000")]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is None
        assert lines[0].movement == Decimal("0")

    def test_threshold_dollar_boundary_gte(self) -> None:
        # Movement of exactly 5000 with threshold 5000 → should be flagged
        prior = [_make_row(70001, ytd_actual="100000")]
        current = [_make_row(70001, ytd_actual="105000")]
        lines = compare_opstat(prior, current, Decimal("5000"), 100)
        assert lines[0].exceeds_threshold is True

    def test_threshold_below_both_not_flagged(self) -> None:
        # Movement 100, threshold $5000 & 10% — below both
        prior = [_make_row(70001, ytd_actual="100000")]
        current = [_make_row(70001, ytd_actual="100100")]
        lines = compare_opstat(prior, current, Decimal("5000"), 10)
        assert lines[0].exceeds_threshold is False

    def test_threshold_pct_boundary(self) -> None:
        # 10% movement exactly should be flagged at threshold_pct=10
        prior = [_make_row(70001, ytd_actual="100000")]
        current = [_make_row(70001, ytd_actual="110000")]
        # Dollar threshold is high so only pct matters
        lines = compare_opstat(prior, current, Decimal("999999"), 10)
        assert lines[0].exceeds_threshold is True

    def test_capital_expenditure_direction(self) -> None:
        prior = [
            _make_row(26001, section="CAPITAL EXPENDITURE", subsection="Assets", ytd_actual="50000")
        ]
        current = [
            _make_row(26001, section="CAPITAL EXPENDITURE", subsection="Assets", ytd_actual="40000")
        ]
        lines = compare_opstat(prior, current, Decimal("1"), 1)
        assert lines[0].is_favourable is True  # lower capex = favourable


# ---------------------------------------------------------------------------
# 4. generate_opstat_comparison — end-to-end on real PDFs
# ---------------------------------------------------------------------------


class TestGenerateOpstatComparison:
    def test_produces_xlsx(self, tmp_path: Path) -> None:
        output = tmp_path / "compare.xlsx"
        progress_calls: list[tuple[int, str]] = []

        def progress(pct: int, msg: str) -> None:
            progress_calls.append((pct, msg))

        summary = generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("5000"),
            threshold_pct=10,
            progress=progress,
        )

        assert output.exists(), "Output XLSX was not created"
        assert isinstance(summary, OpStatSummary)
        assert len(summary.lines) >= 50

    def test_period_labels_in_summary(self, tmp_path: Path) -> None:
        output = tmp_path / "compare.xlsx"
        summary = generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("5000"),
            threshold_pct=10,
            progress=lambda p, m: None,
        )
        assert summary.period_prior == "28 February 2026"
        assert summary.period_current == "31 March 2026"

    def test_progress_callbacks_fired(self, tmp_path: Path) -> None:
        output = tmp_path / "compare.xlsx"
        pcts: list[int] = []

        def progress(pct: int, msg: str) -> None:
            pcts.append(pct)

        generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("5000"),
            threshold_pct=10,
            progress=progress,
        )

        assert 10 in pcts
        assert 30 in pcts
        assert 50 in pcts
        assert 70 in pcts
        assert 100 in pcts

    def test_has_favourable_and_adverse_rows(self, tmp_path: Path) -> None:
        output = tmp_path / "compare.xlsx"
        summary = generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("1"),
            threshold_pct=1,
            progress=lambda p, m: None,
        )

        favourable = [
            ln for ln in summary.lines if ln.exceeds_threshold and ln.is_favourable is True
        ]
        adverse = [ln for ln in summary.lines if ln.exceeds_threshold and ln.is_favourable is False]
        # With threshold $1 / 1% the real data should have both directions
        assert favourable, "Expected at least 1 favourable row"
        assert adverse, "Expected at least 1 adverse row"

    def test_xlsx_fill_colours(self, tmp_path: Path) -> None:
        """Favourable rows get HL_SOURCE_ONLY fill; adverse rows get HL_MISMATCH."""
        output = tmp_path / "compare_fills.xlsx"
        summary = generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("1"),
            threshold_pct=1,
            progress=lambda p, m: None,
        )

        fav_line = next(
            (ln for ln in summary.lines if ln.exceeds_threshold and ln.is_favourable is True),
            None,
        )
        adv_line = next(
            (ln for ln in summary.lines if ln.exceeds_threshold and ln.is_favourable is False),
            None,
        )
        assert fav_line is not None, "No favourable row found for fill check"
        assert adv_line is not None, "No adverse row found for fill check"

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws is not None

        fav_code = str(fav_line.gl_code).zfill(5)
        adv_code = str(adv_line.gl_code).zfill(5)

        fav_found = False
        adv_found = False

        for row_idx in range(3, (ws.max_row or 3) + 1):
            cell_val = ws.cell(row_idx, 1).value
            if cell_val is None:
                continue
            cell_str = str(cell_val)
            fill = ws.cell(row_idx, 1).fill
            if fill and fill.fgColor and fill.fgColor.rgb:
                rgb = fill.fgColor.rgb.upper()
                if cell_str == fav_code and HL_SOURCE_ONLY in rgb:
                    fav_found = True
                if cell_str == adv_code and HL_MISMATCH in rgb:
                    adv_found = True

        assert fav_found, f"Favourable row {fav_code} does not have HL_SOURCE_ONLY fill"
        assert adv_found, f"Adverse row {adv_code} does not have HL_MISMATCH fill"

    def test_no_fill_for_unchanged_rows(self, tmp_path: Path) -> None:
        """A row with zero movement should have no fill (or a neutral fill)."""
        output = tmp_path / "compare_nofill.xlsx"
        summary = generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("5000"),
            threshold_pct=10,
            progress=lambda p, m: None,
        )

        no_fill_lines = [ln for ln in summary.lines if not ln.exceeds_threshold]
        if not no_fill_lines:
            pytest.skip("All rows exceed threshold at these settings")

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws is not None

        sample_code = str(no_fill_lines[0].gl_code).zfill(5)
        for row_idx in range(3, (ws.max_row or 3) + 1):
            if str(ws.cell(row_idx, 1).value) == sample_code:
                fill = ws.cell(row_idx, 1).fill
                has_fill = bool(
                    fill
                    and fill.fill_type == "solid"
                    and fill.fgColor
                    and fill.fgColor.rgb
                    and fill.fgColor.rgb not in ("00000000", "FFFFFFFF", "")
                )
                assert not has_fill, f"Row {sample_code} below threshold should not be highlighted"
                break

    def test_xlsx_has_correct_headers(self, tmp_path: Path) -> None:
        output = tmp_path / "compare_hdrs.xlsx"
        generate_opstat_comparison(
            current_file=_CURRENT_PDF,
            prior_file=_PRIOR_PDF,
            output_file=output,
            threshold_dollars=Decimal("5000"),
            threshold_pct=10,
            progress=lambda p, m: None,
        )

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(1, c).value for c in range(1, 10)]
        assert headers[0] == "Account"
        assert headers[1] == "Description"
        assert "Section" in headers
        assert "Movement" in headers
