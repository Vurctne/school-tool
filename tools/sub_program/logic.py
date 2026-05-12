from __future__ import annotations

import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from toolkit.base_tool import ProgressFn
from toolkit.fills import argb
from toolkit.tokens import HL_MISMATCH

# Pink row fill for over-budget rows in the XLSX output.
# Uses argb() from toolkit.fills to derive the correct openpyxl ARGB literal
# from the canonical HL_MISMATCH token — same approach as master_budget/logic.py.
_OVER_FILL = PatternFill(fill_type="solid", fgColor=argb(HL_MISMATCH))

# Round 39 — comment-column header synonyms.  Module-scope so that the
# fall-through "no comments column found" error path can quote them
# even when every input sheet was empty (otherwise we hit
# UnboundLocalError before surfacing the helpful message).
_COMMENT_SYNONYMS: tuple[str, ...] = (
    "comment",
    "commentary",
    "note",
    "remark",
    "memo",
)

# ---------------------------------------------------------------------------
# Round 51 Phase D — structured commentary
# ---------------------------------------------------------------------------
# Replaces freeform commentary with three Combobox-backed dropdowns plus
# a free-text Notes paragraph. The XLSX cell carries an optional prefix
# of the shape ``[Driver: X | Outlook: Y | Action: Z] notes``; only set
# fields appear in the prefix, blank fields are omitted. When all three
# dropdowns are blank the prefix is suppressed entirely so pre-Phase-D
# files round-trip as Notes-only.

_DRIVER_VALUES: tuple[str, ...] = (
    "One-time",
    "Ongoing",
    "Structural",
    "Timing-early",
    "Timing-late",
    "Investigating",
)
_OUTLOOK_VALUES: tuple[str, ...] = (
    "One-time",
    "Expected to continue",
    "Improving",
    "Deteriorating",
)
_ACTION_VALUES: tuple[str, ...] = (
    "None",
    "Monitor",
    "Investigate",
    "Update forecast",
)

# Greedy-up-to-first-`]` body match. The body must contain at least one
# of the three keys to be treated as a structured prefix; otherwise the
# whole cell falls through to freeform Notes (e.g. ``"[NOTE TO SELF]
# review later"`` is preserved verbatim, NOT parsed as a prefix).
_COMMENTARY_PREFIX_RE = re.compile(
    r"^\[(?P<body>[^\]\n]*)\](?P<rest>.*)$",
    re.DOTALL,
)
_PREFIX_FIELD_RE = re.compile(r"\s*(?P<key>Driver|Outlook|Action)\s*:\s*(?P<val>[^|]*?)\s*(?:\||$)")


def encode_commentary(
    notes: str,
    driver: str = "",
    outlook: str = "",
    action: str = "",
) -> str:
    """Encode structured commentary into a single string.

    Used as the value written to the XLSX Comments cell. The prefix is
    only emitted when at least one structured field is non-empty;
    otherwise we return the bare notes (preserving backward
    compatibility with pre-Phase-D files). Blank fields are omitted
    from the prefix entirely — e.g. ``encode_commentary("n", action="Monitor")``
    → ``"[Action: Monitor]\\nn"``.

    Edge case: when all three structured fields are blank but the notes
    paragraph happens to start with ``[``, we emit an empty-body
    prefix ``[]`` as an escape so :func:`decode_commentary` can later
    distinguish "user typed a literal bracket" from "structured prefix
    we couldn't parse".

    Round 1 fix (R51): the prefix and notes are separated by a NEWLINE
    so when ``wrap_text`` is on the XLSX cell, notes start on their own
    visual line — the prefix doesn't visually merge into the paragraph
    when the cell wraps mid-prefix in a 50-char column.
    :func:`decode_commentary` strips exactly one separator either way,
    so the round-trip is unaffected.

    Excel-formula-injection guard: this encoder NEVER prepends
    ``=``/``+``/``-``/``@`` (the Excel formula sigils) — those checks
    happen at the cell-write site since not every encoder consumer
    writes to a cell. See ``_write_monthly_sub_program_sheet``.
    """
    fields: list[str] = []
    if driver:
        fields.append(f"Driver: {driver}")
    if outlook:
        fields.append(f"Outlook: {outlook}")
    if action:
        fields.append(f"Action: {action}")
    if not fields:
        if notes.lstrip().startswith("["):
            return f"[]\n{notes}"
        return notes
    prefix = "[" + " | ".join(fields) + "]"
    if not notes:
        return prefix
    return f"{prefix}\n{notes}"


def decode_commentary(text: str) -> tuple[str, str, str, str]:
    """Inverse of :func:`encode_commentary`.

    Returns ``(notes, driver, outlook, action)``. Anything that doesn't
    look like a Phase-D prefix is treated as freeform notes — so a
    pre-Phase-D file with unstructured text round-trips as
    ``(text, "", "", "")``.

    Round 1 fixes (R51):

    * **Whitespace preservation** — only the prefix-adjacent separator
      (the single space or newline immediately after ``]``) is stripped.
      Inner whitespace in the user's notes is preserved so a
      save→reopen cycle is idempotent.
    * **Unknown-value validation** — extracted Driver / Outlook / Action
      values are checked against the canonical
      :data:`_DRIVER_VALUES` / :data:`_OUTLOOK_VALUES` /
      :data:`_ACTION_VALUES` tuples. If any value is unknown, the
      whole text is preserved verbatim as Notes — protects users
      whose pre-Phase-D commentary happened to contain literal
      ``[Driver: foo]`` text from being silently mis-parsed and losing
      part of their note. Also the Editor's Combobox preload then
      relies on this contract.

    Special cases:

    * ``[] rest`` — empty-body prefix is the encoder's escape for
      "notes started with ``[`` and dropdowns were blank". We strip
      ``[]`` and return ``rest`` as Notes.
    * ``[FREE TEXT] more`` — body without any Phase-D key. Preserved
      verbatim as Notes (no information loss).
    """
    if not text:
        return "", "", "", ""
    m = _COMMENTARY_PREFIX_RE.match(text)
    if m is None:
        return text, "", "", ""
    body = m.group("body").strip()
    rest_raw = m.group("rest")
    # Strip exactly one separator (newline or space) immediately
    # following the closing ``]`` — preserves inner whitespace in the
    # user's notes so encode→decode→encode is idempotent.
    if rest_raw.startswith("\n") or rest_raw.startswith(" "):
        rest = rest_raw[1:]
    else:
        rest = rest_raw
    if body == "":
        # Empty-body prefix = encoder's escape for "[..." in Notes.
        return rest, "", "", ""
    # Body must contain at least one Phase-D key, else treat the whole
    # thing as freeform notes (preserves "[FREE TEXT] more" verbatim).
    if not any(k in body for k in ("Driver:", "Outlook:", "Action:")):
        return text, "", "", ""
    driver = outlook = action = ""
    for fm in _PREFIX_FIELD_RE.finditer(body):
        key = fm.group("key")
        val = fm.group("val").strip()
        if key == "Driver":
            driver = val
        elif key == "Outlook":
            outlook = val
        elif key == "Action":
            action = val
    # Round 2 fix (R51): preserve unknown values verbatim instead of
    # discarding the whole prefix. The Round-1 "fall through to Notes
    # on unknown value" rule collided with the editor's Combobox
    # preservation (frame.py:_add_combobox_row) and produced a
    # round-trip data-loss bug — a non-canonical value the editor
    # preserved was stripped on the next XLSX read, demoting Driver
    # back into Notes. Trading that for a small pre-Phase-D risk:
    # freeform text like ``"[Driver: training costs] note"`` will now
    # extract Driver as ``"training costs"`` instead of staying as
    # Notes. The ``_DRIVER_VALUES`` / etc tuples remain the source of
    # truth for the editor's dropdown — a user opening such a row sees
    # the legacy value AND can re-pick a canonical one — so any
    # mis-parse is recoverable. The frequency of pre-Phase-D
    # commentary literally beginning with ``[Driver:`` / ``[Outlook:``
    # / ``[Action:`` is empirically very low; the frequency of
    # hand-edited or schema-drifted values is higher.
    return rest, driver, outlook, action


# ---------------------------------------------------------------------------
# Round 53 F1 — Status pills (Move B)
# ---------------------------------------------------------------------------
# The XLSX output gains a per-sub-program ``Status`` column whose value
# is one of six plain-English pills. The pill replaces — for non-finance
# readers — the unreadable ``Available Balance % YTD`` column (which can
# read ``-2.21`` for a 221% overdraw and which the actual KMAR file
# shows as a literal ``7`` for stale ``=N/A`` cells).

_STATUS_ON_TRACK = "On track"
_STATUS_SLIGHTLY_OVER = "Slightly over"
_STATUS_MATERIAL = "Significant overspend"
_STATUS_URGENT = "Investigate urgently"
_STATUS_SPENT_WITHOUT_BUDGET = "Spent without budget"
# Round 62 — Revenue side surfaced in Status. Pre-R62 the pill only
# inspected the Expenditure side, so sub-programs whose Revenue YTD
# materially exceeded Revenue budget (parent payments collected
# ahead of schedule, fundraisers that beat target) printed pink-
# filled on the Watchlist (per-line is_over+is_material aggregation)
# but with Status="On track" — a council-facing contradiction.
# "Revenue over budget" sits at the bottom of the urgency stack: it
# IS noteworthy (council needs to acknowledge over-collection or
# plan refunds / roll-forward) but never escalates ahead of
# Expense-side overruns. The Expense-side buckets always win when
# both fire on the same row.
# Round 71 — O2 rename: "Revenue over budget" → "Revenue exceeded
# target". In school finance parlance, "over budget" usually means
# bad (overspent); over-collected revenue is generally good. The new
# label removes the negative connotation and reads naturally:
# "Revenue exceeded target" = "we got more than expected".
_STATUS_REVENUE_OVER = "Revenue exceeded target"
# Round 71 — O1: net negative revenue (refunds outpaced collections)
# is the only unambiguous under-collection signal without calendar
# awareness, which Round 56 deliberately removed. Council needs to
# know when a sub-program has more refunds than receipts.
_STATUS_REVENUE_REFUNDED = "Revenue refunded"

# Round 56 — "No spend yet" pill dropped along with all calendar-
# pacing judgements. Round 62 — added "Revenue over budget". Round 71
# — renamed to "Revenue exceeded target" + added "Revenue refunded".
_STATUS_VALUES: tuple[str, ...] = (
    _STATUS_ON_TRACK,
    _STATUS_SLIGHTLY_OVER,
    _STATUS_MATERIAL,
    _STATUS_URGENT,
    _STATUS_SPENT_WITHOUT_BUDGET,
    _STATUS_REVENUE_OVER,
    _STATUS_REVENUE_REFUNDED,
)

# Bucket boundaries for overrun magnitude. Both dollar AND percent
# triggers fire — whichever is larger picks the bucket. This matches
# the variance-analysis skill's "either exceeded" rule.
_STATUS_URGENT_DOLLAR = Decimal("100000")
_STATUS_URGENT_PCT = Decimal("50")
_STATUS_MATERIAL_DOLLAR = Decimal("25000")
_STATUS_MATERIAL_PCT = Decimal("25")


def compute_status_pill(
    *,
    annual_exp_budget: Decimal,
    exp_ytd: Decimal,
    annual_rev_budget: Decimal = Decimal("0"),
    rev_ytd: Decimal = Decimal("0"),
    expense_threshold: float = 101.0,
    revenue_threshold: float = 101.0,
    materiality_dollar: int = 100,
) -> str:
    """Return a plain-English Status pill for one sub-program.

    Round 56 redesign — pacing-free contract:

    * Over-budget gate: ``exp_ytd > expense_threshold% × annual_exp_budget``.
      The threshold is the same Expense over-budget slider the user
      sets (default 101%). Sub-programs whose Expense YTD exceeds
      that fraction of the annual expense budget are flagged.
    * ``Spent without budget`` — truly unbudgeted spend: zero budget
      on BOTH sides, no revenue collected, but expenditure occurred.
      Capital-spend-without-council-approval flag.
    * ``On track`` — used_pct ≤ threshold, OR overrun below the
      materiality floor (both dollar and percent).
    * Bucket above materiality:
        * ``Investigate urgently`` — overrun > $100K OR > 50% past threshold.
        * ``Significant overspend`` — $25K–$100K OR 25–50% past threshold.
        * ``Slightly over`` — anything else past materiality.

    Round 62 — Revenue side checked when Expense passes the threshold.
    Round 71 (O2) — ``Revenue over budget`` renamed to
    ``Revenue exceeded target`` to remove the negative connotation
    of "over budget" for what's usually good news.

    * ``Revenue exceeded target`` — when expense isn't over-budget
      but revenue YTD materially exceeds the revenue threshold
      (default 101% of annual revenue budget). Sits at the bottom of
      the urgency stack — over-collection is generally benign but
      council needs to acknowledge it (refund parents, roll forward,
      or adjust the budget). Expense-side buckets always win when
      both fire on the same row.

    Round 71 (O1) — added ``Revenue refunded`` for the inverse
    direction:

    * ``Revenue refunded`` — when revenue YTD is materially negative
      (refunds outpaced receipts). The only unambiguous under-
      collection signal without calendar-pacing context. Fires
      regardless of budget side or expense status, ahead of the
      "exceeded target" check.

    Round 56 dropped from the prior contract:

    * ``available`` parameter (and the surplus / over-drawn signal
      derived from it) — expense vs threshold replaces it.
    * ``calendar_pct`` parameter (and the pacing-aware Expenditure-
      only branch) — pacing semantics removed entirely.
    * ``No spend yet`` pill — depended on calendar.
    """
    mat = Decimal(str(materiality_dollar))

    # Spent without budget — truly unbudgeted spend.
    # Round 71 F2: previously required rev_ytd == 0 too, which masked
    # the case where a sub-program collected a small token of revenue
    # ($100 donation) and then spent $10K against it (no budget on
    # either side). Now fires when both budgets are zero AND the
    # expense YTD materially exceeds whatever revenue was collected —
    # i.e. someone spent money that wasn't covered by either a budget
    # or by collected revenue. Cost-recovery programs with a real
    # rev budget (rev_b > 0) still don't fire because the rev_b == 0
    # clause above is unchanged.
    if (
        annual_exp_budget == 0
        and annual_rev_budget == 0
        and exp_ytd > 0
        and (exp_ytd - rev_ytd) >= mat
    ):
        return _STATUS_SPENT_WITHOUT_BUDGET

    # Helper closure: revenue-side pill, if any. Returns a status
    # string or None. Defined once and reused at every "else →
    # On track" branch so we don't accidentally skip the check on
    # some code paths.
    def _revenue_status() -> str | None:
        # Round 71 O1: net negative revenue (refunds outpaced
        # collections) is concerning regardless of budget side or
        # calendar. Fires first because a refunded program with an
        # otherwise on-track expense side still needs the council
        # to know about the cash flow direction.
        if rev_ytd < -mat:
            return _STATUS_REVENUE_REFUNDED

        rev_th = Decimal(str(revenue_threshold))
        # Unbudgeted revenue collected (e.g. surprise donation under
        # a sub-program with no rev budget) → flag if positive and
        # material. Negative case was already handled above.
        if annual_rev_budget == 0:
            if rev_ytd > 0 and rev_ytd >= mat:
                return _STATUS_REVENUE_OVER
            return None
        # Budgeted revenue exceeds the threshold AND the dollar
        # overshoot meets the materiality floor.
        used = rev_ytd / annual_rev_budget * Decimal("100")
        if used <= rev_th:
            return None
        threshold_dollars = annual_rev_budget * rev_th / Decimal("100")
        overrun = rev_ytd - threshold_dollars
        if abs(overrun) >= mat:
            return _STATUS_REVENUE_OVER
        return None

    # No expense budget allocated and no spend on expense side. Check
    # the revenue side before concluding "On track".
    if annual_exp_budget == 0:
        rev_pill = _revenue_status()
        return rev_pill if rev_pill else _STATUS_ON_TRACK

    threshold = Decimal(str(expense_threshold))
    used_pct = exp_ytd / annual_exp_budget * Decimal("100")
    if used_pct <= threshold:
        rev_pill = _revenue_status()
        return rev_pill if rev_pill else _STATUS_ON_TRACK

    # Over budget on expense. Compute overrun in dollar AND in percent
    # past the threshold so the materiality floor + bucketing logic can
    # apply the variance-analysis skill's "either exceeded" rule.
    threshold_dollars = annual_exp_budget * threshold / Decimal("100")
    overrun = exp_ytd - threshold_dollars  # positive
    pct_over = used_pct - threshold  # how many pp past threshold

    # Hard $500 noise floor — chart-of-accounts placeholder rows
    # below this magnitude are not council-grade attention.
    if overrun < Decimal("500"):
        rev_pill = _revenue_status()
        return rev_pill if rev_pill else _STATUS_ON_TRACK

    # Materiality floor — overrun below BOTH dollar and percent
    # floors collapses back to On track (or Revenue-over if applicable).
    if overrun < mat and pct_over <= Decimal("50"):
        rev_pill = _revenue_status()
        return rev_pill if rev_pill else _STATUS_ON_TRACK

    # Expense-side urgency tiers — these always win over Revenue.
    if overrun > _STATUS_URGENT_DOLLAR or pct_over > _STATUS_URGENT_PCT:
        return _STATUS_URGENT
    if overrun > _STATUS_MATERIAL_DOLLAR or pct_over > _STATUS_MATERIAL_PCT:
        return _STATUS_MATERIAL
    return _STATUS_SLIGHTLY_OVER


# ---------------------------------------------------------------------------
# Round 53 F1 — Plain-English commentary (Move E)
# ---------------------------------------------------------------------------
# The XLSX Comments cell renders the structured Phase-D triplet as
# plain English instead of the bracketed prefix that lives internally.
# Round-trip fidelity through prior-period files is sacrificed in the
# rendered cell (the visible value is prose, not the prefix); the
# decoder still works on legacy R51 prefix-encoded cells.

_DRIVER_PROSE: dict[str, str] = {
    "One-time": "One-time variance",
    "Ongoing": "Ongoing variance",
    "Structural": "Structural variance",
    "Timing-early": "Spend earlier than planned",
    "Timing-late": "Spend later than planned",
    "Investigating": "Driver under investigation",
}

_OUTLOOK_PROSE: dict[str, str] = {
    "One-time": "won't recur",
    "Expected to continue": "expected to continue",
    "Improving": "improving",
    "Deteriorating": "deteriorating",
}

_ACTION_PROSE: dict[str, str] = {
    "None": "no action needed",
    "Monitor": "being monitored",
    "Investigate": "needs investigation",
    "Update forecast": "forecast update needed",
}


def _capitalize_first(s: str) -> str:
    """Uppercase only the first character. Differs from str.capitalize
    which lower-cases the rest — we want 'Reviewed by HOD' to keep
    'HOD' upper-cased, not become 'Reviewed by hod'."""
    return s[:1].upper() + s[1:] if s else s


def _ensure_terminal_period(s: str) -> str:
    """Ensure the trimmed string ends with terminal punctuation."""
    s = s.strip()
    if not s:
        return s
    if s[-1] in (".", "!", "?"):
        return s
    return s + "."


# Combinations that read as logically contradictory and should NOT
# render together. R1 fix: Round-1 logic skeptic flagged these.
# Format: (driver, outlook) tuples whose outlook is dropped.
_CONTRADICTORY_DRIVER_OUTLOOK: frozenset[tuple[str, str]] = frozenset(
    {
        # Structural variance won't recur — by definition it's permanent.
        ("Structural", "One-time"),
        # One-time variance "expected to continue" — definitionally wrong.
        ("One-time", "Expected to continue"),
        # R2 fix: "we don't know what's driving this AND we know it's
        # getting better" reads as incoherent. ``Investigating`` +
        # ``Deteriorating`` is fine ("we don't know why but it's
        # getting worse").
        ("Investigating", "Improving"),
    }
)

# Driver / Action combinations where the action is dropped (it
# duplicates or contradicts the driver).
_CONTRADICTORY_DRIVER_ACTION: frozenset[tuple[str, str]] = frozenset(
    {
        # "Driver under investigation" + "no action needed" is a direct
        # contradiction — investigating IS an action.
        ("Investigating", "None"),
        # R2 fix: "Driver under investigation. Needs investigation." is
        # repetitive — same word root in two sentences. The driver
        # phrase already implies the investigation; drop the
        # action_prose.
        ("Investigating", "Investigate"),
    }
)


def render_commentary_prose(
    notes: str = "",
    driver: str = "",
    outlook: str = "",
    action: str = "",
) -> str:
    """Render structured commentary as plain-English prose for the
    XLSX Comments cell.

    Returns a 1–2 sentence string. The structured triplet (Driver +
    Outlook + Action) collapses to one or two short sentences (each
    with its own period — R1 fix replaced the em-dash splice that
    read as software-generated); the freeform Notes paragraph is the
    final sentence. All blank returns "".

    Unknown structured values (schema drift, hand edit) fall through
    silently — no crash, no half-sentence.

    Round 1 fix: contradictory triplet combinations have the conflicting
    field dropped (e.g. ``Structural`` + ``One-time`` outlook drops the
    outlook; ``Investigating`` + ``None`` action drops the action). See
    :data:`_CONTRADICTORY_DRIVER_OUTLOOK` / `_CONTRADICTORY_DRIVER_ACTION`.
    """
    driver_prose = _DRIVER_PROSE.get(driver, "")
    outlook_prose = _OUTLOOK_PROSE.get(outlook, "")
    action_prose = _ACTION_PROSE.get(action, "")

    # R1 fix: drop conflicting outlook / action combinations.
    if driver and outlook and (driver, outlook) in _CONTRADICTORY_DRIVER_OUTLOOK:
        outlook_prose = ""
    if driver and action and (driver, action) in _CONTRADICTORY_DRIVER_ACTION:
        action_prose = ""

    # Sentence 1: description (Driver + Outlook joined with ", ").
    # Sentence 2: action_prose, capitalised, its own sentence.
    # Sentence 3: notes verbatim with terminal punctuation guaranteed.
    # R1 fix: replaced em-dash splice ("description — action") with two
    # short sentences ("description. Action."). Reads as natural
    # English instead of template output.
    sentences: list[str] = []

    # R2 fix: when ``Investigating`` driver and outlook are both set,
    # the comma-joined form ("Driver under investigation, improving")
    # parses ambiguously — "improving" attaches to "investigation"
    # rather than to the variance. Render as two unambiguous
    # sentences instead.
    if driver == "Investigating" and outlook_prose and driver_prose:
        sentences.append(_ensure_terminal_period(_capitalize_first(driver_prose)))
        sentences.append(_ensure_terminal_period(_capitalize_first("variance " + outlook_prose)))
    else:
        desc_pieces: list[str] = []
        if driver_prose:
            desc_pieces.append(driver_prose)
        if outlook_prose:
            if desc_pieces:
                desc_pieces.append(outlook_prose)
            else:
                # Outlook leads — "Outlook improving."
                desc_pieces.append("Outlook " + outlook_prose)
        description = ", ".join(desc_pieces)
        if description:
            sentences.append(_ensure_terminal_period(_capitalize_first(description)))

    if action_prose:
        sentences.append(_ensure_terminal_period(_capitalize_first(action_prose)))
    if notes:
        sentences.append(_ensure_terminal_period(_capitalize_first(notes.strip())))

    return " ".join(sentences)


# Round 67 — ``cap_percent_for_display`` deleted along with the
# ``_PERCENT_CAP`` constant. The Round 53 F1 (Move F) ±999% cap was
# dropped per user feedback that the ">999%" / "<-999%" text markers
# hid the actual magnitude. Percent cells now always render the
# Excel formula; the data bar (col C/D, 0–100% pinned) saturates
# visually for extreme values while the number text shows the truth.


# Round 62 — Trend column functions deleted. The Trend feature
# was dropped from the XLSX in Round 57; compute_trend and
# load_prior_period_ytd had no production callers afterwards (only
# their own tests). Removed in Round 63 to clear the dead code.


def load_prior_period_funds(xlsx_path: Path) -> dict[str, Decimal]:
    """Round 57 — return ``{sub_program_code: funds_from_previous_years}``
    for every data row in the supplied prior-period XLSX.

    Reads the "Funds from Previous Years" (or "Funds from Previous Years
    (Funds)") column and pairs it with the sub-program code. Empty dict
    when the file is supplied but doesn't carry the column. Skips the
    Watchlist sheet (filtered subset).

    Used by the writer to populate the same column in the current
    period's output, preserving the carry-forward field across reports
    so a school doesn't have to re-enter it every period.
    """
    if not xlsx_path.exists():
        raise ValueError(f"Prior-period file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        result: dict[str, Decimal] = {}
        code_synonyms = ("code", "sub-prog", "sub prog", "sub program", "sub-program")
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == "watchlist":
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            for header_row_idx in (1, 0):
                if header_row_idx >= len(rows):
                    continue
                header_raw = rows[header_row_idx]
                header = [str(c).strip() if c is not None else "" for c in header_raw]
                header_lower = [h.lower() for h in header]
                code_col: int | None = None
                for i, h in enumerate(header_lower):
                    if any(syn in h for syn in code_synonyms):
                        code_col = i
                        break
                funds_col: int | None = None
                for i, h in enumerate(header_lower):
                    if "funds from previous years" in h:
                        funds_col = i
                        break
                if code_col is None or funds_col is None:
                    continue
                for data_row in rows[header_row_idx + 1 :]:
                    if not data_row:
                        continue
                    sp_raw = (
                        data_row[code_col]
                        if code_col < len(data_row) and data_row[code_col] is not None
                        else None
                    )
                    if sp_raw is None:
                        continue
                    sp = str(sp_raw).strip()
                    if sp.endswith(".0"):
                        sp = sp[:-2]
                    if not sp.replace(".", "").isdigit():
                        continue
                    funds_raw = data_row[funds_col] if funds_col < len(data_row) else None
                    if funds_raw is None:
                        continue
                    try:
                        result[sp] = Decimal(str(funds_raw))
                    except (InvalidOperation, ValueError):
                        continue
                break
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Faculty inference
# ---------------------------------------------------------------------------
# Sub-program codes are numeric (e.g. 4001, 8599).
# Faculties are inferred from the leading digit(s) as visible in the PDF.
# The CASES21 GL21157 export groups rows by Revenue / Expenditure sections
# rather than by faculty, so we derive faculty from the code prefix.

_FACULTY_MAP: dict[str, str] = {
    "1": "Design & Technology",
    "4": "Curriculum",
    "5": "Student Wellbeing",
    "6": "Facilities",
    "7": "Administration",
    "8": "Programs & Camps",
    "9": "Computing & Curriculum",
}


def _infer_faculty(sub_program: str) -> str | None:
    """Return a faculty label from the first digit of *sub_program*."""
    code = sub_program.strip()
    if code and code[0].isdigit():
        return _FACULTY_MAP.get(code[0])
    return None


# ---------------------------------------------------------------------------
# Currency parsing
# ---------------------------------------------------------------------------

_DASH_RE = re.compile(r"^[—–\-]{1,2}$")  # em-dash, en-dash, hyphen alone


def parse_decimal(raw: str) -> Decimal:
    """Convert a CASES21 currency string to Decimal.

    Handles: ``"1,234.56"``, ``"$1,234.56"``, ``"-500.00"``, ``"(500.00)"``,
    ``"$0.00"``, ``"—"`` (em-dash = zero), blank strings.
    Returns ``Decimal("0")`` for empty / dash inputs.
    """
    text = raw.strip() if raw else ""

    if not text or _DASH_RE.match(text):
        return Decimal("0")

    # parentheses -> negative  e.g. "(500.00)" -> "-500.00"
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]

    # strip leading $
    text = text.lstrip("$").strip()

    # remove thousands separators
    text = text.replace(",", "")

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse {raw!r} as a decimal: {exc}") from exc


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SubProgramLine:
    sub_program: str
    account: str  # section tag: "Revenue" | "Expenditure"
    description: str
    budget: Decimal
    ytd: Decimal
    remaining: Decimal
    used_pct: Decimal  # 0..100+
    faculty: str | None
    is_over: bool
    # Round 51 Phase D — ``commentary`` is now FREEFORM NOTES only.
    # Pre-Phase-D code stored everything in this single field; the new
    # XLSX prefix carries the structured triplet (driver/outlook/action)
    # alongside the notes via :func:`encode_commentary`.
    commentary: str = ""
    # Round 51 Phase D — structured commentary triplet. Each field is
    # one of the ``_DRIVER_VALUES`` / ``_OUTLOOK_VALUES`` /
    # ``_ACTION_VALUES`` module-scope tuples, or empty string for
    # "not categorised" (distinct from the literal ``"None"`` Action,
    # which means "user reviewed and decided no action needed").
    commentary_driver: str = ""
    commentary_outlook: str = ""
    commentary_action: str = ""
    # New fields -- default to zero so existing callers remain valid.
    last_year_actual: Decimal = Decimal("0")
    last_year_budget: Decimal = Decimal("0")
    outstanding_orders: Decimal = Decimal("0")
    # Round 45 Phase A — variance fields. Round 56 dropped the
    # ``pacing`` field per user direction (all pacing computations
    # removed; calendar-aware judgements gone).
    #
    # variance_amount: signed YTD - Budget. Positive means YTD has
    # exceeded the annual budget. Sign is unconditional (same
    # convention for Revenue and Expenditure rows).
    #
    # variance_pct: variance_amount / budget * 100, also signed. Zero
    # when budget is zero.
    variance_amount: Decimal = Decimal("0")
    variance_pct: Decimal = Decimal("0")
    # Round 45 Phase A — materiality flag. True when |variance_amount|
    # meets OR exceeds the materiality dollar floor for this run.
    is_material: bool = False


@dataclass(frozen=True)
class ReportSummary:
    lines: list[SubProgramLine]
    faculty_counts: dict[str, int]
    over_budget_lines: list[SubProgramLine]
    total_budget: Decimal
    total_ytd: Decimal
    output_path: Path
    faculty_budget: dict[str, Decimal] = field(default_factory=dict)
    faculty_ytd: dict[str, Decimal] = field(default_factory=dict)
    faculty_used_pct: dict[str, Decimal] = field(default_factory=dict)
    period_label: str = ""  # e.g. "March 2026" -- extracted from the PDF footer
    over_budget_threshold: float = 101.0  # threshold used for is_over computation
    # Round 21 — separate Revenue / Expense thresholds.  When the user does
    # not split them explicitly, both fields mirror over_budget_threshold
    # (set in generate_report) so existing callers stay valid.
    revenue_threshold: float = 101.0
    expense_threshold: float = 101.0
    # Round 45 Phase A — dollar materiality. Round 56 dropped the
    # ``calendar_pct`` field along with all pacing-based judgements.
    materiality_dollar: int = 100
    # Round 57 — carry-forward funds loaded from a prior-period XLSX
    # via :func:`load_prior_period_funds`. Empty dict when no prior
    # file was supplied. The writer reads this and populates the
    # "Funds from Previous Years" column for matching sub-programs;
    # rows with no match leave the cell blank (no fabricated zeros).
    prior_funds: dict[str, Decimal] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

# Header patterns that mark the start of a section
_REVENUE_HDR = re.compile(r"Revenue Recurrent", re.IGNORECASE)
_EXPENDITURE_HDR = re.compile(r"Expenditure Recurrent", re.IGNORECASE)

# Lines we must skip
_SKIP_RE = re.compile(
    r"^("
    r"\d{4}:\w"  # school header e.g. "8819:Melbourne"
    r"|General Ledger"
    r"|Annual Sub Program"
    r"|From Sub Program"
    r"|Revenue Recurrent"
    r"|Expenditure Recurrent"
    r"|Sub Prog\."  # column header
    r"|Revenue totals"
    r"|Expenditure totals"
    r"|\d+ \w+ \d{4}"  # date footer e.g. "3 March 2026"
    r"|\d+ \[GL"  # page/number footer e.g. "1 [GL21157]"
    r")",
    re.IGNORECASE,
)

# Matches a whitespace-delimited field that is a standalone numeric token:
# optional leading minus/dollar, digits with commas, optional decimal part;
# OR a parenthesised value like (500.00).
_NUM_PART_RE = re.compile(r"^[\-\$]?[\d,]+(\.\d+)?$|^\([\d,]+(\.\d+)?\)$")


# Round 61 — _parse_numeric_tokens_from_parts, _build_line, and
# _parse_text_lines (the heuristic token-list parser) deleted along
# with all pct-based column disambiguation. Replaced by
# _parse_page_positionally below, which uses pdfplumber x-coordinates
# to bin numeric tokens to fixed column right edges.


# Regex to extract the print-date footer: "3 March 2026 13:37 1 [GL21157]"
# Capture group 1: "D Month YYYY" string used as the period label.
# Round 65 widened the capture from "Month YYYY" to "D Month YYYY" so
# the workbook title shows the actual print date (which is the YTD
# cut-off the CASES21 export was run for) rather than just the
# month. Council readers care about WHICH day the snapshot was taken
# — early-month vs end-of-month numbers can differ materially.
_FOOTER_DATE_RE = re.compile(
    r"(\d{1,2}\s+[A-Z][a-z]+\s+\d{4})\s+\d{2}:\d{2}\s+\d+\s+\[GL",
    re.IGNORECASE,
)


def _extract_period_label(text: str) -> str:
    """Return 'D Month YYYY' from the GL21157 page-footer date, or '' if not found.

    The footer format is: ``3 March 2026 13:37 1 [GL21157]``
    This function extracts ``3 March 2026`` from that pattern.
    """
    m = _FOOTER_DATE_RE.search(text)
    if m:
        return m.group(1)
    return ""


# Round 56 — ``calendar_pct_from_period_label`` and the
# ``_MONTH_TO_PCT`` table dropped along with all pacing-based
# judgements. The period label is still extracted (for the report
# header) but no longer drives any classification.


# ---------------------------------------------------------------------------
# Public API -- parse from PDF
# ---------------------------------------------------------------------------


def parse_sub_program_pdf(pdf_path: Path) -> list[SubProgramLine]:
    """Parse a CASES21 GL21157 Annual Sub-Program Budget Report PDF.

    Returns a list of :class:`SubProgramLine` objects, one per data row.
    Skips header/footer rows and total rows.

    Raises :class:`ValueError` if the file appears empty or unrecognised.
    """
    lines, _period = _parse_sub_program_pdf_internal(pdf_path)
    return lines


def parse_sub_program_pdf_with_period(pdf_path: Path) -> tuple[list[SubProgramLine], str]:
    """Parse a GL21157 PDF and return ``(lines, period_label)``.

    ``period_label`` is a string like ``"March 2026"`` extracted from the
    footer date; it is ``""`` if detection fails.
    """
    return _parse_sub_program_pdf_internal(pdf_path)


def _group_words_by_row(
    words: list[dict[str, Any]], tolerance: float = 2.0
) -> list[list[dict[str, Any]]]:
    """Group pdfplumber word dicts into rows by their ``top`` coordinate.

    Words within ``tolerance`` pixels of each other vertically are
    treated as belonging to the same row. Returns rows sorted top-to-
    bottom; within each row, words are sorted left-to-right by ``x0``.
    """
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: (float(w["top"]), float(w["x0"])))
    rows: list[list[dict[str, Any]]] = []
    current_row: list[dict[str, Any]] = []
    current_top: float | None = None
    for w in sorted_words:
        wtop = float(w["top"])
        if current_top is None or abs(wtop - current_top) <= tolerance:
            current_row.append(w)
            if current_top is None:
                current_top = wtop
        else:
            rows.append(sorted(current_row, key=lambda x: float(x["x0"])))
            current_row = [w]
            current_top = wtop
    if current_row:
        rows.append(sorted(current_row, key=lambda x: float(x["x0"])))
    return rows


# Round 61 — column right-edge derivation from the GL21157 PDF header.
# Numbers in the data rows are right-aligned to fixed x positions
# matching these headers, so we bin tokens to columns by x1 (right
# edge) within a small tolerance. Pre-R61 the parser used pct-based
# heuristics to disambiguate which token was which column; this
# repeatedly broke on rows where a column was blank (e.g. zero YTD,
# zero orders) because the heuristic shifted the column window. The
# CASES21 export is generated with consistent layout, so positional
# extraction is the bullet-proof read.

# Each column has a unique, distinctive RIGHTMOST word in its header
# label whose x1 marks where data values right-align. We find that
# word directly; clustering / fuzzy joins were tried first but the
# inter-label gap between "Orders" and "Uncommitted" headers in the
# GL21157 layout is only ~9 px which is too small for a generic
# whitespace-cluster to split safely.
#
# Mapping rule per column key:
#   ly_actual     — word "actual" (only appears once, in "Last year actual")
#   ly_budget     — first "budget" word reading left-to-right
#   annual_budget — second "budget" word
#   ytd           — word "YTD"
#   pct           — last word in the percent column header
#                   ("Expended" for Expenditure pages, "received" for
#                   Revenue pages — both end with "ed", so we take the
#                   word immediately to the right of "Budget" that's
#                   not "budget" itself)
#   orders        — word "Orders"
#   uncommitted   — word "Balance" (right side of "Uncommitted Balance")


def _detect_column_edges(
    header_row: list[dict[str, Any]],
) -> dict[str, float]:
    """Build {column_key → right edge x1} from a header row.

    Round 61 — direct word-match approach. The GL21157 header layout
    has a 9-px gap between "Orders" and "Uncommitted" which trips
    every reasonable whitespace-cluster threshold, so we instead pick
    the distinctive word at each column's right edge.
    """
    if not header_row:
        return {}
    edges: dict[str, float] = {}
    sorted_row = sorted(header_row, key=lambda w: float(w["x0"]))

    # ly_actual — "actual" only appears in this column.
    for w in sorted_row:
        if w["text"].lower() == "actual":
            edges["ly_actual"] = float(w["x1"])
            break

    # ly_budget / annual_budget — the two "budget" words in left-to-
    # right order. (Header text: "Last year actual | Last year budget |
    # Annual budget | YTD ...".)
    budget_words = [w for w in sorted_row if w["text"].lower() == "budget"]
    if len(budget_words) >= 1:
        edges["ly_budget"] = float(budget_words[0]["x1"])
    if len(budget_words) >= 2:
        edges["annual_budget"] = float(budget_words[1]["x1"])

    # ytd — word "YTD".
    for w in sorted_row:
        if w["text"].upper() == "YTD":
            edges["ytd"] = float(w["x1"])
            break

    # pct — rightmost word of the percent column. The label is either
    # "% Budget Expended" (Expenditure page) or "% Budget received"
    # (Revenue page). Pick the rightmost word that comes AFTER the
    # second "Budget" word but BEFORE "Outstanding" (if present).
    pct_anchor_x = edges.get("annual_budget", 0.0) + 1.0 if "annual_budget" in edges else 0.0
    orders_word = next(
        (w for w in sorted_row if w["text"].lower() == "outstanding"),
        None,
    )
    pct_right_bound = float(orders_word["x0"]) if orders_word is not None else 1e9
    pct_label_words = [w for w in sorted_row if pct_anchor_x < float(w["x0"]) < pct_right_bound]
    if pct_label_words:
        edges["pct"] = max(float(w["x1"]) for w in pct_label_words)

    # orders — word "Orders" (Expenditure page only).
    for w in sorted_row:
        if w["text"].lower() == "orders":
            edges["orders"] = float(w["x1"])
            break

    # uncommitted — word "Balance" at the right edge of the page.
    for w in sorted_row:
        if w["text"].lower() == "balance":
            edges["uncommitted"] = float(w["x1"])
            break

    # sub_prog — "Prog." word ends the sub-program header.
    for w in sorted_row:
        if w["text"].lower() == "prog.":
            edges["sub_prog"] = float(w["x1"])
            break

    # title — "Title" word marks the title column's left edge (used
    # downstream to filter out non-title text).
    for w in sorted_row:
        if w["text"].lower() == "title":
            edges["title"] = float(w["x1"])
            break

    return edges


def _is_numeric_word(text: str) -> bool:
    """Cheap check: does this word look like a numeric value?"""
    return bool(_NUM_PART_RE.match(text))


def _parse_page_positionally(page: Any, section: str) -> list[SubProgramLine]:
    """Parse data rows from one PDF page using x-coordinate column binning.

    Round 61 — replaces the pre-R60 token-list heuristic that guessed
    column identity from the percent value. CASES21 GL21157 right-aligns
    each numeric column to a fixed x position; we read those positions
    from the page's column-header row and bin every numeric token into
    the column whose right edge is closest. Blank columns (zero YTD,
    no outstanding orders, no annual budget at all) simply produce no
    token at that x position, which the binning correctly reads as
    "field absent → defaults to zero".
    """
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    if not words:
        return []

    rows = _group_words_by_row(words)
    if not rows:
        return []

    # Find the column-header row: the row whose joined text contains
    # both "Sub Prog" and "Title". There may be a section banner row
    # ("Revenue Recurrent and Capital" / "Expenditure Recurrent and
    # Capital") above it; we skip past banner / report-title rows.
    header_row: list[dict[str, Any]] | None = None
    header_idx = -1
    for idx, row in enumerate(rows):
        joined = " ".join(w["text"] for w in row).lower()
        if "sub prog" in joined and "title" in joined and "actual" in joined:
            header_row = row
            header_idx = idx
            break

    if not header_row:
        return []

    column_edges = _detect_column_edges(header_row)
    # Required edges for any data row to make sense.
    if "sub_prog" not in column_edges or "annual_budget" not in column_edges:
        return []

    results: list[SubProgramLine] = []
    tolerance = 6.0  # px tolerance for token-to-column binning

    # Title left edge: x0 of the title-column header.
    title_x0_edge = next(
        (float(w["x0"]) for w in header_row if w["text"].lower() == "title"),
        column_edges.get("sub_prog", 50.0) + 20.0,
    )

    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        # Skip footer (page number + GL21157 marker) and totals rows.
        text = " ".join(w["text"] for w in row).strip()
        if _SKIP_RE.match(text):
            continue
        # First word must be a 4-digit sub-program code (right-aligned
        # to sub_prog column). Anything else (page footer, blank line)
        # is dropped.
        first = row[0]
        sub_prog = first["text"].strip()
        if not (len(sub_prog) == 4 and sub_prog.isdigit()):
            continue

        # Title: all words after sub_prog whose x1 falls before the
        # last-year-actual column's right edge minus tolerance.
        ly_actual_edge = column_edges.get("ly_actual", 0.0)
        title_words: list[str] = []
        numeric_words: list[dict[str, Any]] = []
        for w in row[1:]:
            wx1 = float(w["x1"])
            wtext = w["text"]
            if wx1 <= ly_actual_edge - tolerance:
                # Title territory.
                if float(w["x0"]) >= title_x0_edge - tolerance:
                    title_words.append(wtext)
                continue
            if _is_numeric_word(wtext):
                numeric_words.append(w)
            # Non-numeric words past the title column are ignored
            # (defensive — shouldn't occur in well-formed CASES21
            # exports).

        title = " ".join(title_words).strip()

        # Bin numeric words into columns by x1 proximity.
        fields: dict[str, Decimal] = {}
        for w in numeric_words:
            wx1 = float(w["x1"])
            wtext = w["text"]
            best_col: str | None = None
            best_dist = tolerance + 1.0
            for col_key, edge in column_edges.items():
                if col_key in ("sub_prog", "title"):
                    continue
                d = abs(wx1 - edge)
                if d < best_dist:
                    best_dist = d
                    best_col = col_key
            if best_col is None:
                continue
            try:
                fields[best_col] = parse_decimal(wtext)
            except (ValueError, InvalidOperation):
                continue

        budget = fields.get("annual_budget", Decimal("0"))
        ytd = fields.get("ytd", Decimal("0"))
        last_year_actual = fields.get("ly_actual", Decimal("0"))
        last_year_budget = fields.get("ly_budget", Decimal("0"))
        outstanding_orders = (
            fields.get("orders", Decimal("0")) if section == "Expenditure" else Decimal("0")
        )
        used_pct = fields.get("pct", Decimal("0"))

        remaining = budget - ytd
        faculty = _infer_faculty(sub_prog)
        is_over = bool(ytd > budget) if budget != Decimal("0") else False

        results.append(
            SubProgramLine(
                sub_program=sub_prog,
                account=section,
                description=title,
                budget=budget,
                ytd=ytd,
                remaining=remaining,
                used_pct=used_pct,
                faculty=faculty,
                is_over=is_over,
                last_year_actual=last_year_actual,
                last_year_budget=last_year_budget,
                outstanding_orders=outstanding_orders,
            )
        )

    return results


def _parse_sub_program_pdf_internal(
    pdf_path: Path,
) -> tuple[list[SubProgramLine], str]:
    """Internal implementation shared by the two public PDF parsers.

    Round 61 — switched from token-list heuristics to positional
    column binning (see :func:`_parse_page_positionally`). Each page
    is processed independently; the section (Revenue / Expenditure)
    is detected from the page's banner row.
    """
    if not pdf_path.exists():
        raise ValueError(f"File not found: {pdf_path}")

    lines: list[SubProgramLine] = []
    period_label = ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                raise ValueError(
                    "Sub-Program Report PDF appears empty or unrecognised; "
                    "check the file is a CASES21 GL21157 export"
                )

            section = "Revenue"  # default; updated per-page
            for page in pdf.pages:
                text = page.extract_text() or ""
                # Section banner is a row that says "Revenue Recurrent
                # and Capital" or "Expenditure Recurrent and Capital".
                # The CASES21 PDF never shows both banners on the same
                # page; use elif so a mid-page match for one section
                # banner can't accidentally co-fire the other (defensive
                # against future format changes).
                if _EXPENDITURE_HDR.search(text):
                    section = "Expenditure"
                elif _REVENUE_HDR.search(text):
                    section = "Revenue"

                # Extract period label from footer (first match wins).
                if not period_label:
                    period_label = _extract_period_label(text)

                page_lines = _parse_page_positionally(page, section)
                lines.extend(page_lines)

    except OSError as exc:
        raise ValueError(
            "Sub-Program Report PDF appears empty or unrecognised; "
            "check the file is a CASES21 GL21157 export"
        ) from exc

    if not lines:
        raise ValueError(
            "Sub-Program Report PDF appears empty or unrecognised; "
            "check the file is a CASES21 GL21157 export"
        )

    return lines, period_label


# ---------------------------------------------------------------------------
# Public API -- parse from XLSX (fallback)
# ---------------------------------------------------------------------------


def _get_cell(row: Sequence[Any], idx: int) -> Decimal:
    """Safely extract a Decimal from a row at *idx*."""
    if idx < len(row) and row[idx] is not None:
        raw = str(row[idx]).strip()
        try:
            return parse_decimal(raw)
        except ValueError:
            return Decimal("0")
    return Decimal("0")


def parse_sub_program_xlsx(xlsx_path: Path) -> list[SubProgramLine]:
    """Fallback parser: read a CASES21 XLSX export of the sub-program report.

    Expects columns in order: Sub-Program, Title, [Last year actual],
    [Last year budget], Annual budget, YTD, % Budget [, Outstanding, Uncommitted].
    The first row is treated as a header.
    """
    if not xlsx_path.exists():
        raise ValueError(f"File not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows: list[Any] = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows) < 2:
        raise ValueError(
            "Sub-Program Report XLSX appears empty or unrecognised; "
            "check the file is a CASES21 GL21157 export"
        )

    lines: list[SubProgramLine] = []
    section = "Expenditure"

    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        sub_prog = str(row[0]).strip()
        if not sub_prog.isdigit():
            continue

        title = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        budget = _get_cell(row, 4)
        ytd = _get_cell(row, 5)
        pct = _get_cell(row, 6)
        remaining = budget - ytd
        faculty = _infer_faculty(sub_prog)
        is_over = bool(ytd > budget) if budget != Decimal("0") else False

        lines.append(
            SubProgramLine(
                sub_program=sub_prog,
                account=section,
                description=title,
                budget=budget,
                ytd=ytd,
                remaining=remaining,
                used_pct=pct,
                faculty=faculty,
                is_over=is_over,
            )
        )

    if not lines:
        raise ValueError(
            "Sub-Program Report XLSX appears empty or unrecognised; "
            "check the file is a CASES21 GL21157 export"
        )

    return lines


# ---------------------------------------------------------------------------
# Public API -- prior-period comments
# ---------------------------------------------------------------------------


def load_prior_period_comments(xlsx_path: Path) -> dict[tuple[str, str], str]:
    """Load per-row commentary from a prior-period comments XLSX.

    Reads every worksheet in the workbook (Revenue + Expenditure sheets in
    a typical export both carry comments), and returns
    ``{(sub_program, second_key): commentary_text}``.  The second key is
    either the account code (if an "Account" header is found) or the
    line title / description (if not).  This keeps the join robust against
    files exported by this tool itself, which intentionally drops the raw
    Account column from the published workbook.

    Bug fixed in Round 21
    ---------------------
    Earlier versions silently defaulted to ``txt_col = 2`` when no
    "comments" header was found.  Column 2 in our own export is
    "Last year actual" — a dollar amount — so users saw last-year-actual
    values copied into the new report's comments column.  We now raise a
    descriptive ValueError instead, listing the headers we did find so
    the user can rename the column in their source file.

    Accepted comment-column synonyms: comment, comments, commentary,
    note, notes, remark, remarks, memo.

    Accepted account-column synonyms: account, account code, gl, gl code.
    If none match, the line title / description column is used as the
    secondary join key.

    Phase D round-trip
    ------------------
    Cells written by Round 51+ carry an optional structured prefix of
    the form ``[Driver: X | Outlook: Y | Action: Z] notes``. We return
    the cell verbatim — :func:`decode_commentary` is applied at the
    consumer side (``generate_report``) so that pre-Phase-D files
    (no prefix) still round-trip cleanly as Notes-only.
    """
    if not xlsx_path.exists():
        raise ValueError(f"Comments file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        # Walk EVERY sheet — exports of this tool produce Revenue and
        # Expenditure as separate sheets and both carry comments.
        per_sheet_rows: list[list[list[Any]]] = [
            [list(r) for r in wb[name].iter_rows(values_only=True)] for name in wb.sheetnames
        ]
    finally:
        wb.close()

    result: dict[tuple[str, str], str] = {}

    # Track whether ANY sheet had a usable header.  If none did, we raise
    # at the end — silent fall-through used to copy budget data into the
    # comments column.
    found_any_comment_col = False
    seen_headers: list[str] = []

    for sheet_rows in per_sheet_rows:
        if not sheet_rows:
            continue
        header = [
            str(c).strip().lower().replace("_", " ") if c is not None else "" for c in sheet_rows[0]
        ]
        seen_headers.extend(h for h in header if h)

        # --- Sub-program column -------------------------------------------------
        sp_col: int | None = None
        for i, h in enumerate(header):
            # "sub prog" matches "Sub Prog." and "Sub-Program" (after the
            # underscore-to-space normalisation above).
            if "sub" in h and ("prog" in h or "program" in h):
                sp_col = i
                break
        if sp_col is None:
            sp_col = 0  # safe fallback — first column is almost always sub-program

        # --- Comment column (synonyms) ------------------------------------------
        # Synonyms list is module-scope (``_COMMENT_SYNONYMS``) so that the
        # bottom-of-function ``raise ValueError`` block can reference it
        # even when every sheet was empty / blank — otherwise we'd hit
        # an UnboundLocalError before surfacing the real "no comments
        # column" message to the user.
        txt_col: int | None = None
        for i, h in enumerate(header):
            if any(syn in h for syn in _COMMENT_SYNONYMS):
                txt_col = i
                break
        if txt_col is None:
            # No comment column on THIS sheet — skip it; we'll raise at the
            # bottom only if NO sheet had one.
            continue
        found_any_comment_col = True

        # --- Account column (synonyms) ------------------------------------------
        # Exported workbooks intentionally drop the raw Account column, so
        # if no "account" header is found we fall back to the Title /
        # Description column (col 1).  This matches our own exports
        # without forcing the user to hand-edit headers.
        acc_synonyms = ("account", " gl", "gl code", "code")
        sec_col: int | None = None
        for i, h in enumerate(header):
            if any(syn in h for syn in acc_synonyms):
                sec_col = i
                break
        if sec_col is None:
            # Fall back to Title / Description column — same column used
            # as the second join key when generating the new report.
            sec_col = 1

        # --- Walk rows ----------------------------------------------------------
        for row in sheet_rows[1:]:
            if not row:
                continue
            sp = str(row[sp_col]).strip() if sp_col < len(row) and row[sp_col] is not None else ""
            sec = (
                str(row[sec_col]).strip() if sec_col < len(row) and row[sec_col] is not None else ""
            )
            txt = (
                str(row[txt_col]).strip() if txt_col < len(row) and row[txt_col] is not None else ""
            )
            # Round 1 fix (R51): strip the formula-injection-guard
            # apostrophe written by ``_write_monthly_sub_program_sheet``
            # when an encoded cell happened to start with ``=``/``+``/
            # ``-``/``@`` (Excel's formula sigils). Excel renders the
            # apostrophe as a "force text" marker but openpyxl returns
            # it verbatim — we strip it back so :func:`decode_commentary`
            # sees the original encoded value and round-trips cleanly
            # without accumulating apostrophes across save→reopen→save
            # cycles.
            if len(txt) >= 2 and txt[0] == "'" and txt[1] in ("=", "+", "-", "@"):
                txt = txt[1:]
            # Skip rows that are clearly not data — sub-program codes are
            # purely numeric.  This stops "Total" / blank header artefacts
            # from polluting the dict.
            if not sp or not sp.isdigit():
                continue
            if sp or sec:
                # If two sheets disagree on the same key, the later one
                # wins — matches the visual reading order in the workbook.
                result[(sp, sec)] = txt

    if not found_any_comment_col:
        seen = ", ".join(sorted(set(h for h in seen_headers if h))) or "(none)"
        raise ValueError(
            "Could not find a comments column in the prior-period file. "
            "Looked for any header containing one of: "
            f"{', '.join(_COMMENT_SYNONYMS)}. "
            f"Headers we found: {seen}. "
            "Rename your comments column to 'Comments' (or one of the "
            "synonyms above) and try again."
        )

    return result


# ---------------------------------------------------------------------------
# XLSX output writer
# ---------------------------------------------------------------------------

# Excel Accounting format. Round 58 added [Red] colouring to the
# negative branch so negative numbers (over-spends, deficits) print
# in red and read at-a-glance as "below zero". The pre-R58 format
# matched the Jan26 reference file exactly but was monochrome — a
# council reader had to scan the leading minus to spot a deficit.
# Format is 4 sections: positive ; negative ; zero ; text.
_ACCOUNTING_FMT = '_-"$"* #,##0_-;[Red]\\-"$"* #,##0_-;_-"$"* "-"??_-;_-@_-'
_PERCENT_FMT = "0.00"
# Round 47 — proper percent format for the new Monthly shape's
# % columns. Cells store the value as a fraction (0.398...) and
# Excel renders it as "39.8%". Round 58 added [Red] colouring to
# the negative branch.
_PERCENT_AS_PERCENT_FMT = "0.0%;[Red]-0.0%"
_TITLE_FONT = Font(bold=True, size=14)
# Green data bar colour -- matches Jan26 reference (#63C384 with full alpha).
_DATA_BAR_COLOR = "FF63C384"

# Revenue sheet: 8 columns -- no Outstanding Orders column.
_REV_HEADERS = [
    "Sub Prog.",
    "Title",
    "Last year actual",
    "Last year budget",
    "Annual budget",
    "YTD",
    "% Budget received",
    "Comments",
]
_REV_WIDTHS = [10, 43, 13, 14, 17, 15, 13, 60]

# Expenditure sheet: 9 columns -- Outstanding Orders is parsed from the PDF
# and used in the Uncommitted Balance computation (Annual - YTD - Outstanding)
# but NOT displayed as its own column, per user's Q3 spec direction.
_EXP_HEADERS = [
    "Sub Prog.",
    "Title",
    "Last year actual",
    "Last year budget",
    "Annual budget",
    "YTD",
    "% Budget Expended",
    "Uncommitted Balance",
    "Comments",
]
_EXP_WIDTHS = [10, 43, 13, 14, 17, 15, 14, 18, 60]


def _recompute_is_over(
    lines: list[SubProgramLine],
    threshold: float,
    *,
    revenue_threshold: float | None = None,
    expense_threshold: float | None = None,
    materiality_dollar: int = 100,
) -> list[SubProgramLine]:
    """Return new SubProgramLine list with is_over + variance + materiality
    recomputed.

    Round 21 added optional ``revenue_threshold`` / ``expense_threshold``
    keyword-only params: a Revenue line is over-budget if
    ``used_pct > revenue_threshold``, Expenditure if
    ``used_pct > expense_threshold``.

    Round 45 Phase A also (re)computes:

    * ``variance_amount = ytd - budget`` — signed.
    * ``variance_pct = variance_amount / budget * 100`` — signed.
    * ``is_material`` — True when ``abs(variance_amount) >= materiality_dollar``.

    Round 56 dropped the ``calendar_pct`` parameter and the ``pacing``
    field along with all pacing-based judgements.
    """
    from dataclasses import replace as _replace

    rev_th = revenue_threshold if revenue_threshold is not None else threshold
    exp_th = expense_threshold if expense_threshold is not None else threshold
    mat = Decimal(str(materiality_dollar))

    result: list[SubProgramLine] = []
    for ln in lines:
        is_revenue = ln.account.lower().startswith("revenue")
        section_th = rev_th if is_revenue else exp_th
        # Round 47 — a sub-program with $0 budget but $X spent has
        # used_pct = 0 from the parser (division-by-zero defended) but
        # is unambiguously over budget. Flag those rows directly so
        # they don't slip past the percentage gate.
        # Round 71 O3: extend the unbudgeted-spend guard to cover
        # negative-budget rows too. CASES21 doesn't typically emit
        # negative budgets, but the pre-R71 ``== 0`` check would have
        # silently skipped them and let a row with budget=-1000 and
        # ytd=500 fall through to "is_over=False". The new ``<= 0``
        # check fires on any row where there's no positive budget
        # backing the spend.
        unbudgeted_with_spend = ln.budget <= Decimal("0") and ln.ytd != Decimal("0")
        new_is_over = float(ln.used_pct) > section_th or unbudgeted_with_spend

        new_variance_amount = ln.ytd - ln.budget
        new_variance_pct = (
            (new_variance_amount / ln.budget * Decimal("100"))
            if ln.budget != Decimal("0")
            else Decimal("0")
        )
        new_is_material = abs(new_variance_amount) >= mat

        ln = _replace(
            ln,
            is_over=new_is_over,
            variance_amount=new_variance_amount,
            variance_pct=new_variance_pct,
            is_material=new_is_material,
        )
        result.append(ln)
    return result


def _write_sheet(
    ws: Any,
    title: str,
    headers: list[str],
    widths: list[int],
    lines: list[SubProgramLine],
    is_revenue: bool,
) -> None:
    """Populate a single Revenue or Expenditure worksheet."""
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)
    n_cols = len(headers)

    # Row 1: merged title -- bold, size 14, centred.
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Row 2: column headers.
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header).alignment = Alignment(
            horizontal="left", wrap_text=True
        )

    # Freeze panes so title + header row both stay visible.
    ws.freeze_panes = "A3"

    # Column widths.
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows start at row 3.
    percent_col = 7
    for row_idx, line in enumerate(lines, start=3):
        # Round 51 Phase D — encode structured triplet + notes into the
        # single Comments cell. Cells where all four fields are blank
        # come out as empty string (matches pre-Phase-D behaviour).
        encoded_comment = encode_commentary(
            line.commentary,
            driver=line.commentary_driver,
            outlook=line.commentary_outlook,
            action=line.commentary_action,
        )
        if is_revenue:
            row_values: list[Any] = [
                line.sub_program,
                line.description,
                float(line.last_year_actual),
                float(line.last_year_budget),
                float(line.budget),
                float(line.ytd),
                float(line.used_pct),
                encoded_comment,
            ]
            # Accounting format on Annual budget (col 5) and YTD (col 6).
            currency_cols = {5, 6}
        else:
            # Uncommitted Balance = Annual budget - YTD - Outstanding Orders.
            # Outstanding Orders is parsed from the PDF but not surfaced as a
            # column (per user spec Q3); it only contributes to this derivation.
            uncommitted = line.budget - line.ytd - line.outstanding_orders
            row_values = [
                line.sub_program,
                line.description,
                float(line.last_year_actual),
                float(line.last_year_budget),
                float(line.budget),
                float(line.ytd),
                float(line.used_pct),
                float(uncommitted),
                encoded_comment,
            ]
            # Accounting format on Annual budget (col 5), YTD (col 6).
            currency_cols = {5, 6}

        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in currency_cols:
                c.number_format = _ACCOUNTING_FMT
            elif col_idx == percent_col:
                c.number_format = _PERCENT_FMT
            # Pink row fill for over-budget rows (threshold-aware is_over).
            if line.is_over:
                c.fill = _OVER_FILL

    # Conditional formatting on the % Budget column (G = col 7).
    if lines:
        last_data_row = 2 + len(lines)
        pct_col_letter = get_column_letter(percent_col)
        rng = f"{pct_col_letter}3:{pct_col_letter}{last_data_row}"

        # Data bar: green, 0--110 (matches Revenue sheet in Jan26 reference).
        ws.conditional_formatting.add(
            rng,
            DataBarRule(  # type: ignore[no-untyped-call]
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=110,
                color=_DATA_BAR_COLOR,
                showValue=True,
            ),
        )


def _sheet_title(base: str, period_label: str, suffix: str) -> str:
    """Compose a sheet title, gracefully omitting the period when absent."""
    if period_label:
        return f"{base} - {period_label} {suffix}"
    return f"{base} - {suffix}"


def _write_xlsx(
    lines: list[SubProgramLine],
    output_file: Path,
    period_label: str = "",
    over_budget_threshold: float = 101.0,
    *,
    include_combined: bool = False,
    materiality_dollar: int = 100,
    prior_ytd: dict[str, Decimal] | None = None,
    prior_funds: dict[str, Decimal] | None = None,
    revenue_threshold: float | None = None,
    expense_threshold: float | None = None,
) -> None:
    """Write the report to an XLSX with the Monthly Sub Program Report shape.

    Round 38 — replaces the prior 2-sheet (Revenue / Expenditure) layout
    with a single sheet matching the school's own Monthly Sub Program
    Report workbook.

    F2 (Round 54): the workbook contains TWO sheets:
    1. ``Sub Program Report`` — the per-sub-program detail. Round 57
       collapsed the F2 Trend column, leaving 13 cols total
       (Status at col 3, financials at cols 4..13).
    2. ``Watchlist`` — a filtered subset of sub-programs whose Status
       is not "On track", sorted by absolute variance descending.
       Council-targeted view.

    Round 57 dropped the Trend column entirely. The ``prior_ytd``
    parameter is still accepted for backward compatibility with the
    existing call sites + tests but no longer affects the output.

    Round 57 added ``prior_funds`` — when supplied, the writer
    populates the "Funds from Previous Years (Funds)" column from
    the prior-period XLSX so the carry-forward field rolls forward
    automatically.

    Round 62 added ``revenue_threshold`` / ``expense_threshold``. The
    Status pill now flags Revenue-over-budget rows (was Expense-only
    pre-R62), so the writer needs the per-section thresholds to
    compute pills consistently with the in-app view. Both default to
    ``over_budget_threshold`` for backward compatibility — existing
    callers that pass only the combined threshold keep working.

    The ``include_combined`` argument is accepted for backward
    compatibility with existing call sites but no longer affects the
    output.
    """
    if revenue_threshold is None:
        revenue_threshold = over_budget_threshold
    if expense_threshold is None:
        expense_threshold = over_budget_threshold
    del include_combined, prior_ytd  # no-op for the new shape
    del over_budget_threshold  # superseded by per-section thresholds above

    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    ws = wb.create_sheet("Sub Program Report")
    _write_monthly_sub_program_sheet(
        ws,
        lines,
        period_label,
        materiality_dollar,
        prior_funds=prior_funds,
        revenue_threshold=revenue_threshold,
        expense_threshold=expense_threshold,
        # pass-through is handled by the function signature
    )

    # F2: Watchlist sheet — filtered subset for council.
    watchlist_ws = wb.create_sheet("Watchlist")
    _write_watchlist_sheet(
        watchlist_ws,
        lines,
        period_label,
        materiality_dollar,
        prior_funds=prior_funds,
        revenue_threshold=revenue_threshold,
        expense_threshold=expense_threshold,
    )

    # Round 64 — restored the legacy Revenue + Expenditure detail
    # sheets per user request. Round 38 collapsed both into the single
    # "Sub Program Report" tab, but a user who wants to drill into a
    # specific account section (e.g. "all revenue lines in one place
    # for an auditor's read") was left without that view. The two
    # sheets carry the per-Account-row data with the green % data bar
    # and pink-fill for over-budget rows, matching the Jan26 reference
    # workbook's shape exactly.
    revenue_lines = [ln for ln in lines if ln.account.lower().startswith("revenue")]
    if revenue_lines:
        rev_ws = wb.create_sheet("Revenue")
        rev_title = _sheet_title("Annual Sub-Program Budget Report", period_label, "Revenue")
        _write_sheet(rev_ws, rev_title, _REV_HEADERS, _REV_WIDTHS, revenue_lines, is_revenue=True)
    expense_lines = [ln for ln in lines if ln.account.lower().startswith("expenditure")]
    if expense_lines:
        exp_ws = wb.create_sheet("Expenditure")
        exp_title = _sheet_title("Annual Sub-Program Budget Report", period_label, "Expenditure")
        _write_sheet(exp_ws, exp_title, _EXP_HEADERS, _EXP_WIDTHS, expense_lines, is_revenue=False)

    # R1 fix: pin the main sheet as active so Excel opens to it by
    # default. Without this, openpyxl may surface Watchlist first
    # (which is sorted/filtered and looks like a partial export).
    wb.active = wb.index(ws)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)

    # Round 70 — post-save: inject the Excel x14 dataBar extension on
    # any percent column whose data has a negative value. openpyxl's
    # basic v1 dataBar always grows from the cell's LEFT edge so the
    # Round 68 "axis at 1/10 from left" design didn't render in
    # Excel — negatives just appeared as nothing. The x14 extension
    # adds proper axisPosition + negativeFillColor support, so
    # negatives now render leftward from the axis in red while
    # positives render rightward from the axis in green.
    main_sheet_lines = lines
    watchlist_lines = [
        ln
        for ln in lines
        if (ln.is_over and abs(ln.ytd - ln.budget) >= Decimal(str(materiality_dollar)))
    ]
    _inject_x14_data_bars_for_sheet(
        output_file,
        sheet_xml_basename="sheet1.xml",  # Sub Program Report
        lines=main_sheet_lines,
    )
    if watchlist_lines:
        # The Watchlist sheet has its own (filtered) line set; recompute
        # the per-column negative-presence over that subset so the x14
        # extension is added only where the column actually contains a
        # negative value.
        _inject_x14_data_bars_for_sheet(
            output_file,
            sheet_xml_basename="sheet2.xml",  # Watchlist
            lines=watchlist_lines,
        )


# ---------------------------------------------------------------------------
# Round 70 — x14 dataBar extension (axis at 1/10 from left for mixed-sign
# columns). openpyxl's basic v1 dataBar always grows bars from the cell's
# left edge, so a positive 50% renders as half-width-from-left and a
# negative -50% renders as zero-width — no axis, no leftward bar. The
# Excel 2010+ x14 extension supports axisPosition + negativeFillColor for
# proper axis-aware rendering; openpyxl's high-level API doesn't expose
# it so we inject the XML directly on the saved file.
# ---------------------------------------------------------------------------

_X14_DATABAR_REF_URI = "{B025F937-C7B1-47D3-B67F-A62EFF666E3E}"
_X14_CF_URI = "{78C0D931-6437-407d-A8EE-F0AAD7539E65}"
_X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_XM_NS = "http://schemas.microsoft.com/office/excel/2006/main"
# Red palette for the negative-bar fill: a softer rose (not alarm-red)
# so the negative bar reads as "below zero" not "system error".
_DATA_BAR_NEG_COLOR = "FFD46A6A"


def _has_negative_avail_pct(lines: list[SubProgramLine]) -> bool:
    """Return True if any sub-program's (Annual Exp Budget − Exp YTD −
    Outstanding Orders) is negative — i.e. col L renders < 0 and col C
    (Available Balance %) shows a negative percent."""
    exp_b: dict[str, Decimal] = {}
    exp_y: dict[str, Decimal] = {}
    orders: dict[str, Decimal] = {}
    for ln in lines:
        if not ln.account.lower().startswith("expenditure"):
            continue
        sp = ln.sub_program
        exp_b[sp] = exp_b.get(sp, Decimal("0")) + ln.budget
        exp_y[sp] = exp_y.get(sp, Decimal("0")) + ln.ytd
        if ln.outstanding_orders:
            orders[sp] = orders.get(sp, Decimal("0")) + ln.outstanding_orders
    for sp, eb in exp_b.items():
        if eb <= 0:
            continue
        ey = exp_y.get(sp, Decimal("0"))
        oo = orders.get(sp, Decimal("0"))
        if (eb - ey - oo) < 0:
            return True
    return False


def _has_negative_rev_pct(lines: list[SubProgramLine]) -> bool:
    """Return True if any sub-program has Revenue YTD < 0 against a
    non-zero Revenue budget — i.e. col D (Revenue Budget % Received)
    shows a negative percent."""
    rev_b: dict[str, Decimal] = {}
    rev_y: dict[str, Decimal] = {}
    for ln in lines:
        if not ln.account.lower().startswith("revenue"):
            continue
        sp = ln.sub_program
        rev_b[sp] = rev_b.get(sp, Decimal("0")) + ln.budget
        rev_y[sp] = rev_y.get(sp, Decimal("0")) + ln.ytd
    for sp, rb in rev_b.items():
        if rb <= 0:
            continue
        if rev_y.get(sp, Decimal("0")) < 0:
            return True
    return False


def _inject_x14_data_bars_for_sheet(
    xlsx_path: Path,
    *,
    sheet_xml_basename: str,
    lines: list[SubProgramLine],
) -> None:
    """Inject Excel x14 dataBar extensions on cols C (Available Balance
    %) and D (Revenue Budget % Received) when the relevant column
    has at least one negative value.

    Re-opens the saved XLSX zip, modifies the sheet XML to:
      1. Append an ``<extLst>`` inside each target ``<cfRule>`` with a
         unique GUID pointing to the x14 dataBar definition.
      2. Add a worksheet-level ``<extLst>`` carrying the x14
         conditional formatting blocks themselves (axisPosition="automatic",
         negativeFillColor, etc.).
      3. Add ``x14`` to the worksheet root's ``mc:Ignorable`` so old
         Excel versions silently skip the extension.

    Sheets that have all-positive data in both columns get no x14
    injection — the basic v1 dataBar already renders them correctly.
    """
    import re
    import uuid
    import zipfile

    has_neg_c = _has_negative_avail_pct(lines)
    has_neg_d = _has_negative_rev_pct(lines)
    if not (has_neg_c or has_neg_d):
        # Pure positive data on both columns; basic dataBar suffices.
        return

    sheet_path = f"xl/worksheets/{sheet_xml_basename}"

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        entries: dict[str, bytes] = {n: zin.read(n) for n in zin.namelist()}

    if sheet_path not in entries:
        return

    xml = entries[sheet_path].decode("utf-8")

    targets: list[tuple[str, bool]] = []
    if has_neg_c:
        targets.append(("C", True))
    if has_neg_d:
        targets.append(("D", True))

    x14_blocks: list[str] = []
    for col_letter, _ in targets:
        bar_id = "{" + str(uuid.uuid4()).upper() + "}"
        # Find the matching <conditionalFormatting sqref="C3:..."> +
        # <cfRule type="dataBar" ...><dataBar>...</dataBar> and insert
        # <extLst> before </cfRule>.
        pattern = re.compile(
            r'(<conditionalFormatting sqref="'
            + re.escape(col_letter)
            + r"\d+(?::"
            + re.escape(col_letter)
            + r'\d+)?"[^>]*>\s*'
            + r'<cfRule[^>]*type="dataBar"[^>]*>\s*'
            + r"<dataBar[^>]*>.*?</dataBar>\s*"
            + r")(</cfRule>)",
            re.DOTALL,
        )
        ext_inject = (
            "<extLst>"
            f'<ext xmlns:x14="{_X14_NS}" uri="{_X14_DATABAR_REF_URI}">'
            f"<x14:id>{bar_id}</x14:id>"
            "</ext>"
            "</extLst>"
        )
        xml, n_subs = pattern.subn(rf"\1{ext_inject}\2", xml, count=1)
        if n_subs == 0:
            continue

        # Extract the actual sqref from the modified XML so we mirror it
        # in the x14 block (handles e.g. "C3:C92" or single "C3").
        sqref_match = re.search(
            r'<conditionalFormatting sqref="('
            + re.escape(col_letter)
            + r"\d+(?::"
            + re.escape(col_letter)
            + r'\d+)?)"',
            xml,
        )
        if sqref_match is None:
            continue
        sqref = sqref_match.group(1)

        x14_blocks.append(
            f'<x14:conditionalFormatting xmlns:xm="{_XM_NS}">'
            f'<x14:cfRule type="dataBar" id="{bar_id}">'
            '<x14:dataBar minLength="0" maxLength="100" '
            'border="1" negativeBarColorSameAsPositive="0" '
            'axisPosition="automatic">'
            f'<x14:cfvo type="num"><xm:f>{-1.0 / 9.0}</xm:f></x14:cfvo>'
            '<x14:cfvo type="num"><xm:f>1</xm:f></x14:cfvo>'
            f'<x14:borderColor rgb="{_DATA_BAR_COLOR}"/>'
            f'<x14:negativeFillColor rgb="{_DATA_BAR_NEG_COLOR}"/>'
            f'<x14:negativeBorderColor rgb="{_DATA_BAR_NEG_COLOR}"/>'
            '<x14:axisColor rgb="FF000000"/>'
            "</x14:dataBar>"
            "</x14:cfRule>"
            f"<xm:sqref>{sqref}</xm:sqref>"
            "</x14:conditionalFormatting>"
        )

    if not x14_blocks:
        # All target injections failed to find their cfRule — skip.
        entries[sheet_path] = xml.encode("utf-8")
        return

    # Append the worksheet-level <extLst> with all x14 blocks. Inserts
    # right before </worksheet>.
    worksheet_ext = (
        "<extLst>"
        f'<ext uri="{_X14_CF_URI}" xmlns:x14="{_X14_NS}">'
        "<x14:conditionalFormattings>"
        f"{''.join(x14_blocks)}"
        "</x14:conditionalFormattings>"
        "</ext>"
        "</extLst>"
    )
    xml = xml.replace("</worksheet>", f"{worksheet_ext}</worksheet>", 1)

    entries[sheet_path] = xml.encode("utf-8")

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def _write_monthly_sub_program_sheet(
    ws: Any,
    lines: list[SubProgramLine],
    period_label: str,
    materiality_dollar: int = 100,
    *,
    prior_funds: dict[str, Decimal] | None = None,
    filter_to_non_ok: bool = False,
    sort_by_variance_desc: bool = False,
    sheet_title_override: str | None = None,
    revenue_threshold: float = 101.0,
    expense_threshold: float = 101.0,
) -> None:
    """Round 57 — Monthly Sub Program Report shape (13-col).

    Per-sub-program output matching the school's own "Monthly Sub
    Program Report" workbook. Round 66 moved the two percent columns
    between PROGRAM NAME and Status so the at-a-glance bars sit
    immediately next to the program name:

    1.  CODE                                 sub_program code
    2.  PROGRAM NAME                         description
    3.  Available Balance % YTD              =L{r}/H{r}
    4.  Revenue Budget % Received YTD        =I{r}/G{r}
    5.  Status                                computed pill
    6.  Funds from Previous Years (Funds)    carry-forward (from prior file)
    7.  Budget Revenue {year}                annual revenue budget
    8.  Total Budget Allocation Expenditure  annual expenditure budget
    9.  Revenue YTD                          revenue collected so far
    10. Expenditure YTD                      expenditure spent so far
    11. Less outstanding orders              committed-but-not-paid
    12. Available Balance YTD                =H{r}-J{r}-K{r}
    13. Comments                             commentary

    Round 57: the three derived numeric columns (Available Balance YTD,
    Available Balance %, Revenue Budget % Received) are written as
    Excel formulas so a school auditor can see HOW each number is
    derived. Round 67 dropped the ±999% display cap — the actual
    percentage renders, however large, so a 2,021% revenue over-
    collection reads as "2021.3%" instead of being collapsed to
    ">999%". The 0–100% data bar still caps visually at 100% via
    Excel's standard clipping, so the bar + number together convey
    "extreme over-collection, scale 2,021%" at a glance.

    Rows whose Status pill is non-OK get the pink ``_OVER_FILL``
    (HL_MISMATCH) across all 13 cells — the canonical "needs
    attention" indicator across the toolkit. Round 58 changed the
    pink-fill condition from the raw ``available < 0`` signal to
    ``status != "On track"`` so the fill never contradicts the pill.

    The "Funds from Previous Years" column is populated from the
    optional ``prior_funds`` dict (loaded via
    :func:`load_prior_period_funds`). When no prior file is supplied
    or a sub-program isn't found in the prior data, the cell stays
    blank — a blank cell signals "not known", which is honest; a
    zero would be a wrong number.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)

    # ----- Aggregate per-sub-program -----
    rev_b: dict[str, Decimal] = {}
    exp_b: dict[str, Decimal] = {}
    rev_y: dict[str, Decimal] = {}
    exp_y: dict[str, Decimal] = {}
    orders: dict[str, Decimal] = {}
    desc: dict[str, str] = {}
    # Round 60 — track which sub-programs have ANY constituent line
    # (Revenue OR Expenditure) flagged as is_over+is_material. The
    # in-app Watchlist filters on this exact signal per-line; the
    # XLSX writer aggregates that to "is THIS sub-program on the
    # Watchlist?" so the pink fill + Watchlist sheet match the
    # in-app view exactly.
    sps_on_watchlist: set[str] = set()
    # Round 51 Phase D Round-1 fix — atomic per-sub-program commentary
    # aggregation. Pre-fix the four fields (notes / driver / outlook /
    # action) were each independently "first non-empty wins" across
    # Account-rows of the same sub-program — which produced
    # cross-row data fabrication (Action from row B alongside Notes
    # from row A). Now we adopt the WHOLE 4-tuple from the first row
    # that contributes ANY commentary content, keeping the
    # categorisation atomic and traceable to a single source row.
    commentary_tuple: dict[str, tuple[str, str, str, str]] = {}

    for ln in lines:
        sp = ln.sub_program
        kind = ln.account.lower()
        if kind.startswith("revenue"):
            rev_b[sp] = rev_b.get(sp, Decimal("0")) + ln.budget
            rev_y[sp] = rev_y.get(sp, Decimal("0")) + ln.ytd
        elif kind.startswith("expenditure"):
            exp_b[sp] = exp_b.get(sp, Decimal("0")) + ln.budget
            exp_y[sp] = exp_y.get(sp, Decimal("0")) + ln.ytd
            # Outstanding orders are an expenditure-side concept; aggregate
            # across all expenditure account-rows for the sub-program.
            if ln.outstanding_orders:
                orders[sp] = orders.get(sp, Decimal("0")) + ln.outstanding_orders
        # Round 60 — match in-app Watchlist filter: this sub-program
        # is on the Watchlist if ANY of its lines is over+material.
        # Recompute is_material in-place so synthetic callers (tests,
        # direct _write_xlsx invocations) don't need to call
        # _recompute_is_over first to populate the flag — the writer
        # is self-contained against materiality_dollar.
        is_material_now = abs(ln.ytd - ln.budget) >= Decimal(str(materiality_dollar))
        if ln.is_over and is_material_now:
            sps_on_watchlist.add(sp)
        # First non-empty description / commentary wins (consistent
        # with how the in-app Combined view picks its description).
        if ln.description and sp not in desc:
            desc[sp] = ln.description
        # Atomic 4-tuple adoption: the row that wins is the first one
        # that contributes ANY commentary content for the sub-program.
        if sp not in commentary_tuple and (
            ln.commentary or ln.commentary_driver or ln.commentary_outlook or ln.commentary_action
        ):
            commentary_tuple[sp] = (
                ln.commentary,
                ln.commentary_driver,
                ln.commentary_outlook,
                ln.commentary_action,
            )

    sub_programs = sorted(set(rev_b.keys()) | set(exp_b.keys()))

    # Round 60 — Watchlist sheet filter follows the in-app Watchlist
    # tab exactly. The in-app filters on per-line ``is_over and
    # is_material``; ``sps_on_watchlist`` (built above) is the
    # aggregated set of sub-programs with at least one such line.
    # Pre-R60 the filter used ``status != "On track"`` which checks
    # only the Expenditure side and missed Revenue-over-budget rows.
    if filter_to_non_ok or sort_by_variance_desc:
        # Round 63 — per-sub-program max |variance_amount| across the
        # sub-program's FLAGGED lines (is_over AND is_material). Used
        # below as the sort key so the XLSX Watchlist order matches
        # the in-app Watchlist tab, which sorts per-line by
        # ``-abs(variance_amount)``. Restricting to flagged lines is
        # essential: each in-app row IS a flagged line, so the in-app's
        # |variance| is the flagged-line variance — not the max across
        # both Revenue + Expenditure sides. Without this restriction
        # an over-collected Revenue line ($2K over) on a sub-program
        # with a large under-spent Expenditure side ($30K unspent)
        # would sort by the unspent $30K, putting the row much higher
        # than the in-app does. The same in-place is_material recompute
        # used for ``sps_on_watchlist`` (above) gives the writer self-
        # contained behaviour against synthetic test callers.
        sp_max_variance: dict[str, Decimal] = {}
        for ln in lines:
            is_material_now = abs(ln.ytd - ln.budget) >= Decimal(str(materiality_dollar))
            if not (ln.is_over and is_material_now):
                continue
            sp_max_variance[ln.sub_program] = max(
                sp_max_variance.get(ln.sub_program, Decimal("0")),
                abs(ln.variance_amount),
            )

        annotated: list[tuple[str, Decimal, str]] = []  # (sp, max_variance, status)
        for sp in sub_programs:
            ry = rev_y.get(sp, Decimal("0"))
            ey = exp_y.get(sp, Decimal("0"))
            rb = rev_b.get(sp, Decimal("0"))
            eb = exp_b.get(sp, Decimal("0"))
            status = compute_status_pill(
                annual_exp_budget=eb,
                exp_ytd=ey,
                annual_rev_budget=rb,
                rev_ytd=ry,
                revenue_threshold=revenue_threshold,
                expense_threshold=expense_threshold,
                materiality_dollar=materiality_dollar,
            )
            annotated.append((sp, sp_max_variance.get(sp, Decimal("0")), status))
        if filter_to_non_ok:
            annotated = [t for t in annotated if t[0] in sps_on_watchlist]
        if sort_by_variance_desc:
            # Round 63 — sort by |max variance| descending, tied-break
            # by sub_program ascending. Mirrors the in-app Watchlist
            # tab key ``(-abs(ln.variance_amount), ln.sub_program)``.
            annotated.sort(key=lambda t: (-t[1], t[0]))
        sub_programs = [t[0] for t in annotated]

    # ----- Title row (merged across 13 cols in R57) -----
    base = sheet_title_override or "Monthly Sub Program Report"
    title = _sheet_title(base, period_label, "")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    # R57: title spans 13 cols (Trend column dropped — Status at col 3,
    # Funds from Previous Years at col 4, financials at cols 5..13).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)

    # ----- Header row -----
    # Year label for the budget headers — extracted from period_label
    # (e.g. "Apr 2026" → "2026"); falls back to a blank year suffix.
    year_label = ""
    if period_label:
        for tok in period_label.split():
            if tok.isdigit() and len(tok) == 4:
                year_label = tok
                break
    rev_budget_header = f"Budget Revenue {year_label}".strip()
    exp_budget_header = f"Total Budget Allocation Expenditure {year_label}".strip()

    headers = [
        # Round 66 layout: the two percent columns (Available Balance %
        # YTD and Revenue Budget % Received YTD) moved between PROGRAM
        # NAME and Status. Council reader scans left-to-right and the
        # at-a-glance percent bars now sit immediately next to the
        # program name — they're the headline signal, even before the
        # Status pill. Status still leads the dollar columns; the
        # ordering of Funds / Budgets / YTDs / Orders / Avail $ /
        # Comments is unchanged.
        "CODE",  # A
        "PROGRAM NAME",  # B
        "Available Balance % YTD",  # C  (moved from old K)
        "Revenue Budget % Received YTD",  # D  (moved from old L)
        "Status",  # E  (was C)
        "Funds from Previous Years (Funds) ",  # F  (was D)
        rev_budget_header,  # G  (was E)
        exp_budget_header,  # H  (was F)
        "Revenue YTD",  # I  (was G)
        "Expenditure YTD",  # J  (was H)
        "Less outstanding orders",  # K  (was I)
        "Available Balance YTD",  # L  (was J)
        "Comments",  # M  (unchanged)
    ]
    # Round 66 — width slots reshuffled to match. Percent cols (12, 14)
    # are narrower than the now-rightmost Avail Balance YTD ($, 16);
    # Status (22) keeps its old slot. Total still ~230 units.
    widths = [8, 32, 12, 14, 22, 14, 16, 22, 14, 14, 14, 16, 32]
    for col_idx, (header, width) in enumerate(zip(headers, widths, strict=True), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ----- Data rows -----
    # Round 66 column layout (percent columns moved to C/D):
    #   A = CODE
    #   B = PROGRAM NAME
    #   C = Available Balance % YTD     (=L/H)
    #   D = Revenue Budget % Received   (=I/G)
    #   E = Status pill                 (was C)
    #   F = Funds from Previous Years    (was D)
    #   G = Budget Revenue               (was E)
    #   H = Budget Expenditure           (was F)
    #   I = Revenue YTD                  (was G)
    #   J = Expenditure YTD              (was H)
    #   K = Less outstanding orders      (was I)
    #   L = Available Balance YTD        =H-J-K  (was J, =F-H-I)
    #   M = Comments                     (unchanged)
    for row_idx, sp in enumerate(sub_programs, start=3):
        ry = rev_y.get(sp, Decimal("0"))
        ey = exp_y.get(sp, Decimal("0"))
        rb = rev_b.get(sp, Decimal("0"))
        eb = exp_b.get(sp, Decimal("0"))
        oo = orders.get(sp, Decimal("0"))
        # Round 57 — carry-forward is now populated from ``prior_funds``
        # when the user supplied a prior-period XLSX. When no prior
        # file or this sub-program isn't in it, the cell stays blank
        # (a blank cell signals "not known", which is honest; a zero
        # would be a wrong number).
        carry_fwd: Decimal | None = prior_funds.get(sp) if prior_funds is not None else None
        cf_for_calc = carry_fwd if carry_fwd is not None else Decimal("0")
        # Python-side available used only for status pill + pink fill;
        # the cell itself uses an Excel formula so the user can audit
        # the math.
        available = cf_for_calc + ry - ey - oo

        ws.cell(row=row_idx, column=1, value=_to_int_or_str(sp))
        ws.cell(row=row_idx, column=2, value=desc.get(sp, ""))
        # Status (col 5) is populated below — we write it after
        # computing the pill so the cell write happens in one place.

        # Funds from Previous Years (col 6/F) — value when known, blank
        # when not. carry_fwd is None when no prior file supplied or
        # this sub-program wasn't in it.
        if carry_fwd is not None:
            c = ws.cell(row=row_idx, column=6, value=float(carry_fwd))
            c.number_format = _ACCOUNTING_FMT

        # Cols 7..11 are raw inputs (parsed from the PDF). Col 12 is
        # the Available Balance YTD formula; cols 3 & 4 are the
        # percent formulas — both rendered to the left of Status so
        # the council reader sees the at-a-glance bars first.
        for col_idx, value in (
            (7, rb),  # Budget Revenue
            (8, eb),  # Budget Expenditure
            (9, ry),  # Revenue YTD
            (10, ey),  # Expenditure YTD
            (11, oo),  # Outstanding orders
        ):
            c = ws.cell(row=row_idx, column=col_idx, value=float(value))
            c.number_format = _ACCOUNTING_FMT

        # Round 65 — Available Balance YTD (col 12/L) formula:
        #   = Annual Expenditure Budget − Expenditure YTD − Outstanding Orders
        # i.e. "how much expenditure budget is left to spend after
        # committed orders settle". Mirrors the KMAR reference
        # workbook. Round 66: col letters shifted by 2 (H/J/K were
        # F/H/I); formula structure unchanged.
        avail_formula = f"=H{row_idx}-J{row_idx}-K{row_idx}"
        avail_cell = ws.cell(row=row_idx, column=12, value=avail_formula)
        avail_cell.number_format = _ACCOUNTING_FMT

        # Round 67 — always write the Excel formula and let Excel
        # display the actual percentage, however large. Pre-R67 the
        # writer capped values outside ±999% to text markers
        # (">999%" / "<-999%") with a cell comment carrying the real
        # number — a 2,021% revenue over-collection rendered as
        # ">999%" which (per user feedback on D27) hid the actual
        # magnitude. The data bar (col C / col D, R65/R66) already
        # caps visually at 100% via Excel's standard clipping, so a
        # raw "2021.3%" number alongside a saturated bar gives the
        # full information.
        def _write_percent_formula(
            r: int,
            cell_col: int,
            raw_pct: Decimal | None,
            formula: str,
        ) -> None:
            if raw_pct is None:
                # Divisor is zero — leave blank rather than #DIV/0!.
                return
            cell = ws.cell(row=r, column=cell_col, value=formula)
            cell.number_format = _PERCENT_AS_PERCENT_FMT

        avail_pct: Decimal | None = available / eb if eb != 0 else None
        rev_pct: Decimal | None = ry / rb if rb != 0 else None
        # Available Balance % formula: =L{r}/H{r}  (Avail YTD ÷ Budget Exp)
        _write_percent_formula(row_idx, 3, avail_pct, f"=L{row_idx}/H{row_idx}")
        # Revenue Budget % Received formula: =I{r}/G{r}  (Rev YTD ÷ Budget Rev)
        _write_percent_formula(row_idx, 4, rev_pct, f"=I{row_idx}/G{row_idx}")

        # Round 53 F1 (Move B) — Status pill at col 5 (was col 3
        # pre-R66; the two percent columns moved before it). Round 56:
        # pacing-free contract; compute_status_pill gates on
        # ``exp_ytd > expense_threshold% × annual_exp_budget``.
        # Round 62 — Revenue-side check added.
        status = compute_status_pill(
            annual_exp_budget=eb,
            exp_ytd=ey,
            annual_rev_budget=rb,
            rev_ytd=ry,
            revenue_threshold=revenue_threshold,
            expense_threshold=expense_threshold,
            materiality_dollar=materiality_dollar,
        )
        status_cell = ws.cell(row=row_idx, column=5, value=status)
        # Bold the call-for-attention pills so they stand out on print.
        if status in (
            _STATUS_URGENT,
            _STATUS_MATERIAL,
            _STATUS_SPENT_WITHOUT_BUDGET,
        ):
            status_cell.font = Font(bold=True)

        # Round 53 F1 (Move E) — render structured triplet + notes as
        # plain-English prose. Replaces the Round-51 prefix encoding for
        # the visible cell. Pre-Round-53 prefix-encoded cells still
        # round-trip through ``load_prior_period_comments`` +
        # ``decode_commentary`` (the reader handles both forms).
        # R1 fix: when prose is empty AND the Status is non-OK, auto-
        # fill with a "(no commentary recorded)" placeholder so the
        # cell doesn't print as a contradiction (urgent status with
        # blank Comments looks like the BM ignored the alert).
        c_notes, c_driver, c_outlook, c_action = commentary_tuple.get(sp, ("", "", "", ""))
        prose = render_commentary_prose(
            notes=c_notes,
            driver=c_driver,
            outlook=c_outlook,
            action=c_action,
        )
        if not prose and sp in sps_on_watchlist:
            # R2 fix: imperative cue for attention statuses so the
            # cell doesn't print as a contradiction (urgent status
            # with blank Comments looks like the BM ignored the
            # alert). Round 56 dropped the No-spend-yet special-case
            # (the pill itself was removed). Round 62 — gate the
            # auto-fill on Watchlist membership (the same signal that
            # drives the pink fill) rather than the Status pill
            # buckets. Any row with pink fill now also gets the
            # prompt — no pink-with-blank-comments contradictions.
            prose = "Action needed: add commentary."
        # Defensive Excel-formula-injection guard. A user typing
        # ``=SUM(...)`` etc. into Notes would otherwise get
        # auto-evaluated by Excel on file open, surfacing as ``#NAME?``
        # or unexpected numerics. We prepend an apostrophe — Excel's
        # canonical "force text" sigil. The apostrophe doesn't render
        # in Excel display but is preserved in the stored value, and
        # ``load_prior_period_comments`` strips it on read.
        # Round 57: Comments at col 13 (was col 14 in F2 — Trend dropped).
        if prose and prose[0] in ("=", "+", "-", "@"):
            comment_cell = ws.cell(row=row_idx, column=13, value="'" + prose)
        else:
            comment_cell = ws.cell(row=row_idx, column=13, value=prose)
        # Round 47: long commentary cells need wrap_text or they
        # overflow horizontally and push the print width past one page.
        comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Round 51 R1 fix: force text format so a user pasting
        # "01/04/2026 reviewed" doesn't get auto-coerced into an Excel
        # date by General-format inference on re-save.
        comment_cell.number_format = "@"
        # R1 fix: explicit row height so multi-line prose doesn't clip
        # on print. Default Excel row height is 15pt — shows only the
        # first line. Heuristic: ~50 chars per visual line at column
        # width 50, ~15pt per line.
        # F2 R1: Comments column width is now 32 (was 40 in R2; further
        # reduced to relieve print-width compression after F2 added
        # Status + Trend columns). Row-height heuristic tracks the
        # same width.
        col_12_width = 32
        prose_visual_lines = max(1, (len(prose) + col_12_width - 1) // col_12_width)
        if prose_visual_lines > 1:
            ws.row_dimensions[row_idx].height = 15 * prose_visual_lines

        # Round 60 — pink fill matches the in-app Watchlist filter.
        # The in-app filters on per-line ``is_over and is_material``;
        # ``sps_on_watchlist`` is the set of sub-programs whose row
        # appears on that tab. Pre-R60 the fill used the per-sub-
        # program Status pill instead, but Status only looks at the
        # Expenditure side — Revenue-over-budget rows showed up in
        # the in-app Watchlist but did NOT print pink, leaving the
        # two views inconsistent.
        # R58 history: the pre-R58 condition used ``available < 0``
        # which had its own divergence problem (lumpy revenue
        # arrival). R60 settles on the in-app signal as the single
        # source of truth.
        if sp in sps_on_watchlist:
            for col_idx in range(1, 14):
                ws.cell(row=row_idx, column=col_idx).fill = _OVER_FILL

    # Freeze panes below the header so the column titles stay
    # visible while the user scrolls through sub-programs.
    ws.freeze_panes = "A3"

    # F2 R1: explicit print_area pins what gets printed. Without this,
    # Excel's "Print Entire Workbook" mode (one-click in the print
    # dialog) would paginate both sheets and double the council
    # paper trail. Round 57: 13 cols (A:M) after Trend dropped.
    last_data_row = max(2, 2 + len(sub_programs))
    ws.print_area = f"A1:M{last_data_row}"

    # Round 68 — asymmetric data-bar range when the column contains
    # at least one negative value.
    #
    # All-positive column: range is [0, 1.0] → 0% sits flush at the
    # left edge, 100% at the right (the existing R65/R66 behaviour).
    #
    # Mixed-sign column: range is [-1/9, 1.0] → 0 sits at 1/10 from
    # the left. The 10% slice to the left of zero hosts negative
    # values (anything below -1/9 saturates against the leftmost
    # edge); the 90% slice to the right hosts positives 0–100% (and
    # >100% saturates at full width). This intentionally squeezes
    # negatives 9× so the common positive range stays legible while
    # any deficit is still visually obvious.
    #
    # Round 65 history: pre-R65 tried auto-scale (min..max) but a
    # 2,021% over-collection or a 500% unbudgeted-rev outlier
    # squashed every normal row to a sliver. Pinned ranges (R65
    # onwards) give a stable visual scale across runs.
    if sub_programs:
        from openpyxl.formatting.rule import DataBarRule

        # Available Balance % = (Annual Exp Budget − Exp YTD − Orders) ÷ Annual Exp Budget
        has_neg_avail = False
        for sp in sub_programs:
            eb = exp_b.get(sp, Decimal("0"))
            if eb <= 0:
                continue
            ey = exp_y.get(sp, Decimal("0"))
            oo = orders.get(sp, Decimal("0"))
            if (eb - ey - oo) < 0:
                has_neg_avail = True
                break

        # Revenue Budget % Received = Revenue YTD ÷ Annual Rev Budget
        has_neg_rev = False
        for sp in sub_programs:
            rb = rev_b.get(sp, Decimal("0"))
            if rb <= 0:
                continue
            if rev_y.get(sp, Decimal("0")) < 0:
                has_neg_rev = True
                break

        # 0 maps to 1/10 of the bar width when start = -1/9 and end = 1
        # (because 0 / (1 - (-1/9)) = (1/9) / (10/9) = 1/10).
        data_bar_neg_start = -1.0 / 9.0  # ≈ -0.1111

        for col_letter, has_negative in (("C", has_neg_avail), ("D", has_neg_rev)):
            start_value = data_bar_neg_start if has_negative else 0.0
            ws.conditional_formatting.add(
                f"{col_letter}3:{col_letter}{last_data_row}",
                DataBarRule(  # type: ignore[no-untyped-call]
                    start_type="num",
                    start_value=start_value,
                    end_type="num",
                    end_value=1,  # 1.0 = 100% (cells store fractions)
                    color=_DATA_BAR_COLOR,
                    showValue=True,
                ),
            )

    # F2 R1: AutoFilter + tab colour on the Watchlist sheet so the
    # council-targeted view is interactive and visually distinct.
    # F2 R2 fix: AutoFilter only when there are data rows to filter —
    # an empty Watchlist (every sub-program on track) was producing a
    # degenerate "A2:M2" header-only filter that rendered an inert
    # dropdown to a council reader.
    if filter_to_non_ok:
        # `filter_to_non_ok` is the Watchlist sheet's tell.
        # Tab colour = the canonical pink for "needs attention".
        ws.sheet_properties.tabColor = "F4CCCC"  # HL_MISMATCH
        if sub_programs:
            ws.auto_filter.ref = f"A2:M{last_data_row}"

    # Round 47 — print page setup. Without this the 12-column report
    # overflows portrait A4 onto 3-4 pages and rows split across page
    # breaks. Schools print these for council meetings.
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    page_setup_pr = ws.sheet_properties.pageSetUpPr
    if page_setup_pr is not None:
        page_setup_pr.fitToPage = True
    ws.print_title_rows = "1:2"  # repeat title + header on every printed page
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)
    # Period in the centre of the page header; page-N-of-N at right
    # of the footer; tool-version footer left so the artefact is
    # forensically traceable.
    # mypy can't prove ws.oddHeader / ws.oddFooter are non-None (openpyxl
    # types them as Optional even though they're always populated on a
    # real Worksheet), so we guard explicitly to avoid noise.
    odd_header = ws.oddHeader
    odd_footer = ws.oddFooter
    if period_label and odd_header is not None and odd_header.center is not None:
        odd_header.center.text = f"Sub-Program Budget Report — {period_label}"
    if odd_footer is not None:
        if odd_footer.left is not None:
            # Round 57: the F2 trend-warning footer was dropped along
            # with the Trend column. Always show the plain attribution.
            odd_footer.left.text = "Generated by School Tool"
        if odd_footer.right is not None:
            odd_footer.right.text = "Page &P of &N"


def _write_watchlist_sheet(
    ws: Any,
    lines: list[SubProgramLine],
    period_label: str,
    materiality_dollar: int = 100,
    *,
    prior_funds: dict[str, Decimal] | None = None,
    revenue_threshold: float = 101.0,
    expense_threshold: float = 101.0,
) -> None:
    """F2: write the Watchlist sheet — a filtered subset of the main
    sheet showing only sub-programs whose Status is not "On track",
    sorted by absolute variance descending (biggest concerns first).

    Council members read this sheet to find what needs attention
    without scanning past every healthy row in the main report. Same
    13-column shape as the main sheet (Round 57 dropped Trend) so a
    reader looking at a flagged row finds the exact cell positions
    they're used to.
    """
    _write_monthly_sub_program_sheet(
        ws,
        lines,
        period_label,
        materiality_dollar,
        prior_funds=prior_funds,
        filter_to_non_ok=True,
        sort_by_variance_desc=True,
        sheet_title_override="Watchlist — sub-programs needing attention",
        revenue_threshold=revenue_threshold,
        expense_threshold=expense_threshold,
    )


def _to_int_or_str(code: str) -> Any:
    """Render a sub-program code as an int when possible, else as str.

    The Monthly Sub Program Report stores codes as integers (4101,
    1320, …) which Excel sorts and right-aligns.  Codes that are
    non-numeric (rare — e.g. "EI/SP" sentinels) stay as strings.
    """
    text = str(code).strip()
    if text.isdigit():
        return int(text)
    return text


def _write_combined_sheet(ws: Any, lines: list[SubProgramLine], period_label: str) -> None:
    """Round 26 — append a per-sub-program Combined / YTD sheet.

    Columns: Sub-program · Description · Revenue YTD · Expense YTD ·
    Net YTD · Annual budget net.  Sorted by Net YTD ascending so the
    biggest school-funded gaps surface first.  Subsidised rows
    (Net YTD < 0) carry the canonical pink HL_MISMATCH fill.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    assert isinstance(ws, Worksheet)

    rev_b: dict[str, Decimal] = {}
    exp_b: dict[str, Decimal] = {}
    rev_y: dict[str, Decimal] = {}
    exp_y: dict[str, Decimal] = {}
    desc: dict[str, str] = {}
    for ln in lines:
        sp = ln.sub_program
        if ln.account.lower().startswith("revenue"):
            rev_b[sp] = rev_b.get(sp, Decimal("0")) + ln.budget
            rev_y[sp] = rev_y.get(sp, Decimal("0")) + ln.ytd
        elif ln.account.lower().startswith("expenditure"):
            exp_b[sp] = exp_b.get(sp, Decimal("0")) + ln.budget
            exp_y[sp] = exp_y.get(sp, Decimal("0")) + ln.ytd
        if ln.description and sp not in desc:
            desc[sp] = ln.description

    sub_programs = sorted(set(rev_b.keys()) | set(exp_b.keys()))
    # Pre-compute & sort by Net YTD ascending (biggest deficit first).
    rows: list[tuple[str, str, Decimal, Decimal, Decimal, Decimal]] = []
    for sp in sub_programs:
        ry = rev_y.get(sp, Decimal("0"))
        ey = exp_y.get(sp, Decimal("0"))
        rb = rev_b.get(sp, Decimal("0"))
        eb = exp_b.get(sp, Decimal("0"))
        rows.append((sp, desc.get(sp, ""), ry, ey, ry - ey, rb - eb))
    rows.sort(key=lambda r: r[4])

    # Title row (merged).
    title = _sheet_title("Annual Sub-Program Budget Report", period_label, "Combined")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    # Header row.
    headers = [
        "Sub-program",
        "Description",
        "Revenue YTD",
        "Expense YTD",
        "Net YTD",
        "Annual budget net",
    ]
    widths = [12, 38, 14, 14, 16, 18]
    for col_idx, (header, width) in enumerate(zip(headers, widths, strict=True), start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows.
    for row_idx, (sp, description, ry, ey, net_ytd, net_budget) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=sp)
        ws.cell(row=row_idx, column=2, value=description)
        for col_idx, value in enumerate(
            (float(ry), float(ey), float(net_ytd), float(net_budget)), start=3
        ):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.number_format = _ACCOUNTING_FMT

        # Pink fill on subsidised rows (Net YTD < 0) — same HL_MISMATCH
        # token used by the over-budget rows on the other sheets.
        if net_ytd < 0:
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = _OVER_FILL

    # Freeze panes below the headers.
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Public API -- generate_report
# ---------------------------------------------------------------------------


def generate_report(
    report_file: Path,
    comments_file: Path | None,
    output_file: Path,
    progress: ProgressFn,
    over_budget_threshold: float = 101.0,
    write_xlsx: bool = True,
    *,
    revenue_threshold: float | None = None,
    expense_threshold: float | None = None,
    materiality_dollar: int = 100,
) -> ReportSummary:
    """Orchestrate parse + comment join + optional XLSX write.

    Parameters
    ----------
    report_file:
        CASES21 GL21157 PDF (or XLSX fallback).
    comments_file:
        Optional prior-period commentary workbook.
    output_file:
        Destination ``.xlsx`` path (used only when ``write_xlsx=True``).
    progress:
        Callback ``(percent: int, message: str) -> None``.
    over_budget_threshold:
        Combined threshold applied to BOTH Revenue and Expenditure when
        the per-section overrides below are not supplied.  Kept for
        backward compatibility with all earlier callers.
    revenue_threshold, expense_threshold:
        Round 21 — optional per-section thresholds.  When supplied they
        override ``over_budget_threshold`` for that section.  Schools
        often want a different tolerance on Revenue (over-collecting
        is rarely a problem) than on Expenditure (over-running is the
        whole point of the report).
    write_xlsx:
        When True (default), write the XLSX workbook to ``output_file``.
        Pass False to skip the write step (e.g. for the preview-then-export
        two-phase flow in Sub-Program tool v2).
    """
    rev_th = revenue_threshold if revenue_threshold is not None else over_budget_threshold
    exp_th = expense_threshold if expense_threshold is not None else over_budget_threshold
    progress(10, "Reading PDF…")

    period_label = ""
    suffix = report_file.suffix.lower()
    if suffix == ".pdf":
        lines, period_label = parse_sub_program_pdf_with_period(report_file)
    elif suffix in {".xlsx", ".xlsm"}:
        lines = parse_sub_program_xlsx(report_file)
    else:
        raise ValueError(
            f"Unsupported report file format: {suffix!r}. Please supply a .pdf or .xlsx file."
        )

    progress(40, "Joining commentary…")

    comments: dict[tuple[str, str], str] = {}
    if comments_file is not None:
        comments = load_prior_period_comments(comments_file)

    final_lines: list[SubProgramLine] = []
    from dataclasses import replace as _replace

    for ln in lines:
        # Round 51 Phase D — decode the prior-period cell into the
        # structured triplet + notes. Pre-Phase-D files (no prefix)
        # round-trip as Notes-only with the three dropdowns blank.
        raw = comments.get((ln.sub_program, ln.account), "")
        if not raw:
            raw = comments.get((ln.sub_program, ln.description), "")
        notes, driver, outlook, action = decode_commentary(raw)
        if (
            notes != ln.commentary
            or driver != ln.commentary_driver
            or outlook != ln.commentary_outlook
            or action != ln.commentary_action
        ):
            ln = _replace(
                ln,
                commentary=notes,
                commentary_driver=driver,
                commentary_outlook=outlook,
                commentary_action=action,
            )
        final_lines.append(ln)

    # Round 56: pacing dropped — _recompute_is_over no longer takes
    # calendar_pct. used_pct vs threshold is the sole over-budget gate.
    final_lines = _recompute_is_over(
        final_lines,
        over_budget_threshold,
        revenue_threshold=rev_th,
        expense_threshold=exp_th,
        materiality_dollar=materiality_dollar,
    )

    # Round 57 — when a prior-period XLSX was supplied via the
    # ``Prior-period comments`` picker, also pull the carry-forward
    # field (``Funds from Previous Years``) per sub-program so the
    # output workbook keeps that column populated across reports.
    # Failure to read is non-fatal: an old comments-only file simply
    # produces an empty dict and the column stays blank.
    prior_funds: dict[str, Decimal] = {}
    if comments_file is not None:
        try:
            prior_funds = load_prior_period_funds(comments_file)
        except (ValueError, OSError, InvalidOperation):
            prior_funds = {}

    if write_xlsx:
        progress(70, "Writing workbook...")
        _write_xlsx(
            final_lines,
            output_file,
            period_label=period_label,
            over_budget_threshold=over_budget_threshold,
            materiality_dollar=materiality_dollar,
            prior_funds=prior_funds,
            # Round 62 — thread the per-section thresholds through so
            # compute_status_pill's Revenue-side check uses the same
            # value the in-app slider produced.
            revenue_threshold=rev_th,
            expense_threshold=exp_th,
        )
    else:
        progress(70, "Preparing preview...")

    over_budget_lines = [ln for ln in final_lines if ln.is_over]
    total_budget = sum((ln.budget for ln in final_lines), Decimal("0"))
    total_ytd = sum((ln.ytd for ln in final_lines), Decimal("0"))

    progress(100, "Done")

    faculty_counts: dict[str, int] = {}
    faculty_budget: dict[str, Decimal] = {}
    faculty_ytd: dict[str, Decimal] = {}
    for ln in final_lines:
        key = ln.faculty or "Unknown"
        faculty_counts[key] = faculty_counts.get(key, 0) + 1
        faculty_budget[key] = faculty_budget.get(key, Decimal("0")) + ln.budget
        faculty_ytd[key] = faculty_ytd.get(key, Decimal("0")) + ln.ytd

    faculty_used_pct: dict[str, Decimal] = {
        k: (faculty_ytd[k] / faculty_budget[k] * Decimal("100"))
        if faculty_budget[k] != Decimal("0")
        else Decimal("0")
        for k in sorted(faculty_budget.keys())
    }

    return ReportSummary(
        lines=final_lines,
        faculty_counts=faculty_counts,
        over_budget_lines=over_budget_lines,
        total_budget=total_budget,
        total_ytd=total_ytd,
        output_path=output_file,
        faculty_budget=faculty_budget,
        faculty_used_pct=faculty_used_pct,
        period_label=period_label,
        over_budget_threshold=over_budget_threshold,
        revenue_threshold=revenue_threshold or over_budget_threshold,
        expense_threshold=expense_threshold or over_budget_threshold,
        materiality_dollar=materiality_dollar,
        prior_funds=prior_funds,
    )
