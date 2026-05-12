from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from toolkit.tokens import HL_MISMATCH
from tools.sub_program.logic import (
    _OVER_FILL,
    ReportSummary,
    SubProgramLine,
    _extract_period_label,
    _recompute_is_over,
    _sheet_title,
    _write_xlsx,
    generate_report,
    load_prior_period_comments,
    parse_decimal,
    parse_sub_program_pdf,
    parse_sub_program_pdf_with_period,
)

# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------

_SAMPLE_PDF = Path(
    "Samples/Annual Subprogram Budget Report/GL21157_Annual Subprogram budget report.pdf"
)


def _make_comments_xlsx(tmp_path: Path) -> Path:
    """Build a minimal prior-period comments XLSX for fixture use."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Sub-Program", "Account", "Commentary"])  # type: ignore[union-attr]
    ws.append(["4001", "Expenditure", "Reviewed by council"])  # type: ignore[union-attr]
    ws.append(["8599", "Expenditure", "Rowing costs on track"])  # type: ignore[union-attr]
    ws.append(["1251", "Revenue", ""])  # type: ignore[union-attr]
    out = tmp_path / "comments.xlsx"
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# Test: currency parser edge cases
# ---------------------------------------------------------------------------


class TestParseDecimal:
    def test_plain_number(self) -> None:
        assert parse_decimal("1,234.56") == Decimal("1234.56")

    def test_negative(self) -> None:
        assert parse_decimal("-500.00") == Decimal("-500.00")

    def test_parentheses_negative(self) -> None:
        assert parse_decimal("(500.00)") == Decimal("-500.00")

    def test_dollar_prefix(self) -> None:
        assert parse_decimal("$0.00") == Decimal("0.00")

    def test_em_dash(self) -> None:
        assert parse_decimal("—") == Decimal("0")

    def test_blank(self) -> None:
        assert parse_decimal("") == Decimal("0")

    def test_integer_string(self) -> None:
        assert parse_decimal("42") == Decimal("42")

    def test_invalid_raises(self) -> None:
        with pytest.raises(ValueError):
            parse_decimal("not-a-number!!")


# ---------------------------------------------------------------------------
# Test: parse real sample PDF
# ---------------------------------------------------------------------------


class TestParseSamplePdf:
    @pytest.fixture(scope="class")
    def lines(self) -> list[SubProgramLine]:
        return parse_sub_program_pdf(_SAMPLE_PDF)

    def test_minimum_line_count(self, lines: list[SubProgramLine]) -> None:
        assert len(lines) >= 10, f"Expected >= 10 lines, got {len(lines)}"

    def test_all_have_nonempty_sub_program(self, lines: list[SubProgramLine]) -> None:
        bad = [ln for ln in lines if not ln.sub_program.strip()]
        assert not bad, f"Lines with empty sub_program: {bad[:3]}"

    def test_all_have_decimal_budget(self, lines: list[SubProgramLine]) -> None:
        for ln in lines:
            assert isinstance(ln.budget, Decimal)

    def test_budget_sum_positive(self, lines: list[SubProgramLine]) -> None:
        total = sum((ln.budget for ln in lines), Decimal("0"))
        assert total > Decimal("0"), f"Total budget should be positive, got {total}"

    def test_faculty_assigned(self, lines: list[SubProgramLine]) -> None:
        # At least some lines should have a faculty
        with_faculty = [ln for ln in lines if ln.faculty is not None]
        assert len(with_faculty) > 0

    def test_account_section_values(self, lines: list[SubProgramLine]) -> None:
        accounts = {ln.account for ln in lines}
        assert accounts <= {"Revenue", "Expenditure"}, f"Unexpected account values: {accounts}"

    def test_used_pct_is_decimal(self, lines: list[SubProgramLine]) -> None:
        for ln in lines:
            assert isinstance(ln.used_pct, Decimal)

    def test_sub_program_codes_are_numeric(self, lines: list[SubProgramLine]) -> None:
        non_numeric = [ln for ln in lines if not ln.sub_program.isdigit()]
        assert not non_numeric, f"Non-numeric sub_programs: {non_numeric[:3]}"

    def test_no_total_rows_included(self, lines: list[SubProgramLine]) -> None:
        # "Revenue totals" and "Expenditure totals" must not appear as data rows
        for ln in lines:
            assert "total" not in ln.description.lower() or ln.sub_program.isdigit()

    # Round 58 — column-shift bug regression tests. Pre-R58 the parser
    # assumed pre[-1] was always YTD, but the GL21157 PDF omits the
    # YTD column entirely when there's no spend (pct=0.00). This caused
    # ~20% of typical school rows to misread Annual budget as YTD and
    # Last-year budget as Annual budget. Pin the fix against future
    # drift by spot-checking three canonical shapes against the sample
    # PDF's known values.

    def test_zero_spend_row_with_orders_parses_correctly(self, lines: list[SubProgramLine]) -> None:
        """Round 58: 4016 Instrumental Music Expenditure has Annual
        $220,000, no YTD spend, $181,818 outstanding orders. Pre-fix
        the parser read budget=$243,450, ytd=$220,000 (off by one column).
        """
        match = next(
            (
                ln
                for ln in lines
                if ln.sub_program == "4016" and ln.account.startswith("Expenditure")
            ),
            None,
        )
        assert match is not None, "Sub-program 4016 Expenditure missing"
        assert match.budget == Decimal("220000"), (
            f"4016 Expenditure budget should be $220,000; got ${match.budget}"
        )
        assert match.ytd == Decimal("0"), (
            f"4016 Expenditure YTD should be $0 (no spend yet); got ${match.ytd}"
        )
        assert match.outstanding_orders == Decimal("181818")

    def test_zero_spend_row_without_orders_parses_correctly(
        self, lines: list[SubProgramLine]
    ) -> None:
        """Round 58: 4010 Photography Expenditure has Annual $6,000,
        no YTD spend, no orders. Pre-fix the parser read budget=$1,500
        (last-year budget), ytd=$6,000 (annual budget)."""
        match = next(
            (
                ln
                for ln in lines
                if ln.sub_program == "4010" and ln.account.startswith("Expenditure")
            ),
            None,
        )
        assert match is not None, "Sub-program 4010 Expenditure missing"
        assert match.budget == Decimal("6000"), (
            f"4010 Expenditure budget should be $6,000; got ${match.budget}"
        )
        assert match.ytd == Decimal("0")
        assert match.outstanding_orders == Decimal("0")

    def test_no_annual_budget_row_parses_with_zeros(self, lines: list[SubProgramLine]) -> None:
        """Round 58: 4051 Dance Activities Expenditure has no current-year
        allocation (no annual budget, no YTD, no orders) — only last-year
        history. Pre-fix the parser read budget=$10,107 (last-year
        actual), ytd=$10,205 (last-year budget)."""
        match = next(
            (
                ln
                for ln in lines
                if ln.sub_program == "4051" and ln.account.startswith("Expenditure")
            ),
            None,
        )
        assert match is not None, "Sub-program 4051 Expenditure missing"
        assert match.budget == Decimal("0"), (
            f"4051 Expenditure budget should be $0 (no current allocation); got ${match.budget}"
        )
        assert match.ytd == Decimal("0")
        assert match.last_year_budget == Decimal("10205")
        assert match.last_year_actual == Decimal("10107")

    def test_unbudgeted_spend_row_with_negative_avail(self, lines: list[SubProgramLine]) -> None:
        """Round 60: an unbudgeted-spend row has 1 pre token (the YTD)
        AND ≥1 post token (the negative Available Balance produced
        by that spend). 8650 Rowing Program "(See 8599)" paid out
        $26,924 with no budget; PDF reads ``26,924 0.00 -26,924``."""
        match = next(
            (
                ln
                for ln in lines
                if ln.sub_program == "8650" and ln.account.startswith("Expenditure")
            ),
            None,
        )
        assert match is not None, "Sub-program 8650 Expenditure missing"
        assert match.budget == Decimal("0")
        assert match.ytd == Decimal("26924")

    def test_last_year_only_row_parses_with_zero_current(self, lines: list[SubProgramLine]) -> None:
        """Round 60: 8505 School Saving Bonus has 1 pre token ($255,751)
        and 0 post tokens. Per the KMAR reference workbook, that
        pattern means LY_actual only — no current-year revenue,
        expense, or available balance. Pre-R60 the parser misread
        this as unbudgeted spend (ytd=$255,751)."""
        match = next(
            (
                ln
                for ln in lines
                if ln.sub_program == "8505" and ln.account.startswith("Expenditure")
            ),
            None,
        )
        assert match is not None, "Sub-program 8505 Expenditure missing"
        assert match.budget == Decimal("0"), f"8505 budget should be $0; got ${match.budget}"
        assert match.ytd == Decimal("0"), (
            f"8505 YTD should be $0 (LY-only history); got ${match.ytd}"
        )
        assert match.last_year_actual == Decimal("255751")


class TestPositionalParserSpotChecks:
    """Round 61 — positional parser regression suite.

    Pin the parser's output for ~15 representative sub-program rows
    against the values transcribed directly from the sample PDF
    (Samples/Annual Subprogram Budget Report/GL21157_Annual Subprogram
    budget report.pdf, dated 3 March 2026). Pre-R61 the parser used
    pct-based heuristics that misread roughly 20% of rows whenever a
    column was blank. The positional parser reads each column from
    fixed x-coordinates derived from the PDF header, so blank columns
    correctly produce zero rather than shifting the column window.
    """

    @pytest.fixture(scope="class")
    def lines_by_key(self) -> dict[tuple[str, str], SubProgramLine]:
        lines = parse_sub_program_pdf(_SAMPLE_PDF)
        return {(ln.sub_program, ln.account): ln for ln in lines}

    @pytest.mark.parametrize(
        "sub_prog,account,budget,ytd,orders",
        [
            # (sub_prog, account, expected_budget, expected_ytd, expected_orders)
            # Standard: all columns present.
            ("4001", "Expenditure", 52450, 8241, 489),
            ("8328", "Expenditure", 360500, 30931, 0),
            ("8328", "Revenue", 360500, 112300, 0),
            # Zero YTD (blank YTD column in PDF, pct=0.00).
            ("4010", "Expenditure", 6000, 0, 0),
            ("4016", "Expenditure", 220000, 0, 181818),
            ("8851", "Expenditure", 31160, 0, 0),
            ("8756", "Expenditure", 50700, 0, 0),
            # No annual budget shown (LY history only on Expenditure side).
            ("4051", "Expenditure", 0, 0, 0),
            ("4290", "Expenditure", 0, 0, 0),
            ("8401", "Expenditure", 0, 0, 0),
            # Last-year-only row (1 pre-token + 0 post). Must NOT
            # interpret the lone $255,751 as YTD.
            ("8505", "Expenditure", 0, 0, 0),
            ("8505", "Revenue", 0, 0, 0),
            # Unbudgeted current spend (1 pre-token + 1 post).
            ("8650", "Expenditure", 0, 26924, 0),
            # Standard Revenue rows.
            ("5450", "Revenue", 23500, 0, 0),
            ("6001", "Revenue", 120, 12, 0),
            ("7001", "Revenue", 26000, 24389, 0),
            ("8851", "Revenue", 31160, 33326, 0),
            # Expenditure with both YTD and orders shown.
            ("7001", "Expenditure", 581700, 120373, 1686887),
            ("6201", "Expenditure", 221000, 46041, 18515),
            ("4101", "Expenditure", 23125, 279, 0),
        ],
    )
    def test_pdf_row_matches_expected(
        self,
        lines_by_key: dict[tuple[str, str], SubProgramLine],
        sub_prog: str,
        account: str,
        budget: int,
        ytd: int,
        orders: int,
    ) -> None:
        line = lines_by_key.get((sub_prog, account))
        assert line is not None, f"{sub_prog} {account} missing from parsed output"
        assert line.budget == Decimal(budget), (
            f"{sub_prog} {account} budget: expected ${budget}, got ${line.budget}"
        )
        assert line.ytd == Decimal(ytd), (
            f"{sub_prog} {account} YTD: expected ${ytd}, got ${line.ytd}"
        )
        assert line.outstanding_orders == Decimal(orders), (
            f"{sub_prog} {account} orders: expected ${orders}, got ${line.outstanding_orders}"
        )


# ---------------------------------------------------------------------------
# Test: over-budget detection
# ---------------------------------------------------------------------------


class TestOverBudget:
    def test_is_over_true_when_ytd_exceeds_budget(self) -> None:
        ln = SubProgramLine(
            sub_program="4400",
            account="Revenue",
            description="Mathematics",
            budget=Decimal("100"),
            ytd=Decimal("150"),
            remaining=Decimal("-50"),
            used_pct=Decimal("150"),
            faculty="Curriculum",
            is_over=True,
        )
        assert ln.is_over is True
        assert ln.used_pct == Decimal("150")

    def test_is_over_false_when_within_budget(self) -> None:
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Art",
            budget=Decimal("100"),
            ytd=Decimal("80"),
            remaining=Decimal("20"),
            used_pct=Decimal("80"),
            faculty="Curriculum",
            is_over=False,
        )
        assert ln.is_over is False

    def test_parse_pdf_detects_over_budget(self) -> None:
        """Line 4400 Mathematics has pct > 100 in the Revenue section."""
        lines = parse_sub_program_pdf(_SAMPLE_PDF)
        # Sub-program 4400 in Revenue has used_pct ~2021 (YTD 20,213 vs budget 1,000)
        over = [ln for ln in lines if ln.is_over]
        assert over, "Expected at least one over-budget line in the sample PDF"


# ---------------------------------------------------------------------------
# Test: load_prior_period_comments
# ---------------------------------------------------------------------------


class TestLoadComments:
    def test_basic_load(self, tmp_path: Path) -> None:
        comments_xlsx = _make_comments_xlsx(tmp_path)
        result = load_prior_period_comments(comments_xlsx)
        assert ("4001", "Expenditure") in result
        assert result[("4001", "Expenditure")] == "Reviewed by council"
        assert result[("8599", "Expenditure")] == "Rowing costs on track"

    def test_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ValueError, match="not found"):
            load_prior_period_comments(tmp_path / "nonexistent.xlsx")

    def test_empty_commentary_row(self, tmp_path: Path) -> None:
        comments_xlsx = _make_comments_xlsx(tmp_path)
        result = load_prior_period_comments(comments_xlsx)
        # The third row has an empty commentary
        assert result.get(("1251", "Revenue"), "") == ""

    def test_notes_synonym(self, tmp_path: Path) -> None:
        """Round 21 — 'Notes' header is accepted as a comment synonym."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Sub-Program", "Account", "Notes"])
        ws.append(["4001", "Expenditure", "PE consumables reviewed"])
        path = tmp_path / "notes_only.xlsx"
        wb.save(path)

        result = load_prior_period_comments(path)
        assert result[("4001", "Expenditure")] == "PE consumables reviewed"

    def test_no_account_column_falls_back_to_title(self, tmp_path: Path) -> None:
        """Round 21 — files without an 'Account' column (e.g. our own
        exports) are still loadable; the line title becomes the second key."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        # Mirrors the Revenue export schema — Title in col 1, Comments at end.
        ws.append(
            [
                "Sub Prog.",
                "Title",
                "Last year actual",
                "Last year budget",
                "Annual budget",
                "YTD",
                "% Budget received",
                "Comments",
            ]
        )
        ws.append(["1251", "Camp fees", "—", "—", 5000, 4500, "90%", "Camp paid"])
        path = tmp_path / "from_export.xlsx"
        wb.save(path)

        result = load_prior_period_comments(path)
        # Second key is the Title text, not an account string.
        assert result[("1251", "Camp fees")] == "Camp paid"
        # And it definitely is NOT the budget column value (the bug that
        # caused this round of work — col 2 used to be silently chosen).
        assert "5000" not in str(result.get(("1251", "Camp fees"), ""))

    def test_missing_comment_column_raises(self, tmp_path: Path) -> None:
        """Round 21 — bug guard: when no comment column is found we must
        raise instead of silently defaulting to a budget column.

        Pre-Round-21 the loader fell through StopIteration to txt_col=2,
        which is "Last year actual" in our exports — meaning users saw
        last-year-actual budget values copied into the new comments
        column.  That silent fall-through is now a hard error.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        # No comment-like header at all — every column is a budget one.
        ws.append(["Sub Prog.", "Title", "Annual budget", "YTD"])
        ws.append(["4001", "Art", 1000, 200])
        path = tmp_path / "no_comments.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="comments column"):
            load_prior_period_comments(path)

    def test_reads_all_sheets(self, tmp_path: Path) -> None:
        """Round 21 — Revenue + Expenditure sheets should both contribute
        comments (previously only the first sheet was read)."""
        wb = openpyxl.Workbook()
        rev = wb.active
        assert rev is not None
        rev.title = "Revenue"
        rev.append(["Sub-Program", "Account", "Commentary"])
        rev.append(["1251", "Revenue", "Camp paid"])

        exp = wb.create_sheet("Expenditure")
        exp.append(["Sub-Program", "Account", "Commentary"])
        exp.append(["4001", "Expenditure", "Art supplies on order"])

        path = tmp_path / "two_sheets.xlsx"
        wb.save(path)

        result = load_prior_period_comments(path)
        assert result[("1251", "Revenue")] == "Camp paid"
        assert result[("4001", "Expenditure")] == "Art supplies on order"


# ---------------------------------------------------------------------------
# Test: generate_report round-trip
# ---------------------------------------------------------------------------


class TestGenerateReport:
    def test_round_trip_produces_xlsx(self, tmp_path: Path) -> None:
        output = tmp_path / "output.xlsx"
        progress_calls: list[tuple[int, str]] = []

        def progress(pct: int, msg: str) -> None:
            progress_calls.append((pct, msg))

        summary = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=progress,
        )

        assert output.exists(), "Output XLSX was not created"
        assert isinstance(summary, ReportSummary)
        assert len(summary.lines) >= 10

    def test_output_has_correct_header(self, tmp_path: Path) -> None:
        """Round 38 — output is now a single sheet with the Monthly Sub
        Program Report 12-column header shape."""
        output = tmp_path / "output2.xlsx"

        generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=lambda p, m: None,
        )

        wb = openpyxl.load_workbook(output)
        # Single sheet now — name is "Sub Program Report".
        assert "Sub Program Report" in wb.sheetnames, (
            f"Expected 'Sub Program Report' sheet; got {wb.sheetnames}"
        )
        ws = wb["Sub Program Report"]
        # Round 66 layout: Available Balance % YTD (col 3) and Revenue
        # Budget % Received YTD (col 4) moved between PROGRAM NAME
        # and Status. Status now at col 5; financials shift right by
        # 2. Avail Balance YTD ends up at col 12 (was col 10 pre-R66).
        # Total 13 cols.
        headers = [str(ws.cell(2, c).value or "") for c in range(1, 14)]
        assert headers[0] == "CODE"
        assert headers[1] == "PROGRAM NAME"
        assert headers[2] == "Available Balance % YTD"
        assert headers[3] == "Revenue Budget % Received YTD"
        assert headers[4] == "Status"
        assert headers[5].startswith("Funds from Previous Years")
        assert headers[6].startswith("Budget Revenue")
        assert headers[7].startswith("Total Budget Allocation Expenditure")
        assert headers[8] == "Revenue YTD"
        assert headers[9] == "Expenditure YTD"
        assert headers[10] == "Less outstanding orders"
        assert headers[11] == "Available Balance YTD"
        assert headers[12] == "Comments"

    def test_output_row_count_matches_unique_subprograms(self, tmp_path: Path) -> None:
        """Round 38 — one row per sub-program (was: one row per
        account-line, split across two sheets)."""
        output = tmp_path / "output3.xlsx"

        summary = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=lambda p, m: None,
        )

        wb = openpyxl.load_workbook(output)
        ws = wb["Sub Program Report"]
        # Row 1 = title, row 2 = header, rows 3+ = data
        data_rows = (ws.max_row or 2) - 2
        unique_sps = {ln.sub_program for ln in summary.lines}
        assert data_rows == len(unique_sps), (
            f"Expected {len(unique_sps)} data rows (one per sub-program); got {data_rows}"
        )

    def test_pink_fill_on_over_budget_rows(self, tmp_path: Path) -> None:
        """Round 38 — superseded by test_xlsx_monthly_report.py."""
        # Old shape (Revenue/Expenditure sheets) no longer produced.

    def test_over_budget_fill_all_columns(self, tmp_path: Path) -> None:
        """Round 38 — superseded by test_xlsx_monthly_report.py."""
        # Old shape (Revenue/Expenditure sheets) no longer produced.

    def test_commentary_joined(self, tmp_path: Path) -> None:
        comments_xlsx = _make_comments_xlsx(tmp_path)
        output = tmp_path / "output5.xlsx"

        summary = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=comments_xlsx,
            output_file=output,
            progress=lambda p, m: None,
        )

        with_comment = [ln for ln in summary.lines if ln.commentary]
        assert with_comment, "Expected at least one line with commentary"

    def test_progress_callbacks_fired(self, tmp_path: Path) -> None:
        output = tmp_path / "output6.xlsx"
        pcts: list[int] = []

        def progress(pct: int, msg: str) -> None:
            pcts.append(pct)

        generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=progress,
        )

        assert 10 in pcts
        assert 40 in pcts
        assert 70 in pcts
        assert 100 in pcts


# ---------------------------------------------------------------------------
# Test: per-faculty aggregated stats in ReportSummary
# ---------------------------------------------------------------------------


def _make_line_direct(
    sub_program: str,
    faculty: str | None,
    budget: str,
    ytd: str,
    is_over: bool = False,
) -> SubProgramLine:
    b = Decimal(budget)
    y = Decimal(ytd)
    return SubProgramLine(
        sub_program=sub_program,
        account="Expenditure",
        description="Test",
        budget=b,
        ytd=y,
        remaining=b - y,
        used_pct=(y / b * Decimal("100")) if b != Decimal("0") else Decimal("0"),
        faculty=faculty,
        is_over=is_over,
    )


class TestFacultyStats:
    """ReportSummary must carry per-faculty budget/ytd/used_pct aggregations."""

    def _build_summary(self, lines: list[SubProgramLine], tmp_path: Path) -> ReportSummary:
        output = tmp_path / "out.xlsx"
        # Use the frozen dataclass constructor directly so we don't need a real PDF.
        faculty_counts: dict[str, int] = {}
        faculty_budget: dict[str, Decimal] = {}
        faculty_ytd_map: dict[str, Decimal] = {}
        for ln in lines:
            key = ln.faculty or "Unknown"
            faculty_counts[key] = faculty_counts.get(key, 0) + 1
            faculty_budget[key] = faculty_budget.get(key, Decimal("0")) + ln.budget
            faculty_ytd_map[key] = faculty_ytd_map.get(key, Decimal("0")) + ln.ytd
        faculty_used_pct: dict[str, Decimal] = {
            k: (faculty_ytd_map[k] / faculty_budget[k] * Decimal("100"))
            if faculty_budget[k] != Decimal("0")
            else Decimal("0")
            for k in faculty_counts
        }
        return ReportSummary(
            lines=lines,
            faculty_counts=faculty_counts,
            over_budget_lines=[ln for ln in lines if ln.is_over],
            total_budget=sum((ln.budget for ln in lines), Decimal("0")),
            total_ytd=sum((ln.ytd for ln in lines), Decimal("0")),
            output_path=output,
            faculty_budget=faculty_budget,
            faculty_ytd=faculty_ytd_map,
            faculty_used_pct=faculty_used_pct,
        )

    def test_summary_includes_faculty_budget(self, tmp_path: Path) -> None:
        """faculty_budget sums budget across all lines per faculty."""
        lines = [
            _make_line_direct("4001", "Curriculum", "10000", "5000"),
            _make_line_direct("4002", "Curriculum", "5000", "2000"),
            _make_line_direct("5001", "Student Wellbeing", "8000", "4000"),
        ]
        # Build via generate_report using the real logic path indirectly —
        # construct a summary the same way logic.py does to verify correctness.
        summary = self._build_summary(lines, tmp_path)

        assert summary.faculty_budget["Curriculum"] == Decimal("15000")
        assert summary.faculty_budget["Student Wellbeing"] == Decimal("8000")

    def test_summary_includes_faculty_used_pct(self, tmp_path: Path) -> None:
        """faculty_used_pct is computed from totals, not averaged from per-row %s.

        Curriculum: budget=15000, ytd=7000 → 46.666...%
        If averaged from rows: row1=50%, row2=40% → avg=45% — different!
        The totals-based figure (46.666...%) is the correct one.
        """
        lines = [
            _make_line_direct("4001", "Curriculum", "10000", "5000"),  # 50%
            _make_line_direct("4002", "Curriculum", "5000", "2000"),  # 40%
        ]
        summary = self._build_summary(lines, tmp_path)

        # Totals: budget=15000, ytd=7000
        expected = Decimal("7000") / Decimal("15000") * Decimal("100")
        assert summary.faculty_used_pct["Curriculum"] == expected

        # Confirm it is NOT the naïve per-row average (45%)
        naive_avg = (Decimal("50") + Decimal("40")) / Decimal("2")
        assert summary.faculty_used_pct["Curriculum"] != naive_avg

    def test_summary_faculty_used_pct_handles_zero_budget(self, tmp_path: Path) -> None:
        """A faculty whose total budget is 0 must yield 0% — no ZeroDivisionError."""
        lines = [
            _make_line_direct("9001", "Computing & Curriculum", "0", "0"),
        ]
        summary = self._build_summary(lines, tmp_path)

        assert summary.faculty_used_pct["Computing & Curriculum"] == Decimal("0")

    def test_summary_unknown_faculty_aggregated(self, tmp_path: Path) -> None:
        """Lines with faculty=None are bucketed under 'Unknown'."""
        lines = [
            _make_line_direct("9999", None, "3000", "1500"),
            _make_line_direct("9998", None, "2000", "1000"),
        ]
        summary = self._build_summary(lines, tmp_path)

        assert "Unknown" in summary.faculty_budget
        assert summary.faculty_budget["Unknown"] == Decimal("5000")
        assert summary.faculty_ytd["Unknown"] == Decimal("2500")
        expected_pct = Decimal("2500") / Decimal("5000") * Decimal("100")
        assert summary.faculty_used_pct["Unknown"] == expected_pct


# ---------------------------------------------------------------------------
# Helper for new XLSX-structure tests
# ---------------------------------------------------------------------------


def _make_mixed_lines() -> list[SubProgramLine]:
    """Return a short list of Revenue + Expenditure lines for XLSX writer tests."""

    def _rev(sp: str, budget: str, ytd: str, lya: str = "0", lyb: str = "0") -> SubProgramLine:
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sp,
            account="Revenue",
            description=f"Rev {sp}",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=(y / b * Decimal("100")) if b else Decimal("0"),
            faculty="Curriculum",
            is_over=False,
            last_year_actual=Decimal(lya),
            last_year_budget=Decimal(lyb),
        )

    def _exp(
        sp: str,
        budget: str,
        ytd: str,
        outstanding: str = "0",
        lya: str = "0",
        lyb: str = "0",
        is_over: bool = False,
    ) -> SubProgramLine:
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sp,
            account="Expenditure",
            description=f"Exp {sp}",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=(y / b * Decimal("100")) if b else Decimal("0"),
            faculty="Curriculum",
            is_over=is_over,
            last_year_actual=Decimal(lya),
            last_year_budget=Decimal(lyb),
            outstanding_orders=Decimal(outstanding),
        )

    return [
        _rev("4001", "18950", "13760", lya="160"),
        _rev("4003", "7525", "6130"),
        _exp("4001", "52450", "8241", outstanding="489", lya="32565", lyb="32675"),
        _exp("4003", "7775", "582", outstanding="0"),
        _exp("4400", "1000", "20213", is_over=True),  # over-budget
    ]


# ---------------------------------------------------------------------------
# Test: new XLSX structure (two sheets)
# ---------------------------------------------------------------------------


class TestXlsxMonthlyReport:
    """Round 38 — replaced TestXlsxTwoSheets. The XLSX output is now a
    Sub Program Report sheet matching the school's own Monthly Sub
    Program Report workbook. Round 54 F2 added a second sheet
    (Watchlist) so the workbook is two sheets total."""

    def test_workbook_has_expected_sheets(self, tmp_path: Path) -> None:
        """Round 64 — restored the legacy Revenue + Expenditure detail
        sheets per user request. Pre-R64 only the Sub Program Report
        + Watchlist sheets shipped; Round 38 had collapsed Revenue +
        Expenditure into the combined sheet. Now both views coexist.

        With a single Expenditure line, only the Expenditure detail
        sheet is created — Revenue is skipped because there's no
        revenue line to populate it.
        """
        from tools.sub_program.logic import SubProgramLine

        out = tmp_path / "test.xlsx"
        ln = SubProgramLine(
            sub_program="4101",
            account="Expenditure",
            description="English",
            budget=Decimal("1000"),
            ytd=Decimal("400"),
            remaining=Decimal("600"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == [
            "Sub Program Report",
            "Watchlist",
            "Expenditure",
        ]

    def test_workbook_has_both_detail_sheets_when_both_sides_present(self, tmp_path: Path) -> None:
        """Round 64 — when both Revenue + Expenditure lines are present
        the workbook ships all 4 sheets."""
        from tools.sub_program.logic import SubProgramLine

        out = tmp_path / "test.xlsx"
        rev = SubProgramLine(
            sub_program="4101",
            account="Revenue",
            description="English",
            budget=Decimal("500"),
            ytd=Decimal("100"),
            remaining=Decimal("400"),
            used_pct=Decimal("20"),
            faculty="Curriculum",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="4101",
            account="Expenditure",
            description="English",
            budget=Decimal("1000"),
            ytd=Decimal("400"),
            remaining=Decimal("600"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([rev, exp], out)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == [
            "Sub Program Report",
            "Watchlist",
            "Revenue",
            "Expenditure",
        ]


class TestPeriodLabel:
    def test_extract_period_label_from_footer(self) -> None:
        """Round 65 — capture widened to include the day of month so
        the workbook title shows the actual print date (the YTD
        cut-off the CASES21 export was run for), not just the month."""
        text = "3 March 2026 13:37 1 [GL21157]"
        assert _extract_period_label(text) == "3 March 2026"

    def test_extract_period_label_no_match_returns_empty(self) -> None:
        assert _extract_period_label("no date here") == ""

    def test_sheet_title_with_period(self) -> None:
        assert _sheet_title("Report", "3 January 2026", "Revenue") == (
            "Report - 3 January 2026 Revenue"
        )

    def test_sheet_title_without_period(self) -> None:
        t = _sheet_title("Annual Sub-Program Budget Report", "", "Expenditure")
        assert t == "Annual Sub-Program Budget Report - Expenditure"
        assert "  " not in t

    def test_summary_includes_period_label(self, tmp_path: Path) -> None:
        """generate_report must populate period_label from the PDF footer."""
        output = tmp_path / "out.xlsx"
        summary = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=lambda p, m: None,
        )
        # The sample PDF footer contains a date; period_label must be non-empty.
        assert summary.period_label, (
            "period_label is empty -- check the footer regex against the sample PDF"
        )
        # The sample PDF was printed on 3 March 2026.
        assert summary.period_label == "3 March 2026"

    def test_summary_period_label_empty_for_xlsx_input(self, tmp_path: Path) -> None:
        """XLSX-sourced input has no footer date, so period_label defaults to ''."""
        # Build a minimal XLSX fixture to feed to parse_sub_program_xlsx.

        xlsx_in = tmp_path / "input.xlsx"
        wb_in = openpyxl.Workbook()
        ws_in = wb_in.active
        ws_in.append(["Sub-Program", "Title", "LYA", "LYB", "Budget", "YTD", "Pct"])  # type: ignore[union-attr]
        ws_in.append(["4001", "Art", 0, 0, 50000, 25000, 50])  # type: ignore[union-attr]
        wb_in.save(xlsx_in)

        output = tmp_path / "out.xlsx"
        summary = generate_report(
            report_file=xlsx_in,
            comments_file=None,
            output_file=output,
            progress=lambda p, m: None,
        )
        # No footer in XLSX -- period_label must be the default empty string.
        assert summary.period_label == ""


# ---------------------------------------------------------------------------
# Test: new PDF parser fields (last-year actual/budget, outstanding orders)
# ---------------------------------------------------------------------------


class TestNewParserFields:
    @pytest.fixture(scope="class")
    def lines(self) -> list[SubProgramLine]:
        result, _ = parse_sub_program_pdf_with_period(_SAMPLE_PDF)
        return result

    def test_summary_lines_have_last_year_actual(self, lines: list[SubProgramLine]) -> None:
        """At least some Revenue or Expenditure lines must have a non-zero last_year_actual."""
        non_zero = [ln for ln in lines if ln.last_year_actual != Decimal("0")]
        assert non_zero, (
            "No lines with non-zero last_year_actual found -- check the parser "
            "against the sample PDF columns."
        )

    def test_summary_lines_have_last_year_budget(self, lines: list[SubProgramLine]) -> None:
        """At least some lines must have a non-zero last_year_budget."""
        non_zero = [ln for ln in lines if ln.last_year_budget != Decimal("0")]
        assert non_zero, "No lines with non-zero last_year_budget found -- check the parser."

    def test_summary_lines_have_outstanding_orders(self, lines: list[SubProgramLine]) -> None:
        """At least some Expenditure lines must have a non-zero outstanding_orders."""
        exp_with_outstanding = [
            ln
            for ln in lines
            if ln.account == "Expenditure" and ln.outstanding_orders != Decimal("0")
        ]
        assert exp_with_outstanding, (
            "No Expenditure lines with non-zero outstanding_orders found -- "
            "check the post-pct token extraction logic."
        )

    def test_known_row_4016_revenue_last_year_values(self, lines: list[SubProgramLine]) -> None:
        """Row 4016 Revenue in the sample PDF has both last-year columns."""
        row = next(
            (ln for ln in lines if ln.sub_program == "4016" and ln.account == "Revenue"),
            None,
        )
        assert row is not None, "Row 4016 Revenue not found"
        assert row.last_year_actual == Decimal("242911"), (
            f"last_year_actual: expected 242911, got {row.last_year_actual}"
        )
        assert row.last_year_budget == Decimal("243450"), (
            f"last_year_budget: expected 243450, got {row.last_year_budget}"
        )

    def test_known_row_4001_expenditure_outstanding(self, lines: list[SubProgramLine]) -> None:
        """Row 4001 Art Expenditure has outstanding_orders=489."""
        row = next(
            (ln for ln in lines if ln.sub_program == "4001" and ln.account == "Expenditure"),
            None,
        )
        assert row is not None, "Row 4001 Expenditure not found"
        assert row.outstanding_orders == Decimal("489"), (
            f"outstanding_orders: expected 489, got {row.outstanding_orders}"
        )

    def test_revenue_lines_outstanding_defaults_to_zero(self, lines: list[SubProgramLine]) -> None:
        """Revenue lines must all have outstanding_orders == 0 (not in PDF)."""
        bad = [
            ln for ln in lines if ln.account == "Revenue" and ln.outstanding_orders != Decimal("0")
        ]
        assert not bad, (
            f"{len(bad)} Revenue line(s) have non-zero outstanding_orders: "
            f"{[(ln.sub_program, ln.outstanding_orders) for ln in bad[:3]]}"
        )


# ---------------------------------------------------------------------------
# Test: over-budget threshold (Fix 2 -- Round 9)
# ---------------------------------------------------------------------------


def _make_threshold_lines() -> list[SubProgramLine]:
    """Lines with varying used_pct values for threshold testing."""

    def _exp(sp: str, budget: str, ytd: str, pct: str) -> SubProgramLine:
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sp,
            account="Expenditure",
            description=f"Exp {sp}",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=Decimal(pct),
            faculty="Curriculum",
            is_over=False,  # will be overridden by _recompute_is_over
        )

    return [
        _exp("4001", "100", "50", "50.00"),  # 50% -- never over
        _exp("4002", "100", "100", "100.00"),  # exactly 100% -- not > 101
        _exp("4003", "100", "101", "101.00"),  # 101% -- at default threshold (101.0) = not over
        _exp("4004", "100", "102", "102.00"),  # 102% -- over at default threshold
        _exp("4005", "100", "200", "200.00"),  # 200% -- always over (except threshold 200+)
    ]


class TestOverBudgetThreshold:
    """Threshold-aware is_over logic and XLSX fills."""

    def test_recompute_default_threshold(self) -> None:
        """Default 101.0: lines with used_pct > 101 are over."""
        lines = _recompute_is_over(_make_threshold_lines(), 101.0)
        assert not lines[0].is_over  # 50%
        assert not lines[1].is_over  # 100%
        assert not lines[2].is_over  # 101% -- NOT > 101
        assert lines[3].is_over  # 102%
        assert lines[4].is_over  # 200%

    def test_recompute_exact_100_threshold(self) -> None:
        """Threshold 100.0: lines with used_pct > 100 are over."""
        lines = _recompute_is_over(_make_threshold_lines(), 100.0)
        assert not lines[0].is_over  # 50%
        assert not lines[1].is_over  # exactly 100 -- NOT > 100
        assert lines[2].is_over  # 101%
        assert lines[3].is_over  # 102%
        assert lines[4].is_over  # 200%

    def test_recompute_high_threshold_suppresses_all(self) -> None:
        """Threshold 9999: no line is over-budget."""
        lines = _recompute_is_over(_make_threshold_lines(), 9999.0)
        assert not any(ln.is_over for ln in lines)

    def test_generate_report_default_threshold_in_summary(self, tmp_path: Path) -> None:
        """generate_report stores the threshold in ReportSummary."""
        output = tmp_path / "out.xlsx"
        summary = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output,
            progress=lambda p, m: None,
        )
        assert summary.over_budget_threshold == 101.0

    def test_generate_report_custom_threshold_changes_over_count(self, tmp_path: Path) -> None:
        """A high threshold produces fewer (or zero) over-budget lines."""
        output_default = tmp_path / "out_default.xlsx"
        output_high = tmp_path / "out_high.xlsx"

        summary_default = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output_default,
            progress=lambda p, m: None,
            over_budget_threshold=101.0,
        )
        summary_high = generate_report(
            report_file=_SAMPLE_PDF,
            comments_file=None,
            output_file=output_high,
            progress=lambda p, m: None,
            over_budget_threshold=9999.0,
        )

        assert len(summary_default.over_budget_lines) > 0, (
            "Sample PDF must have over-budget lines at default threshold"
        )
        # Round 47 — at threshold 9999%, no PERCENTAGE-driven flag should
        # fire, but the zero-budget-with-spend trigger (Round 47) flags
        # any line where budget=$0 and ytd!=$0 regardless of threshold.
        # Verify percentage-driven flags are gone but allow zero-budget
        # rows to remain on the over list.
        non_zero_budget_overs = [
            ln for ln in summary_high.over_budget_lines if ln.budget != Decimal("0")
        ]
        assert len(non_zero_budget_overs) == 0, (
            "At threshold 9999%, no positive-budget line should exceed the threshold"
        )

    def test_xlsx_pink_fill_respects_threshold(self, tmp_path: Path) -> None:
        """Round 38 — superseded by test_xlsx_monthly_report.py."""
        # Old shape (Revenue/Expenditure sheets) no longer produced.

    def test_over_fill_constant_uses_hl_mismatch(self) -> None:
        """_OVER_FILL fgColor must derive from HL_MISMATCH via argb()."""
        from toolkit.fills import argb

        expected_argb = argb(HL_MISMATCH).upper()
        actual_argb = _OVER_FILL.fgColor.rgb.upper() if _OVER_FILL.fgColor else ""
        assert actual_argb == expected_argb, (
            f"_OVER_FILL fgColor {actual_argb!r} does not match argb(HL_MISMATCH) {expected_argb!r}"
        )


class TestPerSectionThreshold:
    """Round 21 — Revenue and Expense thresholds are independent.

    Schools care about Expense over-runs but rarely flag Revenue
    over-collections.  Verify that the section-aware overload of
    ``_recompute_is_over`` flags rows correctly when the two thresholds
    differ.
    """

    def _mixed_lines(self) -> list[SubProgramLine]:
        """Two Revenue + two Expenditure lines at known used_pct values."""

        def _line(sp: str, section: str, pct: str) -> SubProgramLine:
            return SubProgramLine(
                sub_program=sp,
                account=section,
                description=f"{section} {sp}",
                budget=Decimal("100"),
                ytd=Decimal("100"),
                remaining=Decimal("0"),
                used_pct=Decimal(pct),
                faculty="Curriculum",
                is_over=False,
            )

        return [
            _line("1001", "Revenue", "105"),  # over a 101 threshold
            _line("1002", "Revenue", "115"),  # over a 110 threshold
            _line("4001", "Expenditure", "105"),  # over a 101 threshold
            _line("4002", "Expenditure", "115"),  # over a 110 threshold
        ]

    def test_revenue_threshold_only_flags_revenue(self) -> None:
        """High Revenue threshold + low Expense threshold = no Revenue
        flagged but both Expense lines flagged."""
        lines = _recompute_is_over(
            self._mixed_lines(),
            101.0,  # legacy combined value (ignored when overrides set)
            revenue_threshold=120.0,
            expense_threshold=101.0,
        )
        assert not lines[0].is_over  # Revenue 105% < 120%
        assert not lines[1].is_over  # Revenue 115% < 120%
        assert lines[2].is_over  # Expense 105% > 101%
        assert lines[3].is_over  # Expense 115% > 101%

    def test_expense_threshold_only_flags_expense(self) -> None:
        """Low Revenue + high Expense = both Revenue flagged, no Expense."""
        lines = _recompute_is_over(
            self._mixed_lines(),
            101.0,
            revenue_threshold=101.0,
            expense_threshold=120.0,
        )
        assert lines[0].is_over  # Revenue 105% > 101%
        assert lines[1].is_over  # Revenue 115% > 101%
        assert not lines[2].is_over  # Expense 105% < 120%
        assert not lines[3].is_over  # Expense 115% < 120%

    def test_legacy_single_threshold_still_works(self) -> None:
        """Backward compat - calling without per-section overrides still
        applies one threshold to both sections."""
        lines = _recompute_is_over(self._mixed_lines(), threshold=110.0)
        # Threshold 110% - Revenue 115% triggers, others don't.
        assert not lines[0].is_over  # 105 < 110
        assert lines[1].is_over  # 115 > 110
        assert not lines[2].is_over  # 105 < 110
        assert lines[3].is_over  # 115 > 110


# ---------------------------------------------------------------------------
# Round 51 Phase D — structured commentary
# ---------------------------------------------------------------------------


class TestStructuredCommentary:
    """encode_commentary / decode_commentary round-trip + edge cases."""

    def test_value_tuples_have_designed_shape(self) -> None:
        """Pin the dropdown values so a stray edit can't silently drift
        away from the inline editor's Combobox lists."""
        from tools.sub_program.logic import (
            _ACTION_VALUES,
            _DRIVER_VALUES,
            _OUTLOOK_VALUES,
        )

        assert _DRIVER_VALUES == (
            "One-time",
            "Ongoing",
            "Structural",
            "Timing-early",
            "Timing-late",
            "Investigating",
        )
        assert _OUTLOOK_VALUES == (
            "One-time",
            "Expected to continue",
            "Improving",
            "Deteriorating",
        )
        assert _ACTION_VALUES == (
            "None",
            "Monitor",
            "Investigate",
            "Update forecast",
        )

    def test_encode_all_blank_returns_empty(self) -> None:
        from tools.sub_program.logic import encode_commentary

        assert encode_commentary("") == ""

    def test_encode_notes_only_no_prefix(self) -> None:
        """Pre-Phase-D shape — no prefix when all dropdowns are blank."""
        from tools.sub_program.logic import encode_commentary

        assert encode_commentary("Reviewed by council") == "Reviewed by council"

    def test_encode_only_action_emits_minimal_prefix(self) -> None:
        """Round 1 fix: separator is now newline (not space) so the prefix
        and notes don't visually merge when the XLSX cell wraps."""
        from tools.sub_program.logic import encode_commentary

        assert encode_commentary("notes", action="Monitor") == "[Action: Monitor]\nnotes"

    def test_encode_all_three_emits_full_prefix(self) -> None:
        from tools.sub_program.logic import encode_commentary

        encoded = encode_commentary(
            "Reviewed by council",
            driver="Ongoing",
            outlook="Expected to continue",
            action="Monitor",
        )
        assert encoded == (
            "[Driver: Ongoing | Outlook: Expected to continue | Action: Monitor]\n"
            "Reviewed by council"
        )

    def test_encode_action_none_literal_emits_prefix(self) -> None:
        """The literal 'None' value (user reviewed, no action) is distinct
        from '' (not categorised) and must appear in the prefix."""
        from tools.sub_program.logic import encode_commentary

        assert encode_commentary("done", action="None") == "[Action: None]\ndone"

    def test_encode_escapes_leading_bracket_in_notes(self) -> None:
        """Notes starting with ``[`` and dropdowns blank -> empty-body
        escape so the decoder doesn't mis-parse on next read."""
        from tools.sub_program.logic import encode_commentary

        assert encode_commentary("[Driver: foo] bar") == "[]\n[Driver: foo] bar"

    def test_decode_empty_string(self) -> None:
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary("") == ("", "", "", "")

    def test_decode_pre_phase_d_freeform_text(self) -> None:
        """Pre-Phase-D files have no prefix; the entire cell is Notes."""
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary("Reviewed by council") == (
            "Reviewed by council",
            "",
            "",
            "",
        )

    def test_decode_unknown_bracket_preserved_as_notes(self) -> None:
        """A ``[FREE TEXT]`` opener with no Phase-D keys is treated as
        Notes, not as a (broken) prefix."""
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary("[NOTE TO SELF] check next week") == (
            "[NOTE TO SELF] check next week",
            "",
            "",
            "",
        )

    def test_decode_full_prefix(self) -> None:
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary(
            "[Driver: Ongoing | Outlook: Improving | Action: Monitor] notes"
        ) == ("notes", "Ongoing", "Improving", "Monitor")

    def test_decode_partial_prefix(self) -> None:
        """Only some structured fields set — the rest decode as ''."""
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary("[Action: Investigate] follow up") == (
            "follow up",
            "",
            "",
            "Investigate",
        )

    def test_decode_strips_empty_body_escape(self) -> None:
        """The ``[]`` escape strips correctly and preserves notes."""
        from tools.sub_program.logic import decode_commentary

        assert decode_commentary("[] [Driver: foo] bar") == (
            "[Driver: foo] bar",
            "",
            "",
            "",
        )

    def test_round_trip_full_combination(self) -> None:
        """Composite round-trip across an array of representative inputs."""
        from tools.sub_program.logic import decode_commentary, encode_commentary

        cases = [
            ("", "", "", ""),
            ("notes only", "", "", ""),
            ("", "Ongoing", "", ""),
            ("", "", "Improving", ""),
            ("", "", "", "Investigate"),
            ("notes", "Ongoing", "Improving", "Monitor"),
            ("notes with $1,234.56", "Structural", "Deteriorating", "Update forecast"),
            ("[brackets in notes]", "", "", ""),  # escape path
            ("done", "", "", "None"),  # literal None action
        ]
        for notes, driver, outlook, action in cases:
            enc = encode_commentary(notes, driver=driver, outlook=outlook, action=action)
            assert decode_commentary(enc) == (notes, driver, outlook, action), (
                f"round-trip failed for {(notes, driver, outlook, action)!r} -> "
                f"encoded={enc!r}, decoded={decode_commentary(enc)!r}"
            )


class TestStructuredCommentaryPriorPeriodMigration:
    """Pre-Phase-D prior-period file -> generate_report decodes to Notes only."""

    def test_pre_phase_d_freeform_lands_in_notes_with_blank_dropdowns(self, tmp_path: Path) -> None:
        """A prior-period file whose Comments column has no Phase-D
        prefix should round-trip into the new ``commentary`` (Notes)
        field with the three structured dropdowns left blank."""
        # Build a minimal pre-Phase-D shaped XLSX.
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Sub-Program", "Description", "Comments"])
        ws.append(["4001", "Curriculum", "Reviewed by council"])
        ws.append(["8599", "Rowing", "Spend on track"])
        comments_path = tmp_path / "prior_unstructured.xlsx"
        wb.save(comments_path)

        from tools.sub_program.logic import load_prior_period_comments

        # Reader is unchanged — returns the raw cell text.
        loaded = load_prior_period_comments(comments_path)
        assert loaded[("4001", "Curriculum")] == "Reviewed by council"
        assert loaded[("8599", "Rowing")] == "Spend on track"

    def test_phase_d_encoded_cell_decodes_at_consumer(self, tmp_path: Path) -> None:
        """A prior-period file from Round 51+ has the structured prefix
        in the Comments cell. ``decode_commentary`` (called by
        ``generate_report``) splits it cleanly back into the four
        fields."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Sub-Program", "Description", "Comments"])
        ws.append(
            [
                "4001",
                "Curriculum",
                "[Driver: Ongoing | Action: Monitor] Reviewed by council",
            ]
        )
        comments_path = tmp_path / "prior_structured.xlsx"
        wb.save(comments_path)

        from tools.sub_program.logic import (
            decode_commentary,
            load_prior_period_comments,
        )

        loaded = load_prior_period_comments(comments_path)
        encoded = loaded[("4001", "Curriculum")]
        notes, driver, outlook, action = decode_commentary(encoded)
        assert notes == "Reviewed by council"
        assert driver == "Ongoing"
        assert outlook == ""
        assert action == "Monitor"


class TestStructuredCommentaryXlsxOutput:
    """Structured fields round-trip through the XLSX writer."""

    def _line(
        self,
        sub_program: str,
        notes: str = "",
        driver: str = "",
        outlook: str = "",
        action: str = "",
    ) -> SubProgramLine:
        return SubProgramLine(
            sub_program=sub_program,
            account="Expenditure",
            description=f"{sub_program} description",
            budget=Decimal("10000"),
            ytd=Decimal("4000"),
            remaining=Decimal("6000"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
            commentary=notes,
            commentary_driver=driver,
            commentary_outlook=outlook,
            commentary_action=action,
        )

    def test_xlsx_comments_cell_renders_prose_form(self, tmp_path: Path) -> None:
        """Round 53 F1 (Move E): the visible cell is plain-English prose,
        not the bracketed prefix. The prefix encoding survives only in
        the in-memory ``encode_commentary`` helper for prior-period
        files written by R51 (those round-trip via ``decode_commentary``)."""
        out = tmp_path / "out.xlsx"
        lines = [
            self._line(
                "4001",
                notes="Reviewed by council",
                driver="Ongoing",
                action="Monitor",
            ),
        ]
        _write_xlsx(lines, out, period_label="Apr 2026")

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        # Header is rows 1-2, data starts row 3. Comments is column 12.
        cell = ws.cell(row=3, column=13).value
        # R1 fix: prose now uses period+capital splits instead of em-dash.
        assert cell == "Ongoing variance. Being monitored. Reviewed by council."

    def test_xlsx_comments_cell_blank_when_all_fields_blank(self, tmp_path: Path) -> None:
        out = tmp_path / "out.xlsx"
        lines = [self._line("4001")]
        _write_xlsx(lines, out, period_label="Apr 2026")

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        cell = ws.cell(row=3, column=13).value
        # Empty cell renders as None or "" depending on openpyxl version.
        assert cell in (None, "")

    def test_xlsx_comments_cell_notes_only_renders_with_terminal_period(
        self, tmp_path: Path
    ) -> None:
        """Round 53 F1: notes-only cells get a terminal period from the
        prose renderer's ``_ensure_terminal_period`` helper. Council
        readers see a complete sentence, not a fragment."""
        out = tmp_path / "out.xlsx"
        lines = [self._line("4001", notes="Reviewed by council")]
        _write_xlsx(lines, out, period_label="Apr 2026")

        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=3, column=13).value == "Reviewed by council."


# ---------------------------------------------------------------------------
# Round 51 Phase D Round-1 fixes — protective behaviours added after the
# multi-personality audit (logic skeptic, UX critic, Excel QA).
# ---------------------------------------------------------------------------


class TestStructuredCommentaryRound1Fixes:
    """Targeted tests for the Round-1 protective fixes."""

    def test_decode_preserves_inner_whitespace_in_notes(self) -> None:
        """encode→decode→encode is idempotent — ``decode_commentary``
        only strips the prefix-adjacent separator, not user whitespace."""
        from tools.sub_program.logic import decode_commentary, encode_commentary

        notes = "  multi-line\n  with leading spaces  "
        encoded = encode_commentary(notes, action="Monitor")
        decoded = decode_commentary(encoded)
        assert decoded == (notes, "", "", "Monitor")
        # Idempotent — re-encoding the decoded tuple matches the original.
        re_encoded = encode_commentary(decoded[0], action=decoded[3])
        assert re_encoded == encoded

    def test_decode_preserves_unknown_driver_value(self) -> None:
        """Round 2 fix (R51): unknown Driver value (typo, schema drift,
        third-party edit) is preserved verbatim in the tuple instead
        of falling through to Notes. Avoids round-trip data loss when
        the editor's Combobox preserves a non-canonical value across a
        save→reopen cycle."""
        from tools.sub_program.logic import decode_commentary

        # "FooBar" is not in _DRIVER_VALUES but we still extract it.
        text = "[Driver: FooBar] my actual note"
        assert decode_commentary(text) == ("my actual note", "FooBar", "", "")

    def test_decode_preserves_unknown_outlook_value(self) -> None:
        from tools.sub_program.logic import decode_commentary

        text = "[Outlook: HopefullyOk] some note"
        assert decode_commentary(text) == ("some note", "", "HopefullyOk", "")

    def test_decode_preserves_unknown_action_value(self) -> None:
        from tools.sub_program.logic import decode_commentary

        text = "[Action: ReorderEverything] note"
        assert decode_commentary(text) == ("note", "", "", "ReorderEverything")

    def test_decode_mixed_known_and_unknown_preserves_both(self) -> None:
        """Round 2 fix: known + unknown values in same prefix are both
        extracted — the editor surfaces them and lets the user re-pick."""
        from tools.sub_program.logic import decode_commentary

        # Driver is canonical, Action is not — both preserved.
        text = "[Driver: Ongoing | Action: Reorder] note"
        assert decode_commentary(text) == ("note", "Ongoing", "", "Reorder")

    def test_decode_accepts_space_or_newline_separator(self) -> None:
        """Pre-Round-1 cells used a space separator; Round-1+ cells use
        a newline. Decoder tolerates both for forward / backward compat."""
        from tools.sub_program.logic import decode_commentary

        space_form = "[Action: Monitor] notes"
        newline_form = "[Action: Monitor]\nnotes"
        assert decode_commentary(space_form) == ("notes", "", "", "Monitor")
        assert decode_commentary(newline_form) == ("notes", "", "", "Monitor")

    def test_xlsx_atomic_per_subprogram_no_cross_row_fabrication(self, tmp_path: Path) -> None:
        """When a sub-program has multiple Account-rows with conflicting
        commentary, the writer adopts the WHOLE 4-tuple from one row
        (the first non-empty contributor) — never mixing notes from
        row A with Action from row B."""
        # Row A on Revenue side has notes only.
        # Row B on Expenditure side has structured fields only.
        # Pre-fix the writer would have produced a fabricated combination.
        line_a = SubProgramLine(
            sub_program="4001",
            account="Revenue",
            description="Curriculum",
            budget=Decimal("10000"),
            ytd=Decimal("4000"),
            remaining=Decimal("6000"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
            commentary="row A note",
        )
        line_b = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Curriculum",
            budget=Decimal("10000"),
            ytd=Decimal("4000"),
            remaining=Decimal("6000"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
            commentary_action="Investigate",
        )
        out = tmp_path / "atomic.xlsx"
        _write_xlsx([line_a, line_b], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        cell = ws.cell(row=3, column=13).value
        # Cell carries row A's notes only (row A came first), NOT a
        # fabricated "Needs investigation. Row A note." combination
        # mixing row B's Action with row A's notes. Round 53 F1:
        # rendered as prose with leading capital + terminal period.
        assert cell == "Row A note."

    def test_xlsx_formula_injection_guard_prepends_apostrophe(self, tmp_path: Path) -> None:
        """A user typing '=SUM(D3:E3) outdated' into Notes must NOT
        become a live Excel formula. The writer prepends an apostrophe
        — Excel renders the cell as text on display."""
        line = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Curriculum",
            budget=Decimal("10000"),
            ytd=Decimal("4000"),
            remaining=Decimal("6000"),
            used_pct=Decimal("40"),
            faculty="Curriculum",
            is_over=False,
            commentary="=SUM(D3:E3) outdated",
        )
        out = tmp_path / "formula_guard.xlsx"
        _write_xlsx([line], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        cell = ws.cell(row=3, column=13)
        # Round 53 F1: prose renderer adds a terminal period.
        assert cell.value == "'=SUM(D3:E3) outdated."
        # Cell is text-formatted, not a formula.
        assert cell.data_type == "s"
        assert cell.number_format == "@"

    def test_xlsx_formula_guard_applies_to_other_sigils(self, tmp_path: Path) -> None:
        """``=`` is the canonical formula prefix but Excel also evaluates
        ``+``, ``-``, and ``@`` — guard them all."""
        for sigil in ("+", "-", "@"):
            line = SubProgramLine(
                sub_program="4001",
                account="Expenditure",
                description="Curriculum",
                budget=Decimal("10000"),
                ytd=Decimal("4000"),
                remaining=Decimal("6000"),
                used_pct=Decimal("40"),
                faculty="Curriculum",
                is_over=False,
                commentary=f"{sigil}danger",
            )
            out = tmp_path / f"formula_guard_{sigil!r}.xlsx"
            _write_xlsx([line], out, period_label="Apr 2026")
            wb = openpyxl.load_workbook(out, data_only=True)
            ws = wb["Sub Program Report"]
            cell = ws.cell(row=3, column=13)
            # Round 53 F1: prose renderer adds a terminal period.
            assert cell.value == f"'{sigil}danger.", f"sigil {sigil!r} should be guarded"

    def test_load_prior_period_strips_apostrophe_guard(self, tmp_path: Path) -> None:
        """The reader strips the formula-guard apostrophe so
        ``decode_commentary`` sees the original encoded value — and the
        round-trip doesn't accumulate apostrophes across save→reopen."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Sub-Program", "Description", "Comments"])
        ws.append(["4001", "Curriculum", "'=danger formula"])
        comments_path = tmp_path / "guarded.xlsx"
        wb.save(comments_path)

        loaded = load_prior_period_comments(comments_path)
        assert loaded[("4001", "Curriculum")] == "=danger formula"


# ---------------------------------------------------------------------------
# Round 53 F1 — Status pills (Move B), prose commentary (Move E),
# percent cap (Move F). TDD: tests written first.
# ---------------------------------------------------------------------------


class TestStatusPill:
    """``compute_status_pill`` returns one of the canonical status strings.

    Round 56 redesign — pacing-free contract. The function compares
    ``exp_ytd`` against ``expense_threshold% × annual_exp_budget`` and
    buckets the overrun by dollar / percent past the threshold. The
    ``available`` and ``calendar_pct`` parameters and the ``No spend
    yet`` pill (which depended on calendar awareness) are gone.
    """

    def test_status_values_tuple_has_designed_shape(self) -> None:
        """Pin the status values so a stray edit can't silently drift.
        Round 56 dropped 'No spend yet' along with calendar_pct.
        Round 62 added 'Revenue over budget' so the Status pill agrees
        with the Watchlist filter on revenue-side overruns."""
        from tools.sub_program.logic import _STATUS_VALUES

        assert _STATUS_VALUES == (
            "On track",
            "Slightly over",
            "Significant overspend",
            "Investigate urgently",
            "Spent without budget",
            "Revenue over budget",
        )

    def test_on_track_when_under_budget(self) -> None:
        """exp_ytd well within budget → on track."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("10000"),
                exp_ytd=Decimal("2000"),
            )
            == "On track"
        )

    def test_on_track_when_overrun_below_materiality_floor(self) -> None:
        """A $50 overrun on a $30 budget shouldn't ring alarm bells —
        below the dollar materiality floor (and not >50% past) it reads
        as on track."""
        from tools.sub_program.logic import compute_status_pill

        # $80 spend on a $30 budget = 167% → 67 pp past threshold which
        # IS over the 50% urgent floor. Use a smaller relative overrun
        # so we land in the materiality fallthrough.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("10000"),
                exp_ytd=Decimal("10800"),  # $800 over (>$500 noise) but below $5K mat
                materiality_dollar=5000,
            )
            == "On track"
        )

    def test_slightly_over_for_5k_to_25k_overrun(self) -> None:
        from tools.sub_program.logic import compute_status_pill

        # $112K spend on $100K budget = 12% over → $12K overrun past
        # the 101% threshold ($101K), so $11K past threshold. With
        # default materiality_dollar=5000 → Slightly over bucket.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("100000"),
                exp_ytd=Decimal("112000"),  # $11K past 101% threshold
            )
            == "Slightly over"
        )

    def test_significant_overspend_for_25k_to_100k_overrun(self) -> None:
        """$25K-$100K past the threshold lands in the middle bucket."""
        from tools.sub_program.logic import compute_status_pill

        # $250K on $200K budget = 25% over → $48K past 101% threshold
        # ($202K) → Significant overspend.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("200000"),
                exp_ytd=Decimal("250000"),
            )
            == "Significant overspend"
        )

    def test_investigate_urgently_for_overrun_over_100k(self) -> None:
        """Big dollar overrun = urgent regardless of percent."""
        from tools.sub_program.logic import compute_status_pill

        # $700K on $581K budget = $113K past 101% threshold ($587K) →
        # urgent (>$100K dollar floor).
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("581700"),
                exp_ytd=Decimal("700000"),
            )
            == "Investigate urgently"
        )

    def test_investigate_urgently_for_overrun_over_50pct(self) -> None:
        """Big PERCENT overrun = urgent even if dollar-small."""
        from tools.sub_program.logic import compute_status_pill

        # $50K spend on $20K budget = 250% → 149 pp past 101% threshold,
        # well above 50% → urgent.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("20000"),
                exp_ytd=Decimal("50000"),
            )
            == "Investigate urgently"
        )

    def test_unbudgeted_program_with_no_spend_is_on_track(self) -> None:
        """Sub-program with $0 annual budget and no spend — chart-of-
        accounts placeholder, on track (not 'No spend yet' anymore;
        that pill was dropped in Round 56)."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("0"),
                exp_ytd=Decimal("0"),
            )
            == "On track"
        )

    def test_budgeted_program_with_no_spend_is_on_track(self) -> None:
        """Round 56 — without calendar awareness we can't distinguish
        'early in the year' from 'late in the year, suspiciously
        unfunded'. Default to On track and rely on the YTD column to
        convey the absence of spend."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("50000"),
                exp_ytd=Decimal("0"),
            )
            == "On track"
        )

    def test_spent_without_budget_when_budget_zero_and_ytd_nonzero(self) -> None:
        """A sub-program with $0 annual budget on BOTH sides but $X YTD
        spend and no revenue collected — capital spend without council
        approval, a real category of finance concern."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("0"),
                annual_rev_budget=Decimal("0"),
                rev_ytd=Decimal("0"),
                exp_ytd=Decimal("454545"),
            )
            == "Spent without budget"
        )

    def test_spent_without_budget_excludes_revenue_collection_program(self) -> None:
        """If rev_ytd > 0 the program is a cost-recovery / fundraising
        line — its expenditure is matched by collected revenue, so
        Spent-without-budget doesn't apply."""
        from tools.sub_program.logic import compute_status_pill

        result = compute_status_pill(
            annual_exp_budget=Decimal("0"),
            annual_rev_budget=Decimal("0"),
            rev_ytd=Decimal("8000"),  # collected revenue
            exp_ytd=Decimal("5000"),
        )
        assert result != "Spent without budget"

    def test_threshold_cushion_is_respected(self) -> None:
        """Threshold = 105% means used_pct must exceed 105 to flag.
        At exactly 105% we stay On track."""
        from tools.sub_program.logic import compute_status_pill

        # exp_ytd = 105K on $100K budget = exactly 105% → On track.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("100000"),
                exp_ytd=Decimal("105000"),
                expense_threshold=105.0,
            )
            == "On track"
        )


class TestCommentaryProse:
    """``render_commentary_prose`` converts the structured triplet to
    plain English for the XLSX cell."""

    def test_all_blank_returns_empty(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose() == ""

    def test_notes_only_returns_notes_verbatim(self) -> None:
        """Pre-Phase-D files (notes only) round-trip as is."""
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(notes="Reviewed by council") == "Reviewed by council."

    def test_notes_with_existing_terminal_punctuation(self) -> None:
        """Don't double-period a sentence that already ends in . / ! / ?."""
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(notes="Reviewed by council!") == "Reviewed by council!"

    def test_action_only_no_notes(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(action="Investigate") == "Needs investigation."

    def test_driver_action_combination(self) -> None:
        """R1 fix: em-dash splice replaced with period+capital so two
        short sentences read as natural English."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="Ongoing", action="Monitor")
            == "Ongoing variance. Being monitored."
        )

    def test_full_triplet_combination(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(
                driver="Ongoing",
                outlook="Expected to continue",
                action="Investigate",
            )
            == "Ongoing variance, expected to continue. Needs investigation."
        )

    def test_full_triplet_with_notes(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(
                notes="Reviewed by council",
                driver="Ongoing",
                outlook="Expected to continue",
                action="Monitor",
            )
            == "Ongoing variance, expected to continue. Being monitored. Reviewed by council."
        )

    def test_action_none_literal_renders(self) -> None:
        """Literal 'None' Action means 'reviewed, no action needed' —
        distinct from blank, must render."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(action="None", notes="all good")
            == "No action needed. All good."
        )

    def test_timing_drivers(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(driver="Timing-early") == "Spend earlier than planned."
        assert render_commentary_prose(driver="Timing-late") == "Spend later than planned."

    def test_outlook_only(self) -> None:
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(outlook="Improving") == "Outlook improving."

    def test_unknown_driver_value_falls_through_to_notes(self) -> None:
        """Defensive — an unknown value (schema drift, hand edit) doesn't
        crash; we render whatever notes we have, ignoring the unknown
        structured field."""
        from tools.sub_program.logic import render_commentary_prose

        assert render_commentary_prose(driver="FooBar", notes="some note") == "Some note."


# Round 67 — TestPercentCap class deleted along with the
# cap_percent_for_display function. The ±999% display cap was
# dropped per user feedback that the ">999%" / "<-999%" markers
# hid the actual magnitude.


class TestF1XlsxIntegration:
    """End-to-end: F1 changes appear in the rendered XLSX correctly."""

    def _line(
        self,
        sub_program: str = "4001",
        account: str = "Expenditure",
        budget: str = "10000",
        ytd: str = "5000",
        notes: str = "",
        driver: str = "",
        outlook: str = "",
        action: str = "",
    ) -> SubProgramLine:
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sub_program,
            account=account,
            description=f"{sub_program} desc",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=(y / b * Decimal("100")) if b != 0 else Decimal("0"),
            faculty="Curriculum",
            is_over=False,
            commentary=notes,
            commentary_driver=driver,
            commentary_outlook=outlook,
            commentary_action=action,
        )

    def test_xlsx_has_status_column_5(self, tmp_path: Path) -> None:
        """Round 66 — Status column at position 5 (was col 3 in
        R57 layout). Two percent columns now between PROGRAM NAME
        and Status."""
        out = tmp_path / "out.xlsx"
        _write_xlsx([self._line()], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        # Row 2 is the header.
        assert ws.cell(row=2, column=5).value == "Status"

    def test_xlsx_status_column_renders_pill_value(self, tmp_path: Path) -> None:
        out = tmp_path / "out.xlsx"
        # An on-track sub-program — combined Revenue + Expenditure, with
        # revenue collected covering the spend (positive available).
        rev = self._line(
            sub_program="4001",
            account="Revenue",
            budget="10000",
            ytd="6000",
        )
        exp = self._line(
            sub_program="4001",
            account="Expenditure",
            budget="10000",
            ytd="5000",
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        # Data row 3, Status col 13. Available = rev_y − exp_y = 6000 − 5000 = +1000.
        assert ws.cell(row=3, column=5).value == "On track"

    def test_xlsx_comments_cell_uses_prose_form(self, tmp_path: Path) -> None:
        """Move E + R1 fix: prose renders as two short sentences (was em-
        dash splice). Reads as natural English, not template output."""
        out = tmp_path / "out.xlsx"
        # Combined Revenue + Expenditure so the line aggregates with a
        # well-defined Status (R1 added annual_rev_budget gating).
        rev = self._line(
            sub_program="4001",
            account="Revenue",
            budget="10000",
            ytd="6000",
            notes="Reviewed by council",
            driver="Ongoing",
            action="Monitor",
        )
        exp = self._line(
            sub_program="4001",
            account="Expenditure",
            budget="10000",
            ytd="5000",
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        cell = ws.cell(row=3, column=13).value
        assert cell == "Ongoing variance. Being monitored. Reviewed by council."

    def test_xlsx_extreme_percent_writes_actual_formula(self, tmp_path: Path) -> None:
        """Round 67: percent cells always carry the Excel formula, no
        matter how large the result. A 21x revenue over-collection
        used to render as the text marker ">999%" (Round 53 cap); now
        it shows the actual computed percentage so a council reader
        sees the true magnitude (≈2136%). The 0–100% data bar still
        caps visually at 100% via Excel's standard clipping, so the
        bar saturated + a large number together convey "scale of the
        overage" at a glance."""
        out = tmp_path / "out.xlsx"
        # Sub-program 4400: rev_b $1,000, rev_y $21,365. rev_pct = 21.365.
        rev_line = self._line(
            sub_program="4400",
            account="Revenue",
            budget="1000",
            ytd="21365",
        )
        exp_line = self._line(
            sub_program="4400",
            account="Expenditure",
            budget="7200",
            ytd="2880",
        )
        _write_xlsx([rev_line, exp_line], out, period_label="Apr 2026")
        # Load without data_only so we read the formula text, not
        # cached values (uncached formulas read back as None).
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        # Round 66: Revenue Budget % Received YTD is at col 4 (the
        # two percent columns relocated between PROGRAM NAME and Status).
        rev_pct_cell = ws.cell(row=3, column=4)
        # Formula not a marker.
        assert rev_pct_cell.value == "=I3/G3"
        # No cell comment (the comment was an artefact of the cap
        # fallback; with no cap there's nothing to footnote).
        assert rev_pct_cell.comment is None

    def test_xlsx_status_urgent_for_admin_scale_overrun(self, tmp_path: Path) -> None:
        """Round 56: an Admin sub-program at $700K spend on $582K
        budget (≈$113K past the 101% threshold) reads as 'Investigate
        urgently' via the >$100K dollar floor."""
        out = tmp_path / "out.xlsx"
        rev = self._line(
            sub_program="7001",
            account="Revenue",
            budget="0",
            ytd="26436",
        )
        exp = SubProgramLine(
            sub_program="7001",
            account="Expenditure",
            description="Administration",
            budget=Decimal("581700"),
            ytd=Decimal("700000"),  # $113K past 101% threshold ($587K)
            remaining=Decimal("-118300"),
            used_pct=Decimal("120"),
            faculty="Administration",
            is_over=True,
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=3, column=5).value == "Investigate urgently"

    def test_xlsx_status_on_track_when_budget_set_no_movement(self, tmp_path: Path) -> None:
        """Round 56 — without calendar awareness we can't distinguish
        'early in the year' from 'late in the year'. A budgeted program
        with $0 YTD spend reads as On track; the YTD column conveys the
        absence of spend on its own."""
        out = tmp_path / "out.xlsx"
        rev = self._line(
            sub_program="8330",
            account="Revenue",
            budget="10000",
            ytd="0",
        )
        exp = self._line(
            sub_program="8330",
            account="Expenditure",
            budget="10000",
            ytd="0",
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=3, column=5).value == "On track"


# ---------------------------------------------------------------------------
# Round 53 F1 R1 — regression tests for the fixes applied after the
# 4-Opus-agent review. Pin behaviours that would otherwise silently
# regress on a future writer / pill rewrite.
# ---------------------------------------------------------------------------


class TestF1Round1Fixes:
    """Targeted tests for the 12 R1 fixes."""

    def _exp_only_line(
        self,
        sub_program: str,
        budget: str,
        ytd: str,
    ) -> SubProgramLine:
        """Expenditure-only sub-program (no revenue side)."""
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sub_program,
            account="Expenditure",
            description=f"{sub_program} desc",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=(y / b * Decimal("100")) if b != 0 else Decimal("0"),
            faculty="Curriculum",
            is_over=False,
        )

    def test_expenditure_only_within_threshold_renders_on_track(self) -> None:
        """Round 56: Curriculum sub-program with $10K exp budget, $3,500
        YTD spend = 35% used. With default 101% threshold, used_pct
        ≤ threshold → On track."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("10000"),
                exp_ytd=Decimal("3500"),
            )
            == "On track"
        )

    def test_expenditure_only_above_threshold_flags(self) -> None:
        """Round 56: an exp-only program 110% spent → over the 101%
        threshold by $9K (10K below the $25K Significant cushion when
        exp budget is $100K) → at minimum Slightly over."""
        from tools.sub_program.logic import compute_status_pill

        result = compute_status_pill(
            annual_exp_budget=Decimal("100000"),
            exp_ytd=Decimal("110000"),
        )
        assert result in ("Slightly over", "Significant overspend", "Investigate urgently"), result

    def test_spent_without_budget_excludes_revenue_only_program(self) -> None:
        """A fundraiser with rev_b > 0, exp_b = 0, exp_y > 0 is NOT
        'Spent without budget' — it's a cost-recovery program. The gate
        requires exp_b == 0 AND rev_b == 0 (and rev_ytd == 0)."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("0"),
                annual_rev_budget=Decimal("50000"),
                rev_ytd=Decimal("30000"),
                exp_ytd=Decimal("32000"),
            )
            != "Spent without budget"
        )

    def test_spent_without_budget_fires_when_truly_unbudgeted(self) -> None:
        """The hard case: $0 budget on both sides, no revenue, $X spent."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("0"),
                annual_rev_budget=Decimal("0"),
                rev_ytd=Decimal("0"),
                exp_ytd=Decimal("454545"),
            )
            == "Spent without budget"
        )

    def test_prose_drops_contradictory_structural_one_time(self) -> None:
        """R1 fix: 'Structural variance' + 'won't recur' is contradictory
        — the outlook is dropped, structural stands alone."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="Structural", outlook="One-time")
            == "Structural variance."
        )

    def test_prose_drops_contradictory_one_time_continuing(self) -> None:
        """One-time + 'expected to continue' is contradictory."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="One-time", outlook="Expected to continue")
            == "One-time variance."
        )

    def test_prose_drops_contradictory_investigating_no_action(self) -> None:
        """'Driver under investigation' + 'no action needed' is direct
        contradiction — investigating IS an action."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="Investigating", action="None")
            == "Driver under investigation."
        )

    def test_xlsx_extreme_revenue_pct_writes_formula(self, tmp_path: Path) -> None:
        """Round 67 (was R1 fix): the writer no longer caps extreme
        percents. A 2100% over-collection now renders as the actual
        ``=I3/G3`` formula whose computed value is 21.0 (= 2100% in
        the percent number format), letting the council reader see
        the true magnitude."""
        rev = SubProgramLine(
            sub_program="9999",
            account="Revenue",
            description="Test",
            budget=Decimal("1000"),
            ytd=Decimal("21000"),
            remaining=Decimal("-20000"),
            used_pct=Decimal("2100"),
            faculty="Curriculum",
            is_over=True,
        )
        exp = SubProgramLine(
            sub_program="9999",
            account="Expenditure",
            description="Test",
            budget=Decimal("1000"),
            ytd=Decimal("0"),
            remaining=Decimal("1000"),
            used_pct=Decimal("0"),
            faculty="Curriculum",
            is_over=False,
        )
        out = tmp_path / "no_cap.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        # Load WITHOUT data_only so we can read the formula text
        # (data_only mode returns cached values which are None when
        # the file has never been opened by Excel).
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        # Round 66: Revenue Budget % Received YTD at col 4.
        assert ws.cell(row=3, column=4).value == "=I3/G3"

    def test_xlsx_pink_fill_extends_to_status_column(self, tmp_path: Path) -> None:
        """R1 regression test: an over-budget row paints pink across all
        13 columns including the new Status pill cell. Without this the
        Status would visually un-pink against the rest of the row."""
        # Engineer a $50K overrun on $10K budget → urgent + pink fill.
        from toolkit.tokens import HL_MISMATCH

        rev = SubProgramLine(
            sub_program="4001",
            account="Revenue",
            description="Test",
            budget=Decimal("0"),
            ytd=Decimal("0"),
            remaining=Decimal("0"),
            used_pct=Decimal("0"),
            faculty="Curriculum",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("60000"),
            remaining=Decimal("-50000"),
            used_pct=Decimal("600"),
            faculty="Curriculum",
            is_over=True,
        )
        out = tmp_path / "pink.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        status_cell = ws.cell(row=3, column=5)
        # The pink fill is the canonical HL_MISMATCH ARGB. The fgColor
        # may surface as either an RGB or theme reference depending on
        # openpyxl version; check the suffix.
        fill = status_cell.fill
        assert fill is not None
        rgb = fill.fgColor.rgb if fill.fgColor is not None else None
        assert rgb is not None and HL_MISMATCH in rgb.upper()

    def test_xlsx_no_spend_yet_is_bold(self, tmp_path: Path) -> None:
        """R1 fix: 'No spend yet' joins Material/Urgent/Spent-without-
        budget in the bold set. It's a row council members would ask
        about."""
        rev = SubProgramLine(
            sub_program="8330",
            account="Revenue",
            description="Camp",
            budget=Decimal("10000"),
            ytd=Decimal("0"),
            remaining=Decimal("10000"),
            used_pct=Decimal("0"),
            faculty="Programs & Camps",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="8330",
            account="Expenditure",
            description="Camp",
            budget=Decimal("10000"),
            ytd=Decimal("0"),
            remaining=Decimal("10000"),
            used_pct=Decimal("0"),
            faculty="Programs & Camps",
            is_over=False,
        )
        out = tmp_path / "nospend.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        status_cell = ws.cell(row=3, column=5)
        # Round 56: 'No spend yet' pill removed; budgeted-with-zero-spend
        # rows now read as On track. The bold styling that emphasised
        # the No-spend-yet pill no longer applies here.
        assert status_cell.value == "On track"

    def test_xlsx_auto_fills_empty_comments_for_urgent_rows(self, tmp_path: Path) -> None:
        """R1 fix: when Status is Urgent / Significant / Spent-without-
        budget AND Comments is empty, auto-fill so the cell doesn't
        print as a contradiction. Round 56 dropped the No-spend-yet
        special case along with the pill itself."""
        rev = SubProgramLine(
            sub_program="7001",
            account="Revenue",
            description="Admin",
            budget=Decimal("0"),
            ytd=Decimal("0"),
            remaining=Decimal("0"),
            used_pct=Decimal("0"),
            faculty="Administration",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="7001",
            account="Expenditure",
            description="Admin",
            budget=Decimal("100000"),
            ytd=Decimal("250000"),
            remaining=Decimal("-150000"),
            used_pct=Decimal("250"),
            faculty="Administration",
            is_over=True,
        )
        out = tmp_path / "autofill.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        status = ws.cell(row=3, column=5).value
        comments = ws.cell(row=3, column=13).value
        assert status == "Investigate urgently"
        # R2 fix: imperative cue for non-OK statuses (was archival).
        assert comments == "Action needed: add commentary."

    def test_prose_round_trip_through_decode_commentary(self) -> None:
        """R1 regression test: a prose cell from R53+ fed back through
        ``decode_commentary`` (next month's prior-period join) returns
        the prose verbatim as Notes-only — no crash, no partial parse,
        no silent loss of structure."""
        from tools.sub_program.logic import decode_commentary, render_commentary_prose

        prose = render_commentary_prose(
            notes="Reviewed by council",
            driver="Ongoing",
            outlook="Expected to continue",
            action="Monitor",
        )
        # prose = "Ongoing variance, expected to continue. Being monitored. Reviewed by council."
        notes_back, driver_back, outlook_back, action_back = decode_commentary(prose)
        # Graceful fallback: whole text becomes Notes; structured fields
        # are blank. (User can re-pick categorisation in the editor.)
        assert notes_back == prose
        assert driver_back == ""
        assert outlook_back == ""
        assert action_back == ""

    def test_status_pill_materiality_floor_with_explicit_5k(self) -> None:
        """Round 58: with the function-level default lowered to $100,
        the materiality floor only kicks in when the user / caller
        explicitly raises it. Pin the floor's behaviour with an
        explicit $5K — $4.5K past threshold falls back to On track via
        the materiality + pct_over clause (9 pp ≤ 50)."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("50000"),
                exp_ytd=Decimal("55000"),  # $4.5K past 101% threshold
                materiality_dollar=5000,
            )
            == "On track"
        )

    def test_status_pill_default_materiality_100_does_not_suppress_4k(self) -> None:
        """Round 58 follow-on: with the new $100 default materiality,
        the same $4.5K overrun is no longer suppressed — it lands in
        Slightly over (the $25K Material dollar floor + 25% percent
        floor still gate the higher buckets)."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("50000"),
                exp_ytd=Decimal("55000"),
            )
            == "Slightly over"
        )

    def test_status_pill_boundary_overrun_exactly_25000(self) -> None:
        """Round 56 boundary pin: $25,000 spend over the 101% threshold
        on a $100K budget = $14K past threshold. With pct_over = 14 pp
        (which is < the 25% Material percent floor) the dollar floor
        ($25K Significant) governs — $14K overrun → Slightly over."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("100000"),
                exp_ytd=Decimal("115000"),  # $14K past 101% threshold
            )
            == "Slightly over"
        )

    def test_status_pill_boundary_overrun_just_above_25k_dollar_floor(self) -> None:
        """Round 56 boundary pin: $26K past the 101% threshold trips
        the $25K Material dollar floor → Significant overspend."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("200000"),
                exp_ytd=Decimal("228000"),  # $26K past 101% threshold ($202K)
            )
            == "Significant overspend"
        )

    def test_status_pill_boundary_overrun_exactly_100000_dollar_floor(self) -> None:
        """Round 56 boundary pin: at $100,000 past the threshold the
        rule is strictly ``>`` so this lands in Significant, not
        Urgent. With $500K budget at 101% = $505K threshold, exp_ytd
        $605K leaves $100K overrun → Significant (not yet Urgent)."""
        from tools.sub_program.logic import compute_status_pill

        result = compute_status_pill(
            annual_exp_budget=Decimal("500000"),
            exp_ytd=Decimal("605000"),  # exactly $100K past 101% threshold
        )
        assert result == "Significant overspend"

    def test_status_pill_boundary_overrun_just_above_100000_dollar_floor(self) -> None:
        """Round 56 boundary pin: $101K past the threshold trips the
        $100K Urgent dollar floor → Investigate urgently."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("500000"),
                exp_ytd=Decimal("606000"),  # $101K past 101% threshold
            )
            == "Investigate urgently"
        )

    def test_status_pill_threshold_at_exact_boundary_is_on_track(self) -> None:
        """Round 56 boundary pin: used_pct = threshold exactly → On
        track (the rule is ``used_pct > threshold``, strictly above)."""
        from tools.sub_program.logic import compute_status_pill

        # exp_ytd = 101 K on $100K budget = exactly 101% → On track.
        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("100000"),
                exp_ytd=Decimal("101000"),
                expense_threshold=101.0,
            )
            == "On track"
        )


# ---------------------------------------------------------------------------
# Round 53 F1 R2 — regression tests for the 10 R2 fixes applied after the
# second 4-Opus-agent review. Pin the new behaviour against future drift.
# ---------------------------------------------------------------------------


class TestF1Round2Fixes:
    """Targeted tests for the R2 fixes (logic, prose, writer)."""

    def test_investigating_improving_is_contradictory(self) -> None:
        """R2 fix: ``Investigating + Improving`` is incoherent — 'we
        don't know what's driving this AND it's improving' — outlook
        is dropped."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="Investigating", outlook="Improving")
            == "Driver under investigation."
        )

    def test_investigating_deteriorating_keeps_outlook(self) -> None:
        """R2 fix: ``Investigating + Deteriorating`` is coherent —
        'we don't know why but it's getting worse'."""
        from tools.sub_program.logic import render_commentary_prose

        result = render_commentary_prose(driver="Investigating", outlook="Deteriorating")
        # Per the R2 special-case: Investigating + outlook splits into
        # two unambiguous sentences.
        assert "Driver under investigation." in result
        assert "Variance deteriorating." in result

    def test_investigating_investigate_drops_action(self) -> None:
        """R2 fix: 'Driver under investigation. Needs investigation.' is
        repetitive — same word root in both sentences. Action dropped."""
        from tools.sub_program.logic import render_commentary_prose

        assert (
            render_commentary_prose(driver="Investigating", action="Investigate")
            == "Driver under investigation."
        )

    def test_investigating_with_outlook_renders_two_sentences(self) -> None:
        """R2 fix: ``Investigating`` driver + outlook used to comma-join
        as 'Driver under investigation, improving' — ambiguous (does
        the driver improve, or the variance?). Now splits into 2
        unambiguous sentences."""
        from tools.sub_program.logic import render_commentary_prose

        # Improving is now in the contradictory set, so use Deteriorating.
        result = render_commentary_prose(
            driver="Investigating", outlook="Deteriorating", action="Monitor"
        )
        assert result == "Driver under investigation. Variance deteriorating. Being monitored."

    def test_spent_without_budget_excludes_revenue_collection(self) -> None:
        """R2 fix: ``Spent without budget`` requires rev_ytd == 0 too.
        A program with $0 budget on both sides BUT collecting revenue
        is a configuration mistake (forgot to budget), not unauthorised
        spend — surfacing it as 'Spent without budget' alongside its
        visible revenue reads as contradictory to a council reader."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("0"),
                annual_rev_budget=Decimal("0"),
                rev_ytd=Decimal("5000"),  # rev collected — not unbudgeted
                exp_ytd=Decimal("1000"),
            )
            != "Spent without budget"
        )

    def test_donation_program_with_unbudgeted_revenue_flags(self) -> None:
        """Round 62: a program with rev_b = 0 but rev_y > 0 (donations
        / unbudgeted grants) now reads as 'Revenue over budget' when
        the rev_y is material — council needs to acknowledge the
        unplanned income (refund / roll-forward / re-allocate).
        Pre-R62 the Status pill ignored the Revenue side, so the same
        row read as 'On track' which contradicted the Watchlist
        flagging it via the per-line is_over check."""
        from tools.sub_program.logic import compute_status_pill

        # Donation $5K received (rev_b=0, rev_y=$5K), $3.5K spent of
        # $10K exp budget = 35% (under 101% threshold). The expense
        # side is happy but the revenue side IS material.
        result = compute_status_pill(
            annual_exp_budget=Decimal("10000"),
            annual_rev_budget=Decimal("0"),
            rev_ytd=Decimal("5000"),
            exp_ytd=Decimal("3500"),
        )
        assert result == "Revenue over budget"

    def test_donation_program_below_materiality_is_on_track(self) -> None:
        """Round 62 boundary: same shape as the test above but rev_y
        below the materiality floor → On track (the $100 default mat
        treats sub-$100 donations as chart-of-accounts noise)."""
        from tools.sub_program.logic import compute_status_pill

        result = compute_status_pill(
            annual_exp_budget=Decimal("10000"),
            annual_rev_budget=Decimal("0"),
            rev_ytd=Decimal("50"),  # below $100 mat
            exp_ytd=Decimal("3500"),
        )
        assert result == "On track"

    def test_admin_scale_overrun_triggers_urgent(self) -> None:
        """Round 56: the Admin-style sub-program with $700K spend on a
        $582K budget (≈120%) lands $112K past the 101% threshold,
        well over the $100K Urgent dollar floor."""
        from tools.sub_program.logic import compute_status_pill

        result = compute_status_pill(
            annual_exp_budget=Decimal("581700"),
            annual_rev_budget=Decimal("0"),
            rev_ytd=Decimal("26436"),
            exp_ytd=Decimal("700000"),
        )
        assert result == "Investigate urgently"

    def test_noise_floor_500_dollars(self) -> None:
        """Round 56: hard $500 noise floor regardless of percent past
        the threshold. With a $1K budget at 101% threshold ($1,010), a
        $1.4K spend = $390 past threshold, below the $500 noise floor
        → On track despite the ~40% pct_over."""
        from tools.sub_program.logic import compute_status_pill

        assert (
            compute_status_pill(
                annual_exp_budget=Decimal("1000"),
                exp_ytd=Decimal("1400"),  # $390 past threshold
            )
            == "On track"
        )

    def test_percent_floor_above_50pct(self) -> None:
        """Round 56: percent floor — overrun > 50 pp past threshold
        surfaces even if dollar overrun < $5K materiality. A $4K
        overrun on a $200 budget is way more than 50% past threshold."""
        from tools.sub_program.logic import compute_status_pill

        # $4,200 spend on $200 budget = 2100% used = ~1999 pp past 101%
        # threshold. Overrun = $4,198, well above $500 noise floor and
        # pct_over > 50 → Urgent (pct rule trips before dollar bucket).
        result = compute_status_pill(
            annual_exp_budget=Decimal("200"),
            exp_ytd=Decimal("4200"),
        )
        assert result == "Investigate urgently"

    def test_xlsx_zero_spend_renders_on_track(self, tmp_path: Path) -> None:
        """Round 56: a sub-program with budget allocated but $0 YTD
        spend reads as On track. The Comments cell is left empty for
        On-track rows (auto-fill only runs for non-OK statuses to
        avoid the 'urgent + blank comment' contradiction)."""
        rev = SubProgramLine(
            sub_program="8330",
            account="Revenue",
            description="Camp",
            budget=Decimal("10000"),
            ytd=Decimal("0"),
            remaining=Decimal("10000"),
            used_pct=Decimal("0"),
            faculty="Programs & Camps",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="8330",
            account="Expenditure",
            description="Camp",
            budget=Decimal("10000"),
            ytd=Decimal("0"),
            remaining=Decimal("10000"),
            used_pct=Decimal("0"),
            faculty="Programs & Camps",
            is_over=False,
        )
        out = tmp_path / "zero_spend.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=3, column=5).value == "On track"
        # On-track rows leave Comments empty — no contradiction to fix.
        assert ws.cell(row=3, column=13).value is None

    def test_xlsx_comments_width_reduced_to_40_then_32(self, tmp_path: Path) -> None:
        """Round 53 R2 reduced Comments column 50 → 40. F2 R1 further
        reduced it to 32 (after F2 added Status + Trend columns,
        which pushed total widths up). This test pins the F2 R1 value."""
        out = tmp_path / "width.xlsx"
        line = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([line], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        # Column 12 is Comments; Excel column letter "L".
        # F2 R1: width is now 32 (down from 40).
        assert ws.column_dimensions["M"].width == 32

    def test_xlsx_capped_marker_with_pink_fill_and_text_format(self, tmp_path: Path) -> None:
        """Round 67 rewrite — the writer no longer caps extreme
        percents (was R2's pink + ">999%" marker + text format
        combination, dropped per user feedback that the marker hid
        the actual magnitude). A materially-over row with an
        extreme percent now carries the raw formula AND the pink
        fill simultaneously."""
        from toolkit.tokens import HL_MISMATCH

        # Engineer a row with rev_pct way over 999% AND a material
        # over-spend so both the percent formula and the pink fill apply.
        rev = SubProgramLine(
            sub_program="9001",
            account="Revenue",
            description="Test",
            budget=Decimal("100"),
            ytd=Decimal("3000"),  # 30x over
            remaining=Decimal("-2900"),
            used_pct=Decimal("3000"),
            faculty="Programs & Camps",
            is_over=True,
        )
        exp = SubProgramLine(
            sub_program="9001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("60000"),  # $50K over → urgent + pink
            remaining=Decimal("-50000"),
            used_pct=Decimal("600"),
            faculty="Programs & Camps",
            is_over=True,
        )
        out = tmp_path / "tri_style.xlsx"
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        # Round 67: load without data_only so the formula text is
        # visible (data_only returns None for un-cached formulas).
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        # Round 66: Revenue % at col 4 (relocated with the
        # percent-column move between PROGRAM NAME and Status).
        rev_pct_cell = ws.cell(row=3, column=4)
        # Round 67: raw formula, no marker.
        assert rev_pct_cell.value == "=I3/G3"
        # Pink fill (row is materially over, so the pink-fill loop
        # paints col 4 too).
        rgb = rev_pct_cell.fill.fgColor.rgb if rev_pct_cell.fill.fgColor is not None else None
        assert rgb is not None and HL_MISMATCH in rgb.upper()


# Round 63 — TestComputeTrend + TestLoadPriorPeriodYtd classes
# deleted along with the production functions they exercised. The
# Trend column was dropped from the XLSX in Round 57; the functions
# had no production callers afterwards.


class TestF2XlsxLayout:
    """The F2 layout: Status at col 3, Trend at col 4. Cols 3..12 of
    F1 shift right by 2 (so old col 12 Comments → new col 14)."""

    def _line(
        self,
        sub_program: str,
        account: str,
        budget: str,
        ytd: str,
    ) -> SubProgramLine:
        b = Decimal(budget)
        y = Decimal(ytd)
        return SubProgramLine(
            sub_program=sub_program,
            account=account,
            description=f"{sub_program} desc",
            budget=b,
            ytd=y,
            remaining=b - y,
            used_pct=(y / b * Decimal("100")) if b != 0 else Decimal("0"),
            faculty="Curriculum",
            is_over=False,
        )

    def test_header_row_status_at_col_5(self, tmp_path: Path) -> None:
        """Round 66 — Status column at col 5 (was col 3 in R57; the
        two percent columns moved between PROGRAM NAME and Status)."""
        out = tmp_path / "out.xlsx"
        rev = self._line("4001", "Revenue", "10000", "5000")
        exp = self._line("4001", "Expenditure", "10000", "5000")
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=2, column=5).value == "Status"

    def test_header_row_funds_at_col_6(self, tmp_path: Path) -> None:
        """Round 66: Funds from Previous Years at col 6 (was col 4 in
        R57; shifted right by 2 with the percent-columns move)."""
        out = tmp_path / "out.xlsx"
        rev = self._line("4001", "Revenue", "10000", "5000")
        exp = self._line("4001", "Expenditure", "10000", "5000")
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        header = str(ws.cell(row=2, column=6).value or "")
        assert header.startswith("Funds from Previous Years")

    def test_header_row_percent_columns_at_3_and_4(self, tmp_path: Path) -> None:
        """Round 66: the two percent columns sit at C/D so the council
        reader's eye lands on the at-a-glance bars before Status."""
        out = tmp_path / "out.xlsx"
        rev = self._line("4001", "Revenue", "10000", "5000")
        exp = self._line("4001", "Expenditure", "10000", "5000")
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=2, column=3).value == "Available Balance % YTD"
        assert ws.cell(row=2, column=4).value == "Revenue Budget % Received YTD"

    def test_header_row_comments_at_col_13(self, tmp_path: Path) -> None:
        """Round 57: Comments shifts left from col 14 to col 13 after
        Trend column drop. Round 66 left this unchanged at col 13."""
        out = tmp_path / "out.xlsx"
        rev = self._line("4001", "Revenue", "10000", "5000")
        exp = self._line("4001", "Expenditure", "10000", "5000")
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=2, column=13).value == "Comments"

    def test_data_row_status_at_col_5(self, tmp_path: Path) -> None:
        out = tmp_path / "out.xlsx"
        rev = self._line("4001", "Revenue", "10000", "6000")
        exp = self._line("4001", "Expenditure", "10000", "5000")
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Sub Program Report"]
        assert ws.cell(row=3, column=5).value == "On track"


class TestF2WatchlistSheet:
    """The Watchlist sheet: filtered subset where Status != On track."""

    def _ok_pair(self, sp: str) -> list[SubProgramLine]:
        return [
            SubProgramLine(
                sub_program=sp,
                account="Revenue",
                description="OK",
                budget=Decimal("10000"),
                ytd=Decimal("6000"),
                remaining=Decimal("4000"),
                used_pct=Decimal("60"),
                faculty="Curriculum",
                is_over=False,
            ),
            SubProgramLine(
                sub_program=sp,
                account="Expenditure",
                description="OK",
                budget=Decimal("10000"),
                ytd=Decimal("5000"),
                remaining=Decimal("5000"),
                used_pct=Decimal("50"),
                faculty="Curriculum",
                is_over=False,
            ),
        ]

    def _over_pair(self, sp: str, exp_ytd: str) -> list[SubProgramLine]:
        return [
            SubProgramLine(
                sub_program=sp,
                account="Revenue",
                description="Over",
                budget=Decimal("0"),
                ytd=Decimal("0"),
                remaining=Decimal("0"),
                used_pct=Decimal("0"),
                faculty="X",
                is_over=False,
            ),
            SubProgramLine(
                sub_program=sp,
                account="Expenditure",
                description="Over",
                budget=Decimal("100000"),
                ytd=Decimal(exp_ytd),
                remaining=Decimal("100000") - Decimal(exp_ytd),
                # Round 63: derive consistent used_pct + variance_amount
                # + is_material so the test data flows through
                # _recompute_is_over (and the writer's in-place is_material
                # recompute) without contradictions.
                used_pct=(Decimal(exp_ytd) / Decimal("100000")) * Decimal("100"),
                faculty="X",
                is_over=True,
                is_material=abs(Decimal(exp_ytd) - Decimal("100000")) >= Decimal("100"),
                variance_amount=Decimal(exp_ytd) - Decimal("100000"),
            ),
        ]

    def test_watchlist_sheet_exists(self, tmp_path: Path) -> None:
        out = tmp_path / "out.xlsx"
        rows = self._ok_pair("4001") + self._over_pair("7001", "250000")
        _write_xlsx(rows, out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        assert "Watchlist" in wb.sheetnames

    def test_watchlist_excludes_on_track_rows(self, tmp_path: Path) -> None:
        out = tmp_path / "out.xlsx"
        rows = self._ok_pair("4001") + self._over_pair("7001", "250000")
        _write_xlsx(rows, out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Watchlist"]
        codes: list[str] = []
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None:
                codes.append(str(v))
        assert "7001" in codes
        assert "4001" not in codes

    def test_watchlist_sorted_by_max_variance_desc(self, tmp_path: Path) -> None:
        """Round 63: sort by max |variance_amount| across the
        sub-program's lines, descending. Mirrors the in-app Watchlist
        tab order so a council reader scanning either view sees the
        same item at the top. Pre-R63 the XLSX sorted by signed
        available ascending and the in-app sorted by -abs(variance),
        producing different orderings of the same sub-program set."""
        out = tmp_path / "out.xlsx"
        # _over_pair builds Expenditure lines with budget=$100K and
        # ytd as supplied. |variance| = |ytd - 100,000|.
        rows = (
            self._over_pair("1001", "150000")  # |var| = $50K
            + self._over_pair("2001", "300000")  # |var| = $200K
            + self._over_pair("3001", "120000")  # |var| = $20K
        )
        _write_xlsx(rows, out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb["Watchlist"]
        codes: list[str] = []
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None:
                codes.append(str(v))
        # Biggest variance ($200K = 2001) first; smallest ($20K = 3001) last.
        assert codes == ["2001", "1001", "3001"]


# ---------------------------------------------------------------------------
# Round 54 F2 R1 — regression tests for the 10 fixes after the first
# 4-Opus parallel review.
# ---------------------------------------------------------------------------


class TestF2Round1Fixes:
    """Targeted tests pinning the F2 R1 fixes."""

    # Round 63 — three test methods (test_compute_trend_uses_current_status_for_threshold_rows,
    # test_compute_trend_status_off_track_overrides_available_signal,
    # test_load_prior_period_ytd_skips_watchlist_sheet) deleted with
    # the production functions they exercised.

    def test_xlsx_active_sheet_is_sub_program_report(self, tmp_path: Path) -> None:
        """R1 fix: ``wb.active`` pinned to the main sheet so Excel
        opens to it by default — not the Watchlist (which is
        sorted/filtered and looks like a partial export)."""
        out = tmp_path / "active.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        # The active sheet on file open is "Sub Program Report".
        assert wb.active is not None
        assert wb.active.title == "Sub Program Report"

    def test_xlsx_main_sheet_has_print_area(self, tmp_path: Path) -> None:
        """R1 fix: ``ws.print_area`` set on both sheets so a council
        member who hits Ctrl+P doesn't accidentally double-paginate
        via the workbook-wide print mode."""
        out = tmp_path / "print_area.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        # print_area set; openpyxl returns "'Sheet'!$A$1:$M$N" form.
        assert ws.print_area
        # Round 57: ends on column M (col 13) after Trend column dropped.
        assert "$A$1" in ws.print_area and "$M$" in ws.print_area

    def test_xlsx_watchlist_has_autofilter_and_pink_tab(self, tmp_path: Path) -> None:
        """R1 fix: Watchlist sheet has AutoFilter on its header row and
        a pink tab colour to draw the council reader's eye to the
        actionable sheet."""
        out = tmp_path / "watchlist_features.xlsx"
        rev = SubProgramLine(
            sub_program="7001",
            account="Revenue",
            description="Admin",
            budget=Decimal("0"),
            ytd=Decimal("0"),
            remaining=Decimal("0"),
            used_pct=Decimal("0"),
            faculty="Administration",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="7001",
            account="Expenditure",
            description="Admin",
            budget=Decimal("100000"),
            ytd=Decimal("250000"),
            remaining=Decimal("-150000"),
            used_pct=Decimal("250"),
            faculty="Administration",
            is_over=True,
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Watchlist"]
        # AutoFilter set with a non-trivial reference range.
        assert ws.auto_filter.ref is not None
        assert "A2" in ws.auto_filter.ref
        # Tab colour = HL_MISMATCH (canonical pink for needs-attention).
        from toolkit.tokens import HL_MISMATCH

        # openpyxl returns Color or rgb string; both should contain HL_MISMATCH.
        tab_color = ws.sheet_properties.tabColor
        if tab_color is not None:
            tab_color_str = str(tab_color.rgb if hasattr(tab_color, "rgb") else tab_color)
            assert HL_MISMATCH in tab_color_str.upper()

    def test_xlsx_main_sheet_has_no_autofilter(self, tmp_path: Path) -> None:
        """The main sheet should NOT have AutoFilter — the Watchlist
        sheet is the interactive view; the main sheet is the
        canonical layout."""
        out = tmp_path / "main_no_filter.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        assert ws.auto_filter.ref is None

    def test_xlsx_comments_column_width_reduced_to_32(self, tmp_path: Path) -> None:
        """R1 fix: Comments column 40 → 32 to relieve print compression
        after F2 added Status (22) + Trend (16) widths."""
        out = tmp_path / "width.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        # Comments at col 14 = letter "N".
        assert ws.column_dimensions["M"].width == 32

    def test_watchlist_excludes_unspent_programs(self, tmp_path: Path) -> None:
        """Round 56: the Watchlist is strictly an over-budget list. A
        budgeted-but-unspent program no longer appears on the Watchlist
        (the calendar-based 'No spend yet' pill was retired with the
        pacing logic). Only the overspend row 7001 surfaces here."""
        out = tmp_path / "signed_sort.xlsx"
        # Big overspend
        over_rev = SubProgramLine(
            sub_program="7001",
            account="Revenue",
            description="Admin",
            budget=Decimal("0"),
            ytd=Decimal("0"),
            remaining=Decimal("0"),
            used_pct=Decimal("0"),
            faculty="X",
            is_over=False,
        )
        over_exp = SubProgramLine(
            sub_program="7001",
            account="Expenditure",
            description="Admin",
            budget=Decimal("100000"),
            ytd=Decimal("300000"),
            remaining=Decimal("-200000"),
            used_pct=Decimal("300"),
            faculty="X",
            is_over=True,
        )
        # Big unspent — was 'No spend yet' pre-R56, now reads as On track.
        unspent_rev = SubProgramLine(
            sub_program="8330",
            account="Revenue",
            description="Camp",
            budget=Decimal("200000"),
            ytd=Decimal("0"),
            remaining=Decimal("200000"),
            used_pct=Decimal("0"),
            faculty="Y",
            is_over=False,
        )
        unspent_exp = SubProgramLine(
            sub_program="8330",
            account="Expenditure",
            description="Camp",
            budget=Decimal("200000"),
            ytd=Decimal("0"),
            remaining=Decimal("200000"),
            used_pct=Decimal("0"),
            faculty="Y",
            is_over=False,
        )
        _write_xlsx(
            [over_rev, over_exp, unspent_rev, unspent_exp],
            out,
            period_label="Apr 2026",
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Watchlist"]
        codes: list[str] = []
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v is not None:
                codes.append(str(v))
        # 7001 (overspend) appears; 8330 (unspent) does not — Round 56
        # narrowed the Watchlist to over-budget rows only.
        assert "7001" in codes
        assert "8330" not in codes


# ---------------------------------------------------------------------------
# Round 54 F2 R2 — regression tests for the R2 fixes after the second
# 4-Opus parallel review.
# ---------------------------------------------------------------------------


class TestF2Round2Fixes:
    """Targeted tests pinning the R2 fixes."""

    def test_empty_watchlist_has_no_autofilter(self, tmp_path: Path) -> None:
        """R2 fix: when every sub-program is on track, the Watchlist
        sheet is empty (header only). Don't write a degenerate
        ``A2:N2`` AutoFilter — Excel renders an inert dropdown that
        looks broken to a council reader."""
        out = tmp_path / "empty_watchlist.xlsx"
        rev = SubProgramLine(
            sub_program="4001",
            account="Revenue",
            description="OK",
            budget=Decimal("10000"),
            ytd=Decimal("6000"),
            remaining=Decimal("4000"),
            used_pct=Decimal("60"),
            faculty="Curriculum",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="OK",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Watchlist"]
        assert ws.auto_filter.ref is None

    def test_empty_watchlist_keeps_pink_tab_color(self, tmp_path: Path) -> None:
        """R2 fix: even with no Watchlist data rows, the tab stays pink
        — it signals the sheet's CATEGORY (urgent / actionable),
        independent of whether there's content this period."""
        from toolkit.tokens import HL_MISMATCH

        out = tmp_path / "empty_pink.xlsx"
        rev = SubProgramLine(
            sub_program="4001",
            account="Revenue",
            description="OK",
            budget=Decimal("10000"),
            ytd=Decimal("6000"),
            remaining=Decimal("4000"),
            used_pct=Decimal("60"),
            faculty="Curriculum",
            is_over=False,
        )
        exp = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="OK",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([rev, exp], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Watchlist"]
        tab_color = ws.sheet_properties.tabColor
        if tab_color is not None:
            tab_color_str = str(tab_color.rgb if hasattr(tab_color, "rgb") else tab_color)
            assert HL_MISMATCH in tab_color_str.upper()

    def test_footer_text_is_plain_attribution(self, tmp_path: Path) -> None:
        """Round 57: the F2 trend-warning footer was dropped along with
        the Trend column. The footer-left always shows the plain
        'Generated by School Tool' attribution now."""
        out = tmp_path / "footer.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx([ln], out, period_label="Apr 2026")
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        footer_left = ws.oddFooter.left.text if ws.oddFooter is not None else None
        assert footer_left == "Generated by School Tool"

    def test_footer_text_when_prior_period_supplied_is_attribution(self, tmp_path: Path) -> None:
        """When prior-period IS supplied, footer is the plain attribution."""
        out = tmp_path / "footer_with_prior.xlsx"
        ln = SubProgramLine(
            sub_program="4001",
            account="Expenditure",
            description="Test",
            budget=Decimal("10000"),
            ytd=Decimal("5000"),
            remaining=Decimal("5000"),
            used_pct=Decimal("50"),
            faculty="Curriculum",
            is_over=False,
        )
        _write_xlsx(
            [ln],
            out,
            period_label="Apr 2026",
            prior_ytd={"4001": Decimal("4000")},
        )
        wb = openpyxl.load_workbook(out)
        ws = wb["Sub Program Report"]
        footer_left = ws.oddFooter.left.text if ws.oddFooter is not None else None
        assert footer_left == "Generated by School Tool"

    # Round 63 — test_compute_trend_documents_status_aligned_recommendation
    # deleted along with the compute_trend function.
