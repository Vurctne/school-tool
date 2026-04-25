"""
Renderer-specific colour helpers for the Vic School Finance Toolkit.

Purpose
-------
This module provides pure helper functions that convert an ``HL_*`` hex-RGB
string (as defined in ``toolkit.tokens``) into the encoding required by a
specific renderer.  The two renderers currently supported are:

* **openpyxl** – ``PatternFill(fgColor=...)`` expects an 8-character ARGB
  string with a leading ``"FF"`` alpha byte, e.g. ``"FFF4CCCC"``.
* **Win32 COM** – ``.Interior.Color`` expects a BGR integer, e.g.
  ``13421812`` (``0xCCCCF4``).

Design-system mapping
---------------------
The three data-highlight tokens from ``toolkit.tokens``, their CSS custom
properties, and their roles:

==============================  ====================  =======================
Constant                        CSS custom property   Role / appearance
==============================  ====================  =======================
``HL_EDITED      = "FFF2CC"``   ``--hl-edited``       Yellow – user-edited cell
``HL_MISMATCH    = "F4CCCC"``   ``--hl-mismatch``     Pink/red – value mismatch
``HL_SOURCE_ONLY = "E2F0D9"``   ``--hl-source-only``  Green – source-only row
==============================  ====================  =======================

Note on ``HL_EDITED``: this token is defined in the design-system token set
and appears in the user-facing instruction text for the Master Budget tool
(instruction-sheet legend).  However, **no fill site in the current codebase
actually applies it to cells** at this time.  It is retained here for legend
completeness and possible future use.

Usage example
-------------
::

    from toolkit.tokens import HL_MISMATCH, HL_SOURCE_ONLY
    from toolkit.fills import argb, bgr_int

    # openpyxl
    from openpyxl.styles import PatternFill
    fill = PatternFill("solid", fgColor=argb(HL_MISMATCH))

    # Win32 COM
    ws.Range("A1").Interior.Color = bgr_int(HL_SOURCE_ONLY)
"""

from __future__ import annotations

_HEX_CHARS = frozenset("0123456789ABCDEFabcdef")


def _validate(hex_rgb: str) -> str:
    """Strip an optional leading ``#``, upper-case, and validate length/chars.

    Returns the clean 6-character uppercase hex string, or raises
    ``ValueError``.
    """
    s = hex_rgb.lstrip("#").upper()
    if len(s) != 6 or not all(c in _HEX_CHARS for c in s):
        raise ValueError(
            f"hex_rgb must be exactly 6 hex digits (with or without a leading '#'); got {hex_rgb!r}"
        )
    return s


def argb(hex_rgb: str) -> str:
    """Return an 8-character uppercase ARGB string with full-opacity alpha.

    Prepends ``"FF"`` (fully opaque) to the cleaned RGB value so the result
    is suitable for ``openpyxl.styles.PatternFill(fgColor=...)``.

    Parameters
    ----------
    hex_rgb:
        6-character RGB hex string, case-insensitive, with or without a
        leading ``#``.  E.g. ``"F4CCCC"`` or ``"#f4cccc"``.

    Returns
    -------
    str
        8-character uppercase ARGB string, e.g. ``"FFF4CCCC"``.

    Raises
    ------
    ValueError
        If *hex_rgb* is not exactly 6 hex digits after stripping ``#``.

    Examples
    --------
    >>> argb("F4CCCC")
    'FFF4CCCC'
    >>> argb("#e2f0d9")
    'FFE2F0D9'
    """
    return "FF" + _validate(hex_rgb)


def bgr_int(hex_rgb: str) -> int:
    """Return the Win32 COM BGR integer for ``.Interior.Color``.

    Parses the RR, GG, BB components and returns ``BB << 16 | GG << 8 | RR``
    so that Excel's ``Interior.Color`` property receives the correct value.

    Parameters
    ----------
    hex_rgb:
        6-character RGB hex string, case-insensitive, with or without a
        leading ``#``.  E.g. ``"F4CCCC"`` or ``"#f4cccc"``.

    Returns
    -------
    int
        BGR-encoded integer.  E.g. ``bgr_int("F4CCCC")`` → ``13421812``
        (``0xCCCCF4``).

    Raises
    ------
    ValueError
        If *hex_rgb* is not exactly 6 hex digits after stripping ``#``.

    Examples
    --------
    >>> bgr_int("F4CCCC")
    13421812
    >>> bgr_int("E2F0D9")
    14282978
    """
    s = _validate(hex_rgb)
    rr = int(s[0:2], 16)
    gg = int(s[2:4], 16)
    bb = int(s[4:6], 16)
    return (bb << 16) | (gg << 8) | rr
